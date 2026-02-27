from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import csv
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, validator
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
import jwt
import hashlib
import json
import shutil
import re
import bcrypt as _bcrypt
from collections import defaultdict
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

BOGOTA_TZ = ZoneInfo("America/Bogota")

ROOT_DIR = Path(__file__).parent
# Load .env.local first (for local development with real credentials)
# Then load .env (for default/example values)
load_dotenv(ROOT_DIR / '.env.local')
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
# Use environment variable or default to localhost for local development
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
# Log connection without exposing sensitive information (credentials, ports, etc.)
def redact_mongo_url(url: str) -> str:
    """Redact sensitive information from MongoDB URL for logging"""
    if 'localhost' in url or '127.0.0.1' in url:
        return "localhost"
    # For remote URLs, just indicate it's remote without showing host/credentials
    return "cloud/remote"

logger.info(f"Connecting to MongoDB at: {redact_mongo_url(mongo_url)}")
try:
    client = AsyncIOMotorClient(
        mongo_url,
        maxPoolSize=100,
        minPoolSize=10,
        maxIdleTimeMS=30000,
        connectTimeoutMS=5000,
        serverSelectionTimeoutMS=5000,
    )
    db = client[os.environ.get('DB_NAME', 'WebApp')]
    logger.info(f"MongoDB client initialized for database: {os.environ.get('DB_NAME', 'WebApp')}")
except Exception as e:
    logger.error(f"Failed to initialize MongoDB client: {e}")
    raise

# JWT Secret
JWT_SECRET = os.environ.get('JWT_SECRET')
_JWT_DEFAULT = 'educando_secret_key_2025_CHANGE_ME'
if not JWT_SECRET:
    logger.warning("⚠️ JWT_SECRET not set! Using insecure default. SET THIS IN PRODUCTION!")
    JWT_SECRET = _JWT_DEFAULT
# Block server start in production when using the insecure default JWT_SECRET
if any(os.environ.get(env) for env in ['RENDER', 'RAILWAY_ENVIRONMENT', 'DYNO']) and JWT_SECRET == _JWT_DEFAULT:
    raise RuntimeError(
        "🚫 FATAL: JWT_SECRET is set to the insecure default value. "
        "Cannot start server in production. Set the JWT_SECRET environment variable to a strong random secret."
    )
JWT_ALGORITHM = "HS256"

# Password hashing with bcrypt (using bcrypt directly to avoid passlib compatibility issues)

# Password storage mode: 'plain' for plain text, 'bcrypt' for hashed (default: 'bcrypt' for security)
# For backwards compatibility with existing plain text passwords, set PASSWORD_STORAGE_MODE='plain'
PASSWORD_STORAGE_MODE = os.environ.get('PASSWORD_STORAGE_MODE', 'bcrypt').lower()

# Rate limiting: track login attempts per IP
# WARNING: This is in-memory storage and will be reset on server restart.
# For production with multiple instances or persistence across restarts,
# consider using Redis or another distributed cache for rate limiting.
login_attempts = defaultdict(list)
login_attempts_by_identifier = defaultdict(list)
MAX_LOGIN_ATTEMPTS_PER_IP = 50        # límite alto para WiFi compartido
MAX_LOGIN_ATTEMPTS_PER_USER = 5       # límite estricto por usuario individual
LOGIN_ATTEMPT_WINDOW = 300  # 5 minutes in seconds

# Rate limiting for file uploads: track per user
upload_attempts = defaultdict(list)
MAX_UPLOADS_PER_MINUTE = 20
UPLOAD_WINDOW = 60  # 1 minute in seconds

# Module validation constants
MIN_MODULE_NUMBER = 1
# MAX_MODULE_NUMBER is not enforced – the system supports N modules per program definition.

def validate_module_number(module_num, field_name="module"):
    """Validate that a module number is a positive integer (no upper bound – supports N modules)."""
    if not isinstance(module_num, int) or module_num < MIN_MODULE_NUMBER:
        raise ValueError(f"{field_name} must be >= {MIN_MODULE_NUMBER}, got {module_num}")
    return True

app = FastAPI()

# Detect production environment
IS_PRODUCTION = any(os.environ.get(env) for env in ['RENDER', 'RAILWAY_ENVIRONMENT', 'DYNO'])

# Configure CORS origins
cors_origins_str = os.environ.get('CORS_ORIGINS', '*')
cors_origins = cors_origins_str.split(',') if ',' in cors_origins_str else [cors_origins_str]
allow_credentials = "*" not in cors_origins

if IS_PRODUCTION and '*' in cors_origins:
    import logging as _logging
    _logging.getLogger(__name__).warning(
        "⚠️  SECURITY WARNING: CORS_ORIGINS is set to '*' in production! "
        "Restrict CORS_ORIGINS to your frontend domain to improve security."
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=allow_credentials,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Requested-With"],
)

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        if request.url.scheme == "https":
            response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        if request.url.path.startswith("/api/auth/"):
            response.headers["Cache-Control"] = "no-store"
        return response

app.add_middleware(SecurityHeadersMiddleware)

api_router = APIRouter(prefix="/api")

# Initialize scheduler for automatic module closure
scheduler = AsyncIOScheduler()

async def check_and_close_modules():
    """
    Check all programs and courses for module close dates that have passed and automatically close them.
    This function runs daily at 00:01 AM.
    Also checks recovery close dates: students who haven't passed recovery by the deadline are removed.
    """
    try:
        logger.info("Running automatic module closure check...")
        now = datetime.now(timezone.utc)
        # Use Bogotá timezone for date comparisons so scheduler results are
        # consistent with the dates configured by administrators (UTC-5).
        today_str = datetime.now(BOGOTA_TZ).strftime("%Y-%m-%d")
        
        # Get all programs; check all moduleN_close_date fields in Python for N-module support.
        # (Business rule 2026-02-25: no longer hardcoded to 2 modules)
        programs = await db.programs.find({}, {"_id": 0}).to_list(100)
        
        if not programs:
            logger.info("No programs with close dates to process")
        else:
            for program in programs:
                program_id = program["id"]
                program_name = program["name"]
                
                # Determine how many modules this program defines (minimum 2 for backwards compatibility)
                program_modules_list = program.get("modules") or []
                max_modules = max(len(program_modules_list), 2)
                
                # Check all modules for this program dynamically (supports N modules)
                for mod_num in range(1, max_modules + 1):
                    close_date_field = f"module{mod_num}_close_date"
                    close_date = program.get(close_date_field)
                    if not close_date or close_date > today_str:
                        continue

                    closure_check = await db.module_closures.find_one({
                        "program_id": program_id,
                        "module_number": mod_num,
                        "closed_date": close_date
                    })

                    if not closure_check:
                        logger.info(f"Auto-closing Module {mod_num} for program {program_name} (date: {close_date})")
                        try:
                            result = await close_module_internal(module_number=mod_num, program_id=program_id)

                            await db.module_closures.insert_one({
                                "id": str(uuid.uuid4()),
                                "program_id": program_id,
                                "module_number": mod_num,
                                "closed_date": close_date,
                                "closed_at": now.isoformat(),
                                "result": result
                            })
                            logger.info(
                                f"Module {mod_num} closed for {program_name}: "
                                f"{result['promoted_count']} promoted, {result['graduated_count']} graduated, "
                                f"{result['recovery_pending_count']} in recovery"
                            )
                        except Exception as e:
                            logger.error(f"Error auto-closing Module {mod_num} for {program_name}: {e}", exc_info=True)
        
        # Check course-level recovery close dates: promote winners, remove losers
        all_courses = await db.courses.find({}, {"_id": 0}).to_list(1000)
        all_subjects = await db.subjects.find({}, {"_id": 0, "id": 1, "module_number": 1}).to_list(5000)
        subject_module_map = {s["id"]: (s.get("module_number") or 1) for s in all_subjects}
        # Pre-load programs indexed by id for fallback module_dates construction
        all_programs_map = {p["id"]: p for p in programs} if programs else {}
        removed_count = 0
        promoted_recovery_count = 0
        for course in all_courses:
            module_dates = course.get("module_dates") or {}
            prog_id_for_course = course.get("program_id", "")

            # --- Causa 1: fallback module_dates from program when course has none ---
            if not module_dates and prog_id_for_course:
                prog = all_programs_map.get(prog_id_for_course)
                if prog:
                    prog_modules_list = prog.get("modules") or []
                    max_mod = max(len(prog_modules_list), 2)
                    for mn in range(1, max_mod + 1):
                        close_field = f"module{mn}_close_date"
                        close_val = prog.get(close_field)
                        if close_val:
                            # Use moduleN_close_date as a proxy recovery_close when no
                            # per-course module_dates are configured.
                            module_dates[str(mn)] = {"recovery_close": close_val}
                if module_dates:
                    logger.warning(
                        f"Course {course['id']} has no module_dates; synthesized fallback "
                        f"from program {prog_id_for_course}: {list(module_dates.keys())}"
                    )
                else:
                    logger.warning(
                        f"Course {course['id']} has no module_dates and program "
                        f"{prog_id_for_course!r} has no moduleN_close_date fields – "
                        "skipping recovery check for this course."
                    )
                    continue

            for module_key, dates in module_dates.items():
                recovery_close = dates.get("recovery_close") if dates else None
                if not recovery_close or recovery_close > today_str:
                    continue  # Recovery period not closed yet
                
                # --- Causa 3: retroactively run close_module_internal if it was never executed ---
                # Check whether failed_subjects records already exist for this course/module.
                existing_any = await db.failed_subjects.find_one(
                    {"course_id": course["id"], "module_number": int(module_key)},
                    {"_id": 0, "id": 1}
                )
                if not existing_any:
                    # Check if close_module_internal was already recorded for this program/module.
                    closure_exists = await db.module_closures.find_one({
                        "program_id": prog_id_for_course,
                        "module_number": int(module_key),
                    }) if prog_id_for_course else None
                    if not closure_exists and prog_id_for_course:
                        logger.warning(
                            f"Recovery close passed for course {course['id']} module {module_key} "
                            "but no failed_subjects records found and no module closure recorded – "
                            "running close_module_internal retroactively."
                        )
                        try:
                            retro_result = await close_module_internal(
                                module_number=int(module_key),
                                program_id=prog_id_for_course
                            )
                            await db.module_closures.insert_one({
                                "id": str(uuid.uuid4()),
                                "program_id": prog_id_for_course,
                                "module_number": int(module_key),
                                "closed_date": recovery_close,
                                "closed_at": now.isoformat(),
                                "result": retro_result,
                                "retroactive": True
                            })
                            logger.info(
                                f"Retroactive close_module_internal for program {prog_id_for_course} "
                                f"module {module_key}: {retro_result}"
                            )
                            # Re-fetch records after the retroactive closure
                            existing_any = await db.failed_subjects.find_one(
                                {"course_id": course["id"], "module_number": int(module_key)},
                                {"_id": 0, "id": 1}
                            )
                        except Exception as retro_err:
                            logger.error(
                                f"Error in retroactive close_module_internal for course "
                                f"{course['id']} module {module_key}: {retro_err}",
                                exc_info=True
                            )

                # Find ALL failed_subjects for this course/module that haven't been processed yet.
                all_records = await db.failed_subjects.find({
                    "course_id": course["id"],
                    "module_number": int(module_key),
                    "recovery_expired": {"$ne": True},
                    "recovery_processed": {"$ne": True}
                }, {"_id": 0}).to_list(None)
                
                prog_id = course.get("program_id", "")

                # Group recovery records by student
                students_records = {}
                for record in all_records:
                    sid = record["student_id"]
                    students_records.setdefault(sid, []).append(record)

                # Skip only when there are no recovery records AND no enrolled students.
                # We still need to process courses with no recovery records if they have
                # direct-pass students whose promotion was deferred to recovery_close.
                if not all_records and not course.get("student_ids"):
                    continue
                
                for student_id, records in students_records.items():
                    # Business rule (2026-02-25): if recovery_close has passed and no admin action
                    # has been taken, treat the student as failed: remove from group and mark activo
                    # for re-enrollment so they are never left in limbo indefinitely.
                    has_admin_action = any(
                        r.get("recovery_approved") is True or r.get("recovery_completed") is True
                        for r in records
                    )
                    if not has_admin_action:
                        logger.info(
                            f"Recovery close: no admin action for student {student_id} in course "
                            f"{course['id']} – applying fail flow (removing from group, marking reprobado)"
                        )
                        # Remove student only from the course where recovery was not passed.
                        await _unenroll_student_from_course(student_id, course["id"])
                        student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
                        _ps = (student_doc or {}).get("program_statuses") or {}
                        if prog_id:
                            _ps[prog_id] = "reprobado"
                        _new_estado = derive_estado_from_program_statuses(_ps)
                        _upd: dict = {"estado": _new_estado}
                        if prog_id:
                            _upd["program_statuses"] = _ps
                        await db.users.update_one({"id": student_id, "role": "estudiante"}, {"$set": _upd})
                        logger.info(
                            f"Student {student_id} marked reprobado for program {prog_id} "
                            "(recovery_close passed with no admin action)"
                        )
                        await log_audit(
                            "student_removed_from_group", "system", "system",
                            {"student_id": student_id, "course_id": course["id"],
                             "program_id": prog_id, "trigger": "recovery_close_no_admin_action"}
                        )
                        removed_count += 1
                        for record in records:
                            await db.failed_subjects.update_one(
                                {"id": record["id"]},
                                {"$set": {"recovery_processed": True, "processed_at": now.isoformat()}}
                            )
                        continue

                    # A student passes only if ALL their records are approved by admin AND completed AND approved by teacher
                    all_passed = all(
                        r.get("recovery_approved") is True and
                        r.get("recovery_completed") is True and
                        r.get("teacher_graded_status") == "approved"
                        for r in records
                    )
                    
                    if all_passed:
                        # Promote to next module or graduate
                        student = await db.users.find_one({"id": student_id}, {"_id": 0})
                        if student:
                            program_modules = student.get("program_modules") or {}
                            program_statuses = student.get("program_statuses") or {}
                            program = await db.programs.find_one({"id": prog_id}, {"_id": 0}) if prog_id else None
                            max_modules = max(len(program.get("modules", [])) if program and program.get("modules") else 2, 2)
                            # Use the module being closed (not the student's potentially stale field)
                            # to decide between promotion and graduation
                            if int(module_key) >= max_modules:
                                program_statuses[prog_id] = "egresado"
                                new_global_estado = derive_estado_from_program_statuses(program_statuses)
                                await db.users.update_one(
                                    {"id": student_id},
                                    {
                                        "$set": {"program_statuses": program_statuses, "estado": new_global_estado},
                                        "$unset": {f"program_promotion_pending.{prog_id}": ""}
                                    }
                                )
                                logger.info(f"Student {student_id} graduated (recovery passed all subjects in course {course['id']})")
                                await log_audit("student_graduated", "system", "system", {"student_id": student_id, "course_id": course["id"], "trigger": "recovery_close"})
                            else:
                                next_module = int(module_key) + 1
                                program_statuses[prog_id] = "activo"
                                new_global_estado = derive_estado_from_program_statuses(program_statuses)
                                update_fields = {
                                    "estado": new_global_estado,
                                    "program_statuses": program_statuses
                                }
                                if prog_id:
                                    update_fields[f"program_modules.{prog_id}"] = next_module
                                await db.users.update_one(
                                    {"id": student_id},
                                    {
                                        "$set": update_fields,
                                        "$unset": {f"program_promotion_pending.{prog_id}": ""}
                                    }
                                )
                                logger.info(f"Student {student_id} promoted to module {next_module} (recovery passed in course {course['id']})")
                                await log_audit("student_promoted", "system", "system", {"student_id": student_id, "course_id": course["id"], "next_module": next_module, "trigger": "recovery_close"})
                            promoted_recovery_count += 1
                    else:
                        # At least one subject not passed: remove from group, mark activo for re-enrollment
                        logger.info(f"Recovery close: student {student_id} did not pass all subjects in course {course['id']} – removing from group")
                        # Remove student only from the course where recovery was not passed.
                        await _unenroll_student_from_course(student_id, course["id"])
                        student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
                        student_program_statuses = (student_doc or {}).get("program_statuses") or {}
                        if prog_id:
                            student_program_statuses[prog_id] = "reprobado"
                        new_estado = derive_estado_from_program_statuses(student_program_statuses)
                        update_fields = {"estado": new_estado}
                        if prog_id:
                            update_fields["program_statuses"] = student_program_statuses
                        await db.users.update_one({"id": student_id, "role": "estudiante"}, {"$set": update_fields})
                        logger.info(f"Student {student_id} marked as reprobado for program {prog_id} (recovery closed, did not pass)")
                        await log_audit("student_removed_from_group", "system", "system", {"student_id": student_id, "course_id": course["id"], "program_id": prog_id, "trigger": "recovery_close_failed"})
                        removed_count += 1
                    
                    # Mark all records for this student in this course/module as processed
                    for record in records:
                        await db.failed_subjects.update_one(
                            {"id": record["id"]},
                            {"$set": {"recovery_processed": True, "processed_at": now.isoformat()}}
                        )
                
                # Precompute course/module grades for direct-pass validation and strict fallback.
                subject_ids = course.get("subject_ids") or []
                if not subject_ids and course.get("subject_id"):
                    subject_ids = [course["subject_id"]]
                module_subject_ids = [sid for sid in subject_ids if subject_module_map.get(sid, 1) == int(module_key)]
                course_grades = await db.grades.find(
                    {"course_id": course["id"]},
                    {"_id": 0, "student_id": 1, "subject_id": 1, "value": 1}
                ).to_list(None)
                grades_index = {}
                for g in course_grades:
                    if g.get("value") is None:
                        continue
                    grades_index.setdefault((g.get("student_id"), g.get("subject_id")), []).append(g["value"])

                # Also promote students who passed the module directly (no failed subjects),
                # whose promotion was deferred to recovery_close by close_module_internal.
                for student_id in course.get("student_ids", []):
                    if student_id in students_records:
                        continue  # Already handled above as a recovery student
                    direct_student = await db.users.find_one({"id": student_id}, {"_id": 0})
                    if not direct_student:
                        continue
                    direct_prog_modules = direct_student.get("program_modules") or {}
                    # Only process if still in the module being closed (idempotency guard)
                    if direct_prog_modules.get(prog_id) != int(module_key):
                        continue
                    # Check if student has unresolved failed_subjects in OTHER courses
                    # for the same program and module. If so, skip promotion — the student
                    # must resolve all recovery subjects before advancing.
                    other_pending = await db.failed_subjects.find_one({
                        "student_id": student_id,
                        "program_id": prog_id,
                        "module_number": int(module_key),
                        "recovery_processed": {"$ne": True},
                        "recovery_expired": {"$ne": True},
                    })
                    if other_pending:
                        continue

                    # Defensive rule: never promote at recovery_close when the student still
                    # has failing averages in this course/module, even if failed_subjects is
                    # missing or incomplete.
                    if module_subject_ids:
                        has_failing = False
                        for sid in module_subject_ids:
                            values = grades_index.get((student_id, sid), [])
                            avg = (sum(values) / len(values)) if values else 0.0
                            if avg < 3.0:
                                has_failing = True
                                break
                    else:
                        values = [g.get("value") for g in course_grades if g.get("student_id") == student_id and g.get("value") is not None]
                        avg = (sum(values) / len(values)) if values else 0.0
                        has_failing = avg < 3.0

                    if has_failing:
                        logger.info(
                            f"Recovery close: student {student_id} has failing averages in "
                            f"course {course['id']} module {module_key} without full recovery pass "
                            "– removing from group instead of promoting"
                        )
                        await _unenroll_student_from_course(student_id, course["id"])
                        removed_count += 1
                        student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
                        ps = (student_doc or {}).get("program_statuses") or {}
                        if prog_id:
                            ps[prog_id] = "reprobado"
                        new_estado = derive_estado_from_program_statuses(ps)
                        upd = {"estado": new_estado}
                        if prog_id:
                            upd["program_statuses"] = ps
                        await db.users.update_one({"id": student_id, "role": "estudiante"}, {"$set": upd})
                        await log_audit(
                            "student_removed_from_group", "system", "system",
                            {"student_id": student_id, "course_id": course["id"],
                             "program_id": prog_id, "trigger": "recovery_close_direct_pass_failing_guard"}
                        )
                        continue

                    direct_prog_statuses = direct_student.get("program_statuses") or {}
                    # A student can be reprobado here when they were removed from a different
                    # course in the same program/module (either earlier in this scheduler run
                    # or by _check_and_update_recovery_rejection) but their student_id was not
                    # yet pulled from THIS course's student_ids.  Remove them instead of
                    # promoting them.
                    if direct_prog_statuses.get(prog_id) == "reprobado":
                        # Student is already reprobado (failed in another course or by a prior
                        # scheduler run) but their id was not yet removed from this course's
                        # student_ids.  Remove them now so they are not left as ghosts.
                        logger.info(
                            f"Recovery close: student {student_id} is already reprobado for "
                            f"program {prog_id} – removing from course {course['id']} (CASO 5)"
                        )
                        await _unenroll_student_from_course(student_id, course["id"])
                        removed_count += 1
                        await log_audit(
                            "student_removed_from_group", "system", "system",
                            {"student_id": student_id, "course_id": course["id"],
                             "program_id": prog_id, "trigger": "recovery_close_already_reprobado"}
                        )
                        continue
                    program = await db.programs.find_one({"id": prog_id}, {"_id": 0}) if prog_id else None
                    max_modules = max(len(program.get("modules", [])) if program and program.get("modules") else 2, 2)
                    if int(module_key) >= max_modules:
                        direct_prog_statuses[prog_id] = "egresado"
                        new_global_estado = derive_estado_from_program_statuses(direct_prog_statuses)
                        await db.users.update_one(
                            {"id": student_id},
                            {
                                "$set": {"program_statuses": direct_prog_statuses, "estado": new_global_estado},
                                "$unset": {f"program_promotion_pending.{prog_id}": ""}
                            }
                        )
                        logger.info(f"Student {student_id} graduated (direct pass, deferred to recovery_close for course {course['id']})")
                        await log_audit("student_graduated", "system", "system", {"student_id": student_id, "course_id": course["id"], "trigger": "recovery_close_direct_pass"})
                    else:
                        next_module = int(module_key) + 1
                        direct_prog_statuses[prog_id] = "activo"
                        new_global_estado = derive_estado_from_program_statuses(direct_prog_statuses)
                        update_fields = {
                            "estado": new_global_estado,
                            "program_statuses": direct_prog_statuses
                        }
                        if prog_id:
                            update_fields[f"program_modules.{prog_id}"] = next_module
                        await db.users.update_one(
                            {"id": student_id},
                            {
                                "$set": update_fields,
                                "$unset": {f"program_promotion_pending.{prog_id}": ""}
                            }
                        )
                        logger.info(f"Student {student_id} promoted to module {next_module} (direct pass, deferred to recovery_close for course {course['id']})")
                        await log_audit("student_promoted", "system", "system", {"student_id": student_id, "course_id": course["id"], "next_module": next_module, "trigger": "recovery_close_direct_pass"})
                    promoted_recovery_count += 1

                # Strict fallback at recovery_close: if a student still has failing averages
                # in this course/module and did not fully pass recovery records, remove them.
                # This covers edge cases where failed_subjects records are missing/stale.
                for student_id in list(course.get("student_ids", [])):
                    # Skip if already unenrolled in previous steps
                    current_course_doc = await db.courses.find_one({"id": course["id"]}, {"_id": 0, "student_ids": 1})
                    if not current_course_doc or student_id not in (current_course_doc.get("student_ids") or []):
                        continue

                    # Only enforce for students still in the module being closed
                    stud = await db.users.find_one({"id": student_id}, {"_id": 0, "program_modules": 1})
                    if not stud:
                        continue
                    if prog_id and (stud.get("program_modules") or {}).get(prog_id) != int(module_key):
                        continue

                    # Determine if student has failing performance in this course/module
                    if module_subject_ids:
                        has_failing = False
                        for sid in module_subject_ids:
                            values = grades_index.get((student_id, sid), [])
                            avg = (sum(values) / len(values)) if values else 0.0
                            if avg < 3.0:
                                has_failing = True
                                break
                    else:
                        values = [g.get("value") for g in course_grades if g.get("student_id") == student_id and g.get("value") is not None]
                        avg = (sum(values) / len(values)) if values else 0.0
                        has_failing = avg < 3.0

                    if not has_failing:
                        continue

                    unresolved_records = await db.failed_subjects.find({
                        "student_id": student_id,
                        "course_id": course["id"],
                        "module_number": int(module_key),
                        "recovery_processed": {"$ne": True},
                        "recovery_expired": {"$ne": True},
                    }, {"_id": 0}).to_list(None)

                    # If all unresolved records are fully approved+graded, let normal pass flow apply
                    if unresolved_records and all(
                        r.get("recovery_approved") is True and
                        r.get("recovery_completed") is True and
                        r.get("teacher_graded_status") == "approved"
                        for r in unresolved_records
                    ):
                        continue

                    await _unenroll_student_from_course(student_id, course["id"])
                    removed_count += 1
                    student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
                    ps = (student_doc or {}).get("program_statuses") or {}
                    if prog_id:
                        ps[prog_id] = "reprobado"
                    new_estado = derive_estado_from_program_statuses(ps)
                    upd = {"estado": new_estado}
                    if prog_id:
                        upd["program_statuses"] = ps
                    await db.users.update_one({"id": student_id, "role": "estudiante"}, {"$set": upd})

                    if unresolved_records:
                        for rec in unresolved_records:
                            await db.failed_subjects.update_one(
                                {"id": rec["id"]},
                                {"$set": {"recovery_processed": True, "processed_at": now.isoformat()}}
                            )

                    await log_audit(
                        "student_removed_from_group", "system", "system",
                        {"student_id": student_id, "course_id": course["id"],
                         "program_id": prog_id, "trigger": "recovery_close_strict_fallback_failed"}
                    )

        if removed_count > 0:
            logger.info(f"Recovery check: removed {removed_count} students from groups due to failed recovery")
        if promoted_recovery_count > 0:
            logger.info(f"Recovery check: promoted/graduated {promoted_recovery_count} students after passing recovery")
        
        logger.info("Automatic module closure check completed")
    except Exception as e:
        logger.error(f"Error in automatic module closure check: {e}", exc_info=True)

# Health check endpoint for monitoring
@app.get("/api/health")
async def health_check():
    """
    Health check endpoint for load balancers and monitoring systems.
    Returns basic status without exposing sensitive implementation details.
    """
    try:
        # Ping MongoDB to verify connection
        await db.command('ping')
        return {"status": "healthy"}
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={"status": "unhealthy"}
        )


# Global exception handler
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception: {exc}", exc_info=True)
    # Only expose detailed error in development (when DEBUG env var is set)
    debug_mode = os.environ.get('DEBUG', 'false').lower() == 'true'
    if debug_mode:
        logger.warning("DEBUG mode is enabled - detailed errors are exposed to clients")
    return JSONResponse(
        status_code=500,
        content={
            "detail": "Internal server error",
            "error": str(exc) if debug_mode else "An unexpected error occurred"
        }
    )

# File upload directory
# WARNING: Files are stored on local disk, which is ephemeral in containerized environments
# like Render, Railway, or Docker. Uploaded files will be LOST on redeployment or restart.
# For production, migrate to persistent storage like Cloudinary, AWS S3, or similar services.
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Cloudinary configuration (optional – enables persistent cloud storage for uploads)
CLOUDINARY_CLOUD_NAME = os.environ.get('CLOUDINARY_CLOUD_NAME')
CLOUDINARY_API_KEY = os.environ.get('CLOUDINARY_API_KEY')
CLOUDINARY_API_SECRET = os.environ.get('CLOUDINARY_API_SECRET')
USE_CLOUDINARY = bool(CLOUDINARY_CLOUD_NAME and CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET)

if USE_CLOUDINARY:
    try:
        import cloudinary
        import cloudinary.uploader
        cloudinary.config(
            cloud_name=CLOUDINARY_CLOUD_NAME,
            api_key=CLOUDINARY_API_KEY,
            api_secret=CLOUDINARY_API_SECRET,
            secure=True
        )
        logger.info("Cloudinary configured – uploads will use persistent cloud storage.")
    except ImportError:
        USE_CLOUDINARY = False
        logger.warning("cloudinary package not installed. Falling back to local disk storage.")
else:
    # Check if we're in a production environment and warn about ephemeral storage
    if os.environ.get('RENDER') or os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('DYNO'):
        logger.warning(
            "⚠️  PRODUCTION WARNING: Files are stored on ephemeral disk storage. "
            "Uploaded files will be LOST on redeployment. "
            "Set CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY, and CLOUDINARY_API_SECRET to use persistent storage."
        )

# AWS S3 configuration for PDF storage (persistent across deploys)
AWS_ACCESS_KEY_ID = os.environ.get('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.environ.get('AWS_SECRET_ACCESS_KEY')
AWS_S3_BUCKET_NAME = os.environ.get('AWS_S3_BUCKET_NAME')
AWS_S3_REGION = os.environ.get('AWS_S3_REGION', 'us-east-1')
USE_S3 = bool(AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY and AWS_S3_BUCKET_NAME)

if USE_S3:
    logger.info("AWS S3 configured – PDFs will use persistent S3 storage.")
else:
    logger.warning("AWS S3 not configured – PDFs stored on ephemeral local disk.")

# --- Startup Event - Crear datos iniciales ---
@app.on_event("startup")
async def startup_event():
    try:
        logger.info("Starting application initialization...")
        
        # Log production warnings
        if PASSWORD_STORAGE_MODE == 'plain':
            logger.warning(
                "⚠️  SECURITY WARNING: Password storage mode is set to 'plain'. "
                "Passwords are stored in plain text, which is INSECURE. "
                "Set PASSWORD_STORAGE_MODE='bcrypt' in your environment variables for production."
            )
        
        # Warn about in-memory rate limiting
        if os.environ.get('RENDER') or os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('DYNO'):
            logger.warning(
                "⚠️  PRODUCTION NOTICE: Rate limiting is in-memory and will reset on server restart. "
                "For distributed deployments, consider using Redis for persistent rate limiting."
            )
        
        # Test MongoDB connection
        await db.command('ping')
        logger.info("MongoDB connection successful")
        # Crear índices para rendimiento óptimo
        try:
            from create_indexes import create_indexes
            await create_indexes(db)
            logger.info("Índices MongoDB verificados/creados exitosamente")
        except Exception as e:
            logger.warning(f"No se pudieron crear índices automáticamente: {e}")
        await create_initial_data()
        
        # Start the automatic module closure scheduler
        # Runs daily at 02:00 AM (server local timezone) – low traffic window
        # Note: Uses server's local timezone by default. For production, consider explicitly setting timezone.
        scheduler.add_job(
            check_and_close_modules,
            CronTrigger(hour=2, minute=0),  # Run at 02:00 AM daily (server local time)
            id='auto_close_modules',
            name='Automatic Module Closure',
            replace_existing=True
        )
        scheduler.start()
        logger.info("Automatic module closure scheduler started (runs daily at 02:00 AM server local time)")
        
        # Verify S3 bucket accessibility at startup
        if USE_S3:
            try:
                import boto3
                s3_test = boto3.client(
                    's3',
                    aws_access_key_id=AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                    region_name=AWS_S3_REGION
                )
                s3_test.head_bucket(Bucket=AWS_S3_BUCKET_NAME)
                logger.info(f"✅ AWS S3 bucket '{AWS_S3_BUCKET_NAME}' is accessible.")
            except Exception as e:
                logger.error(f"❌ AWS S3 bucket verification failed: {e}. PDF uploads will fail!")
        
        logger.info("Application startup completed successfully")
    except Exception as e:
        logger.error(f"Startup failed: {e}", exc_info=True)
        if "auth" in str(e).lower() or "connection" in str(e).lower() or "ServerSelectionTimeoutError" in type(e).__name__:
            logger.error(
                "MongoDB connection failed. Please check your MONGO_URL environment variable. "
                "Common causes: invalid credentials, IP not whitelisted in MongoDB Atlas, "
                "or incorrect connection string format. "
                "See backend/.env.example for configuration examples."
            )
        logger.warning(
            "Application started WITHOUT database connection. "
            "API endpoints requiring MongoDB will not work until the connection is restored."
        )

async def create_initial_data():
    """Crea los usuarios y datos iniciales si no existen"""
    logger.info("Verificando y creando datos iniciales...")
    
    # Crear programas con sus módulos y materias SOLO si no existen
    # Usar $setOnInsert para evitar sobrescribir cambios hechos desde el admin panel
    programs = [
        {
            "id": "prog-admin", 
            "name": "Técnico en Asistencia Administrativa", 
            "description": "Formación técnica en gestión administrativa", 
            "duration": "12 meses",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Fundamentos de Administración",
                    "Herramientas Ofimáticas",
                    "Gestión Documental y Archivo",
                    "Atención al Cliente y Comunicación Organizacional",
                    "Legislación Laboral y Ética Profesional"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Contabilidad Básica",
                    "Nómina y Seguridad Social Aplicada",
                    "Control de Inventarios y Logística",
                    "Inglés Técnico / Competencias Ciudadanas",
                    "Proyecto Integrador Virtual"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True
        },
        {
            "id": "prog-infancia", 
            "name": "Técnico Laboral en Atención a la Primera Infancia", 
            "description": "Formación en desarrollo infantil", 
            "duration": "12 meses",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Inglés I",
                    "Proyecto de vida",
                    "Construcción social de la infancia",
                    "Perspectiva del desarrollo infantil",
                    "Salud y nutrición",
                    "Lenguaje y educación infantil",
                    "Juego y otras formas de comunicación",
                    "Educación y pedagogía"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Inglés II",
                    "Construcción del mundo Matemático",
                    "Dificultades en el aprendizaje",
                    "Estrategias del aula",
                    "Trabajo de grado",
                    "Investigación",
                    "Práctica - Informe"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True
        },
        {
            "id": "prog-sst", 
            "name": "Técnico en Seguridad y Salud en el Trabajo", 
            "description": "Formación en prevención de riesgos", 
            "duration": "12 meses",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Fundamentos en Seguridad y Salud en el Trabajo",
                    "Administración en salud",
                    "Condiciones de seguridad",
                    "Matemáticas",
                    "Psicología del Trabajo"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Comunicación oral y escrita",
                    "Sistema de gestión de seguridad y salud del trabajo",
                    "Anatomía y fisiología",
                    "Medicina preventiva del trabajo",
                    "Ética profesional",
                    "Gestión ambiental",
                    "Proyecto de grado"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True
        },
    ]
    # Only insert programs if they don't already exist - never overwrite admin changes
    programs_created = 0
    for p in programs:
        result = await db.programs.update_one(
            {"id": p["id"]},
            {"$setOnInsert": p},
            upsert=True
        )
        if result.upserted_id:
            programs_created += 1
    if programs_created > 0:
        logger.info(f"Creados {programs_created} programas nuevos")
    else:
        logger.info("Todos los programas ya existen - no se sobrescribieron")
    
    # Crear materias basadas en los módulos de los programas
    # IMPORTANTE: Usar $setOnInsert para TODOS los campos para evitar sobrescribir cambios
    # Esto previene que los cursos pierdan la referencia a las materias
    for prog in programs:
        for module in prog["modules"]:
            for subj_name in module["subjects"]:
                subject_data = {
                    "id": str(uuid.uuid4()),
                    "name": subj_name,
                    "program_id": prog["id"],
                    "module_number": module["number"],
                    "description": "",
                    "active": True
                }
                await db.subjects.update_one(
                    {"name": subj_name, "program_id": prog["id"], "module_number": module["number"]},
                    {"$setOnInsert": subject_data},
                    upsert=True
                )
    
    # Verificar y crear usuarios iniciales (seed users)
    # IMPORTANTE: Solo creamos usuarios semilla si NO EXISTEN. No los sobrescribimos.
    # Esto permite que los cambios hechos desde el panel de admin sean permanentes.
    # Para forzar reset de usuarios, establecer RESET_USERS=true en variables de entorno.
    # Para deshabilitar la creación de usuarios semilla (recomendado en producción),
    # establecer CREATE_SEED_USERS=false en variables de entorno.
    
    # Migrate legacy non-UUID user IDs to proper UUID5-based IDs.
    # Old seed users had IDs like "user-editor-1" which are not valid UUIDs.
    # These caused 401 errors on every protected API call after login, logging the user out.
    legacy_id_map = {
        "user-editor-1": str(uuid.uuid5(uuid.NAMESPACE_OID, "user-editor-1")),
        "user-prof-1":   str(uuid.uuid5(uuid.NAMESPACE_OID, "user-prof-1")),
        "user-prof-2":   str(uuid.uuid5(uuid.NAMESPACE_OID, "user-prof-2")),
    }
    for old_id, new_id in legacy_id_map.items():
        result = await db.users.update_one({"id": old_id}, {"$set": {"id": new_id}})
        if result.modified_count > 0:
            logger.info(f"Migrated legacy user ID: {old_id} -> {new_id}")

    reset_users = os.environ.get('RESET_USERS', 'false').lower() == 'true'
    if reset_users:
        logger.warning("⚠️  RESET_USERS=true: Eliminando TODOS los usuarios existentes...")
        deleted_result = await db.users.delete_many({})
        logger.info(f"Eliminados {deleted_result.deleted_count} usuarios")
    
    existing_user_count = await db.users.count_documents({})
    if existing_user_count > 0:
        logger.info(f"Base de datos tiene {existing_user_count} usuarios. Verificando usuarios semilla...")
    else:
        logger.info("Base de datos vacía. Creando usuarios iniciales...")
    
    # CREATE_SEED_USERS: controla si se crean usuarios semilla en startup.
    # Por defecto 'true' para desarrollo local; establecer 'false' en producción (Render/Railway)
    # para evitar que los usuarios semilla se recreen automáticamente.
    create_seed_users = os.environ.get('CREATE_SEED_USERS', 'false').lower() == 'true'
    
    if not create_seed_users:
        logger.info("CREATE_SEED_USERS=false: Omitiendo creación de usuarios semilla (modo producción)")
    else:
        # Definir usuarios semilla (seed users) - solo se crean si no existen
        # Note: Email domains vary by role (@tecnico.com, @estudiante.com, @profesor.com) 
        # as specified in the requirements to clearly distinguish user types
        seed_users = [
            # 1 Editor
            {"id": str(uuid.uuid5(uuid.NAMESPACE_OID, "user-editor-1")), "name": "Editor Principal", "email": "editor@tecnico.com", "cedula": None, "password_hash": hash_password("Editor2024!"), "role": "editor", "program_id": None, "program_ids": [], "subject_ids": [], "phone": None, "active": True, "module": None, "grupo": None, "estado": "activo"},
            
            # 2 Profesores
            {"id": str(uuid.uuid5(uuid.NAMESPACE_OID, "user-prof-1")), "name": "Ana Martínez", "email": "ana.martinez@profesor.com", "cedula": None, "password_hash": hash_password("Profesor1!"), "role": "profesor", "program_id": None, "program_ids": [], "subject_ids": [], "phone": None, "active": True, "module": None, "grupo": None, "estado": "activo"},
            {"id": str(uuid.uuid5(uuid.NAMESPACE_OID, "user-prof-2")), "name": "Juan Rodríguez", "email": "juan.rodriguez@profesor.com", "cedula": None, "password_hash": hash_password("Profesor2!"), "role": "profesor", "program_id": None, "program_ids": [], "subject_ids": [], "phone": None, "active": True, "module": None, "grupo": None, "estado": "activo"},
        ]
        
        # Insertar usuarios semilla solo si no existen (setOnInsert)
        # Esto preserva los cambios hechos desde el admin panel
        created_count = 0
        for u in seed_users:
            result = await db.users.update_one(
                {"id": u["id"]},
                {"$setOnInsert": u},
                upsert=True
            )
            if result.upserted_id:
                created_count += 1
        
        if created_count > 0:
            logger.info(f"Creados {created_count} usuarios semilla nuevos")
        else:
            logger.info("Todos los usuarios semilla ya existen - no se sobrescribieron")
    
    logger.info(f"Total usuarios en sistema: {await db.users.count_documents({})}")
    
    # Migrate existing courses to ensure subject_ids field exists and is properly set
    # This fixes the data persistence issue where subject_ids might be missing or None
    logger.info("Checking and fixing course subject_ids field...")
    courses_to_fix = await db.courses.find(
        {"$or": [
            {"subject_ids": {"$exists": False}},
            {"subject_ids": None},
            {"subject_ids": []}
        ]},
        {"_id": 0}
    ).to_list(1000)
    
    fixed_count = 0
    for course in courses_to_fix:
        if course.get("subject_id") and (not course.get("subject_ids") or len(course.get("subject_ids", [])) == 0):
            # Course has subject_id but no subject_ids - migrate it
            await db.courses.update_one(
                {"id": course["id"]},
                {"$set": {"subject_ids": [course["subject_id"]]}}
            )
            fixed_count += 1
            logger.info(f"Fixed course {course['id']}: added subject_ids=[{course['subject_id']}]")
    
    if fixed_count > 0:
        logger.info(f"Fixed {fixed_count} courses with missing subject_ids field")
    
    # Ensure all users have subject_ids field (for teachers)
    logger.info("Checking and fixing user subject_ids field...")
    users_to_fix = await db.users.find(
        {"$or": [
            {"subject_ids": {"$exists": False}},
            {"subject_ids": None}
        ]},
        {"_id": 0}
    ).to_list(1000)
    
    fixed_user_count = 0
    for user in users_to_fix:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"subject_ids": []}}
        )
        fixed_user_count += 1
    
    if fixed_user_count > 0:
        logger.info(f"Fixed {fixed_user_count} users with missing subject_ids field")
    
    # Soft migration: initialize program_statuses for existing students who don't have it yet
    logger.info("Checking and migrating student program_statuses field...")
    students_to_migrate = await db.users.find(
        {"role": "estudiante", "program_statuses": {"$exists": False}},
        {"_id": 0, "id": 1, "program_ids": 1, "program_id": 1, "estado": 1, "failed_subjects": 1}
    ).to_list(5000)
    
    migrated_count = 0
    for student in students_to_migrate:
        program_ids_list = student.get("program_ids") or []
        if not program_ids_list and student.get("program_id"):
            program_ids_list = [student["program_id"]]
        if not program_ids_list:
            continue
        global_estado = student.get("estado", "activo") or "activo"
        # Map global status to per-program status; all programs get the same status
        program_statuses = {pid: global_estado for pid in program_ids_list}
        await db.users.update_one(
            {"id": student["id"]},
            {"$set": {"program_statuses": program_statuses}}
        )
        migrated_count += 1
    
    if migrated_count > 0:
        logger.info(f"Migrated {migrated_count} students: initialized program_statuses field")
    
    # Purge orphaned group-related data: delete records whose course_id no longer exists
    logger.info("Checking for orphaned group/course data...")
    existing_course_ids = [c["id"] for c in await db.courses.find({}, {"_id": 0, "id": 1}).to_list(5000)]
    if existing_course_ids:
        orphan_filter = {"course_id": {"$nin": existing_course_ids}}
    else:
        orphan_filter = {"course_id": {"$exists": True}}
    orphan_activities = await db.activities.delete_many(orphan_filter)
    orphan_grades = await db.grades.delete_many(orphan_filter)
    orphan_submissions = await db.submissions.delete_many(orphan_filter)
    orphan_videos = await db.class_videos.delete_many(orphan_filter)
    orphan_recovery = await db.recovery_enabled.delete_many(orphan_filter)
    orphan_failed = await db.failed_subjects.delete_many(orphan_filter)
    total_purged = (orphan_activities.deleted_count + orphan_grades.deleted_count +
                    orphan_submissions.deleted_count + orphan_videos.deleted_count +
                    orphan_recovery.deleted_count + orphan_failed.deleted_count)
    if total_purged > 0:
        logger.info(
            f"Purged {total_purged} orphaned records: "
            f"{orphan_activities.deleted_count} activities, "
            f"{orphan_grades.deleted_count} grades, "
            f"{orphan_submissions.deleted_count} submissions, "
            f"{orphan_videos.deleted_count} videos, "
            f"{orphan_recovery.deleted_count} recovery_enabled, "
            f"{orphan_failed.deleted_count} failed_subjects"
        )
    else:
        logger.info("No orphaned group/course data found")

    logger.info("Datos iniciales verificados/creados exitosamente")
    logger.info("5 usuarios semilla disponibles (ver USUARIOS_Y_CONTRASEÑAS.txt)")
    logger.info(f"Modo de almacenamiento de contraseñas: {PASSWORD_STORAGE_MODE}")

# --- Utility Functions ---
def get_current_module_from_dates(module_dates: dict) -> Optional[int]:
    """Determine the current module number based on today's date and module_dates.
    
    Returns the module whose date range (start to recovery_close or end) includes today.
    If today is before all modules, returns the first module.
    If today is after all modules, returns the last module.
    Returns None if module_dates is empty or has no valid date ranges.
    """
    if not module_dates:
        return None
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sorted_keys = sorted(module_dates.keys(), key=lambda k: int(k) if str(k).isdigit() else 0)
    # Check if today falls within any module's date range
    for mod_key in sorted_keys:
        dates = module_dates.get(mod_key) or {}
        start = dates.get("start")
        end = dates.get("recovery_close") or dates.get("end")
        if start and end and start <= today <= end:
            return int(mod_key)
    # No direct match - find nearest module
    modules_with_start = [(int(k), (module_dates.get(k) or {}).get("start")) for k in sorted_keys if (module_dates.get(k) or {}).get("start")]
    if not modules_with_start:
        return None
    modules_with_start.sort()
    # If today is before the first module start, use first module
    if today < modules_with_start[0][1]:
        return modules_with_start[0][0]
    # Otherwise use the last module that has started
    current = modules_with_start[0][0]
    for mod_num, start in modules_with_start:
        if start <= today:
            current = mod_num
    return current


def can_enroll_in_course(course: dict) -> bool:
    """Check if enrollment is still open for a course.

    Enrollment is allowed only if today is strictly before module 1's start date.
    If no module 1 start date is defined, enrollment is always allowed.
    """
    module_dates = course.get("module_dates") or {}
    mod1_dates = module_dates.get("1") or module_dates.get(1) or {}
    mod1_start = mod1_dates.get("start")
    if not mod1_start:
        return True
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return today < mod1_start


def can_enroll_in_module(module_dates: dict, module_number: int) -> bool:
    """Check if enrollment window is open for a specific module number.

    Rules:
    - Module 1: today < start_mod1
    - Module N>1: recovery_close_mod(N-1) <= today < start_mod(N)

    If dates are missing, enrollment is allowed (no restriction).
    """
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    mod_key = str(module_number)
    mod_dates = module_dates.get(mod_key) or module_dates.get(module_number) or {}
    mod_start = mod_dates.get("start")

    if module_number == 1:
        if not mod_start:
            return True
        return today < mod_start

    # Module N > 1: window is recovery_close_mod(N-1) <= today < start_mod(N)
    prev_key = str(module_number - 1)
    prev_dates = module_dates.get(prev_key) or module_dates.get(module_number - 1) or {}
    prev_recovery_close = prev_dates.get("recovery_close") or prev_dates.get("end")

    if not prev_recovery_close and not mod_start:
        return True  # No dates configured – allow
    if prev_recovery_close and today < prev_recovery_close:
        return False  # Previous module recovery hasn't closed yet
    if mod_start and today >= mod_start:
        return False  # Target module has already started
    return True


def get_open_enrollment_module(module_dates: dict) -> Optional[int]:
    """Return the module currently accepting enrollments based on date windows.

    Priority: first module whose enrollment window is currently open according to
    ``can_enroll_in_module``. Also checks the module immediately after the highest
    defined one so that inter-module enrollment windows are detected even when the
    next module's dates have not yet been added to the group (e.g. module 1 recovery
    has closed but module 2 dates are not yet configured – enrollment for module 2
    should still be allowed during that gap).
    Returns None when no module has an open window.
    """
    if not module_dates:
        return None
    mod_numbers = sorted(
        int(k) for k in module_dates.keys() if str(k).isdigit()
    )
    if not mod_numbers:
        return None
    # Also probe the next module beyond the highest defined so that the window
    # between module N's recovery_close and module N+1's start is recognised even
    # when module N+1 dates have not been added to the group yet.
    mods_to_check = mod_numbers + [mod_numbers[-1] + 1]
    for mod_num in mods_to_check:
        if can_enroll_in_module(module_dates, mod_num):
            return mod_num
    return None


def validate_module_dates_order(module_dates: dict) -> Optional[str]:
    """Validate that module N+1 starts after module N's recovery_close (or end) date.
    
    Returns an error message string if invalid, or None if valid.
    """
    if not module_dates or len(module_dates) < 2:
        return None
    sorted_keys = sorted(module_dates.keys(), key=lambda k: int(k) if str(k).isdigit() else 0)
    for i in range(len(sorted_keys) - 1):
        curr_key = sorted_keys[i]
        next_key = sorted_keys[i + 1]
        curr_dates = module_dates.get(curr_key) or {}
        next_dates = module_dates.get(next_key) or {}
        # Use recovery_close as the boundary if available, otherwise use end
        curr_boundary = curr_dates.get("recovery_close") or curr_dates.get("end")
        next_start = next_dates.get("start")
        if curr_boundary and next_start and next_start <= curr_boundary:
            return (f"La fecha de inicio del Módulo {next_key} ({next_start}) debe ser "
                    f"posterior al cierre de recuperaciones del Módulo {curr_key} ({curr_boundary})")
    return None


def validate_module_dates_recovery_close(module_dates: dict) -> Optional[str]:
    """Validate that every module_dates entry includes a recovery_close date.

    Business rule (2026-02-25): all module date windows must define a recovery_close
    so that the scheduler can deterministically resolve every student's status.
    Returns an error message string if any module is missing recovery_close, else None.
    """
    for mod_key, dates in (module_dates or {}).items():
        if not (dates or {}).get("recovery_close"):
            return (
                f"El Módulo {mod_key} no tiene fecha de cierre de recuperaciones "
                "(recovery_close). Todas las entradas de fechas de módulo deben incluirla."
            )
    return None


def derive_estado_from_program_statuses(program_statuses: dict) -> str:
    """Derive the global 'estado' from per-program statuses.

    Priority rules (multi-program independence):
    1. If ANY program is 'activo' → global 'activo' (student can still operate)
    2. If ALL programs are 'egresado' → global 'egresado'
    3. If ANY program is 'pendiente_recuperacion' (no activo) → global 'pendiente_recuperacion'
    4. If ANY program is 'egresado' (mix with retirado, no activo/pendiente) → 'egresado'
    5. All programs 'retirado' → global 'retirado'
    """
    if not program_statuses:
        return "activo"
    statuses = list(program_statuses.values())
    if "activo" in statuses:
        return "activo"
    if all(s == "egresado" for s in statuses):
        return "egresado"
    if "pendiente_recuperacion" in statuses:
        return "pendiente_recuperacion"
    if "egresado" in statuses:
        return "egresado"
    if "reprobado" in statuses:
        return "reprobado"
    return "retirado"


async def _get_program_courses_for_student(
    student_id: str, prog_id: str, fallback_course_id: str
) -> list:
    """Return all courses for prog_id where student_id is still enrolled.

    Falls back to a single-element list with fallback_course_id when:
    - prog_id is empty/falsy, or
    - the DB query returns no courses (e.g. student was already removed).

    This ensures callers always have at least one course to process.
    """
    if prog_id:
        courses = await db.courses.find(
            {"program_id": prog_id, "student_ids": student_id},
            {"_id": 0, "id": 1},
        ).to_list(None)
        return courses or [{"id": fallback_course_id}]
    return [{"id": fallback_course_id}]


async def _unenroll_student_from_course(
    student_id: str, course_id: str
) -> list:
    """Unenroll student only from the specified course and clear stale group label."""
    await db.courses.update_one(
        {"id": course_id},
        {
            "$pull": {"student_ids": student_id},
            "$addToSet": {"removed_student_ids": student_id}
        }
    )

    still_enrolled_anywhere = await db.courses.find_one(
        {"student_ids": student_id},
        {"_id": 0, "id": 1}
    )
    if not still_enrolled_anywhere:
        await db.users.update_one(
            {"id": student_id, "role": "estudiante"},
            {"$unset": {"grupo": ""}}
        )

    return [course_id]


async def _check_and_update_recovery_completion(student_id: str, course_id: str):
    """Check if all recovery subjects for a student in a course are now completed
    and approved by both admin and teacher. If so, immediately update the student's
    program status to 'activo' (or 'egresado' for the last module) without waiting
    for the recovery_close date.

    This is called after a teacher grades a recovery activity as approved so that the student
    exits 'pendiente_recuperacion' as soon as all their recoveries are resolved.

    Important: checks ALL unprocessed failed_subjects (not just admin-approved) so that
    a student is not prematurely promoted when some subjects haven't been admin-approved yet.
    """
    # Fetch the course to get program_id and module info
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        return
    prog_id = course.get("program_id", "")

    # Find ALL unprocessed recovery records for this student+course (including ones
    # where admin hasn't approved yet). This prevents premature promotion when some
    # subjects are still pending admin approval.
    all_records = await db.failed_subjects.find(
        {
            "student_id": student_id,
            "course_id": course_id,
            "recovery_processed": {"$ne": True},
            "recovery_expired": {"$ne": True},
        },
        {"_id": 0},
    ).to_list(None)

    if not all_records:
        return

    # Only proceed if every record is admin-approved, completed, and approved by teacher
    all_passed = all(
        r.get("recovery_approved") is True and
        r.get("recovery_completed") is True and r.get("teacher_graded_status") == "approved"
        for r in all_records
    )
    if not all_passed:
        return

    # Fetch current student document
    student = await db.users.find_one({"id": student_id}, {"_id": 0})
    if not student:
        return

    program_statuses = student.get("program_statuses") or {}

    # Only act when the student is still in pendiente_recuperacion for this program
    if program_statuses.get(prog_id) != "pendiente_recuperacion":
        return

    # Determine new status: egresado if this was the last module, otherwise activo.
    # Use the student's tracked module for this program as the authoritative source;
    # fall back to the module_number stored in the recovery records.
    module_number = (
        (student.get("program_modules") or {}).get(prog_id)
        or student.get("module")
        or all_records[0].get("module_number", 1)
    )
    program = await db.programs.find_one({"id": prog_id}, {"_id": 0}) if prog_id else None
    max_modules = max(len(program.get("modules", [])) if program else 2, 2)

    if module_number >= max_modules:
        program_statuses[prog_id] = "egresado"
    else:
        program_statuses[prog_id] = "activo"

    new_estado = derive_estado_from_program_statuses(program_statuses)
    await db.users.update_one(
        {"id": student_id},
        {"$set": {"estado": new_estado, "program_statuses": program_statuses}},
    )
    logger.info(
        f"Student {student_id} status updated to {program_statuses.get(prog_id)} "
        f"after all recovery subjects approved in course {course_id}"
    )
    await log_audit(
        "recovery_all_approved",
        "system",
        "system",
        {
            "student_id": student_id,
            "course_id": course_id,
            "new_status": program_statuses.get(prog_id),
            "trigger": "all_recoveries_graded_approved",
        },
    )


async def _check_and_update_recovery_rejection(student_id: str, course_id: str):
    """Check if a teacher rejection means the student definitively cannot pass recovery.

    Called after a teacher rejects a recovery subject. Only checks admin-approved records
    (recovery_approved=True) because non-approved subjects can't be teacher-graded.
    If ALL admin-approved records have been teacher-graded and at least one is rejected,
    the student cannot pass all subjects. In that case, immediately remove the student
    from the group and mark them as 'reprobado'.

    Note: this only checks admin-approved records (unlike _check_and_update_recovery_completion
    which checks ALL records) because we're determining if the student can still pass the
    subjects that were approved for recovery — unapproved subjects are already considered
    failed and will be handled at recovery_close by the scheduler.
    """
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        return
    prog_id = course.get("program_id", "")

    # Find all unprocessed, admin-approved recovery records for this student+course
    approved_records = await db.failed_subjects.find(
        {
            "student_id": student_id,
            "course_id": course_id,
            "recovery_approved": True,
            "recovery_processed": {"$ne": True},
            "recovery_expired": {"$ne": True},
        },
        {"_id": 0},
    ).to_list(None)

    if not approved_records:
        return

    # Check if ALL admin-approved records have been teacher-graded (completed)
    all_graded = all(r.get("recovery_completed") is True for r in approved_records)
    if not all_graded:
        return  # Some subjects still pending teacher grading

    # Check if at least one is rejected
    any_rejected = any(r.get("teacher_graded_status") == "rejected" for r in approved_records)
    if not any_rejected:
        return  # All approved — handled by _check_and_update_recovery_completion

    # Student definitively cannot pass: remove from group and mark reprobado
    student_doc = await db.users.find_one({"id": student_id}, {"_id": 0})
    if not student_doc:
        return
    program_statuses = student_doc.get("program_statuses") or {}
    if program_statuses.get(prog_id) != "pendiente_recuperacion":
        return

    # Remove only from the course where teacher rejection occurred.
    await _unenroll_student_from_course(student_id, course_id)
    # Mark reprobado
    program_statuses[prog_id] = "reprobado"
    new_estado = derive_estado_from_program_statuses(program_statuses)
    await db.users.update_one(
        {"id": student_id, "role": "estudiante"},
        {"$set": {"estado": new_estado, "program_statuses": program_statuses}}
    )
    # Keep records visible in admin/student recovery panels until recovery-close,
    # but mark them explicitly as rejected by teacher for status/reporting.
    now = datetime.now(timezone.utc)
    for record in approved_records:
        await db.failed_subjects.update_one(
            {"id": record["id"]},
            {"$set": {
                "recovery_rejected": True,
                "rejected_at": now.isoformat(),
                "rejected_by": "teacher"
            }}
        )
    logger.info(
        f"Student {student_id} marked reprobado and removed from group {course_id} "
        f"(teacher rejected recovery subject, all subjects resolved)"
    )
    await log_audit(
        "student_removed_from_group", "system", "system",
        {"student_id": student_id, "course_id": course_id,
         "program_id": prog_id, "trigger": "teacher_rejected_recovery"}
    )


def sanitize_string(input_str: str, max_length: int = 500) -> str:
    """Sanitize string input to prevent injection attacks"""
    if not input_str or not isinstance(input_str, str):
        return ""
    # Remove potential XSS/injection characters including quotes, parentheses, and script-related chars
    sanitized = re.sub(r'[<>{}()\'"\[\]\\;`]', '', input_str)
    # Remove any non-printable characters and control characters
    sanitized = ''.join(char for char in sanitized if char.isprintable())
    # Limit length
    return sanitized[:max_length]

def hash_password(password: str) -> str:
    """Store password based on PASSWORD_STORAGE_MODE (plain or bcrypt)
    
    WARNING: Plain text storage is insecure and should only be used for
    backwards compatibility with existing data. Set PASSWORD_STORAGE_MODE='bcrypt'
    for production systems.
    """
    if PASSWORD_STORAGE_MODE == 'plain':
        # Store password as plain text (for backwards compatibility with existing data)
        # WARNING: This is insecure! Only use for compatibility with existing data.
        logger.warning("Storing password as plain text (PASSWORD_STORAGE_MODE='plain'). "
                      "This is insecure. Consider using PASSWORD_STORAGE_MODE='bcrypt' for production.")
        return password
    else:
        # Hash password using bcrypt directly (avoids passlib compatibility issues)
        return _bcrypt.hashpw(password.encode('utf-8'), _bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against bcrypt hash, SHA256, or plain text"""
    # Check format first to avoid timing attacks
    # bcrypt hashes start with $2a$, $2b$, or $2y$
    if hashed_password.startswith(('$2a$', '$2b$', '$2y$')):
        # Stored as bcrypt hash - use bcrypt directly (avoids passlib compatibility issues)
        # bcrypt requires bytes input, so we encode strings to UTF-8
        try:
            return _bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))
        except Exception as e:
            logger.error(f"Bcrypt verification error: {type(e).__name__}")
            return False
    
    # Check if it's a SHA256 hash (64 hex characters)
    elif len(hashed_password) == 64 and all(c in '0123456789abcdef' for c in hashed_password.lower()):
        # Stored as SHA256 hash
        try:
            return hashlib.sha256(plain_password.encode()).hexdigest() == hashed_password
        except Exception:
            return False
    
    # Otherwise, try plain text comparison (for backwards compatibility)
    else:
        # Log warning for security audit
        if PASSWORD_STORAGE_MODE == 'plain':
            logger.debug("Plain text password comparison used (PASSWORD_STORAGE_MODE='plain')")
        else:
            logger.warning(f"Plain text password detected in database for backwards compatibility")
        return plain_password == hashed_password

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=2)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

# Lock for thread-safe access to login_attempts
login_attempts_lock = asyncio.Lock()

async def check_rate_limit(ip_address: str, identifier: str = None) -> bool:
    """Verifica límite por IP (50) y por identificador de usuario (5).
    En redes educativas con WiFi compartido, el límite por IP es alto
    para no bloquear a todos por los errores de uno."""
    async with login_attempts_lock:
        current_time = datetime.now().timestamp()
        # Limpiar intentos viejos por IP
        login_attempts[ip_address] = [
            t for t in login_attempts[ip_address]
            if current_time - t < LOGIN_ATTEMPT_WINDOW
        ]
        # Verificar límite por IP (protección anti-bots)
        if len(login_attempts[ip_address]) >= MAX_LOGIN_ATTEMPTS_PER_IP:
            return False
        # Verificar límite por identificador individual (cédula o email)
        if identifier:
            login_attempts_by_identifier[identifier] = [
                t for t in login_attempts_by_identifier[identifier]
                if current_time - t < LOGIN_ATTEMPT_WINDOW
            ]
            if len(login_attempts_by_identifier[identifier]) >= MAX_LOGIN_ATTEMPTS_PER_USER:
                return False
        return True

def log_security_event(event_type: str, details: dict):
    """Log security-related events with sanitized details"""
    # Sanitize details to prevent log injection
    sanitized_details = {}
    for key, value in details.items():
        if isinstance(value, str):
            sanitized_details[key] = sanitize_string(value, 200)
        else:
            sanitized_details[key] = str(value)[:200]
    logger.warning(f"SECURITY: {event_type} - {json.dumps(sanitized_details)}")

async def log_audit(action: str, user_id: str, user_role: str, details: dict):
    """Insert an audit log record into the audit_logs collection."""
    try:
        record = {
            "id": str(uuid.uuid4()),
            "action": action,
            "user_id": user_id,
            "user_role": user_role,
            "details": details,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        await db.audit_logs.insert_one(record)
    except Exception as exc:
        logger.error(f"Failed to write audit log (action={action}): {exc}")

async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="No autorizado")
    token = authorization.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id", "")
        # Validate user_id to prevent NoSQL injection via manipulated tokens.
        # Ensures user_id is a non-empty string without MongoDB operator prefixes;
        # dict/object payloads (e.g. {"$ne": None}) fail the isinstance check.
        if not isinstance(user_id, str) or not user_id or user_id.startswith('$'):
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# --- Pydantic Models ---
class LoginRequest(BaseModel):
    email: Optional[str] = None
    cedula: Optional[str] = None
    password: str = Field(..., min_length=1, max_length=200)
    role: str = Field(..., pattern="^(estudiante|profesor|admin|editor)$")

    @validator('email')
    def sanitize_email(cls, v):
        if v:
            return sanitize_string(v, 200)
        return v
    
    @validator('cedula')
    def sanitize_cedula(cls, v):
        if v:
            # Only allow numbers for cedula (consistent with UserCreate/UserUpdate)
            return re.sub(r'\D', '', v)[:50]
        return v

class AdminCreateByEditor(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    email: str = Field(..., min_length=1, max_length=200)
    password: str = Field(..., min_length=6, max_length=200)

    @validator('name', 'email')
    def sanitize_fields(cls, v):
        return sanitize_string(v, 200)

class AdminUpdateByEditor(BaseModel):
    name: Optional[str] = Field(None, min_length=1, max_length=200)
    email: Optional[str] = Field(None, min_length=1, max_length=200)
    password: Optional[str] = Field(None, min_length=6, max_length=200)
    active: Optional[bool] = None

    @validator('name', 'email')
    def sanitize_fields(cls, v):
        if v is not None:
            return sanitize_string(v, 200)
        return v

class UserCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    email: Optional[str] = Field(None, max_length=200)
    cedula: Optional[str] = Field(None, max_length=50)
    password: str = Field(..., min_length=6, max_length=200)
    role: str = Field(..., pattern="^(estudiante|profesor|admin)$")
    program_id: Optional[str] = None  # For backward compatibility
    program_ids: Optional[List[str]] = None  # Multiple programs support
    course_ids: Optional[List[str]] = None  # Ignored on creation; handled via course enrollment endpoints
    subject_ids: Optional[List[str]] = None  # For professors - subjects they teach
    phone: Optional[str] = Field(None, max_length=50)
    module: Optional[int] = Field(None, ge=1)  # Deprecated: use program_modules (no upper bound – N modules)
    program_modules: Optional[dict] = None  # Maps program_id to module number, e.g., {"prog-admin": 1, "prog-infancia": 2}
    program_statuses: Optional[dict] = None  # Maps program_id to status, e.g., {"prog-admin": "activo"}
    estado: Optional[str] = Field(None, pattern="^(activo|egresado|pendiente_recuperacion|retirado)$")  # Student status

    @validator('name', 'email', 'phone')
    def sanitize_text_fields(cls, v):
        if v:
            return sanitize_string(v, 200)
        return v
    
    @validator('cedula')
    def sanitize_cedula(cls, v):
        if v:
            # Only allow numbers for cedula
            cleaned = re.sub(r'\D', '', v)[:50]
            if not cleaned:
                raise ValueError('La cédula debe contener al menos un número')
            return cleaned
        return v
    
    @validator('email')
    def validate_email(cls, v):
        if v:
            # Email validation: must contain @ and a domain
            email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if not re.match(email_pattern, v):
                raise ValueError('El correo electrónico debe contener @ y un dominio válido')
        return v
    
    @validator('program_modules')
    def validate_program_modules(cls, v):
        if v is not None:
            for prog_id, module_num in v.items():
                validate_module_number(module_num, f"Module number for program {prog_id}")
        return v

class UserUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    email: Optional[str] = Field(None, max_length=200)
    cedula: Optional[str] = Field(None, max_length=50)
    password: Optional[str] = Field(None, min_length=6, max_length=200)  # Allow password updates
    phone: Optional[str] = Field(None, max_length=50)
    program_id: Optional[str] = None  # For backward compatibility
    program_ids: Optional[List[str]] = None  # Multiple programs support
    subject_ids: Optional[List[str]] = None  # For professors - subjects they teach
    active: Optional[bool] = None
    module: Optional[int] = Field(None, ge=1)  # Deprecated: use program_modules (no upper bound – N modules)
    program_modules: Optional[dict] = None  # Maps program_id to module number, e.g., {"prog-admin": 1, "prog-infancia": 2}
    program_statuses: Optional[dict] = None  # Maps program_id to status, e.g., {"prog-admin": "activo"}
    estado: Optional[str] = Field(None, pattern="^(activo|egresado|pendiente_recuperacion|retirado)$")  # Student status

    @validator('name', 'email', 'phone')
    def sanitize_text_fields(cls, v):
        if v:
            return sanitize_string(v, 200)
        return v
    
    @validator('cedula')
    def sanitize_cedula(cls, v):
        if v:
            # Only allow numbers for cedula
            cleaned = re.sub(r'\D', '', v)[:50]
            if not cleaned:
                raise ValueError('La cédula debe contener al menos un número')
            return cleaned
        return v
    
    @validator('email')
    def validate_email(cls, v):
        if v:
            # Email validation: must contain @ and a domain
            email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if not re.match(email_pattern, v):
                raise ValueError('El correo electrónico debe contener @ y un dominio válido')
        return v
    
    @validator('program_modules')
    def validate_program_modules(cls, v):
        if v is not None:
            for prog_id, module_num in v.items():
                validate_module_number(module_num, f"Module number for program {prog_id}")
        return v

class ProgramCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    duration: Optional[str] = "12 meses"
    modules: Optional[list] = []

class ProgramUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    duration: Optional[str] = None
    modules: Optional[list] = None
    active: Optional[bool] = None
    module1_close_date: Optional[str] = None
    module2_close_date: Optional[str] = None

class SubjectCreate(BaseModel):
    name: str
    program_id: str
    module_number: int = 1
    description: Optional[str] = ""

class SubjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    module_number: Optional[int] = None
    program_id: Optional[str] = None
    active: Optional[bool] = None

class CourseCreate(BaseModel):
    name: str
    program_id: str
    subject_id: Optional[str] = None  # For backward compatibility - single subject
    subject_ids: Optional[List[str]] = []  # Multiple subjects per course/group
    teacher_id: Optional[str] = None  # Optional - courses are cohorts/promotions
    year: int = datetime.now().year
    student_ids: Optional[List[str]] = []
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    grupo: Optional[str] = None  # e.g., "ENERO-2026 - TECNICO EN SISTEMAS"
    module_dates: Optional[dict] = {}  # e.g., {"1": {"start": "2026-01-01", "end": "2026-06-30"}, "2": {"start": "2026-07-01", "end": "2026-12-31"}}

class CourseUpdate(BaseModel):
    name: Optional[str] = None
    subject_id: Optional[str] = None  # For backward compatibility
    subject_ids: Optional[List[str]] = None  # Multiple subjects
    teacher_id: Optional[str] = None
    student_ids: Optional[List[str]] = None
    active: Optional[bool] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    grupo: Optional[str] = None  # e.g., "ENERO-2026 - TECNICO EN SISTEMAS"
    module_dates: Optional[dict] = None  # e.g., {"1": {"start": "2026-01-01", "end": "2026-06-30"}, "2": {"start": "2026-07-01", "end": "2026-12-31"}}

class ActivityCreate(BaseModel):
    course_id: str
    subject_id: Optional[str] = None  # Specific subject within the course/group
    title: str
    description: Optional[str] = ""
    start_date: Optional[str] = None
    due_date: str
    files: Optional[list] = []
    is_recovery: Optional[bool] = False

class ActivityUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    files: Optional[list] = None
    active: Optional[bool] = None
    is_recovery: Optional[bool] = None
    subject_id: Optional[str] = None

class GradeCreate(BaseModel):
    student_id: str
    course_id: str
    activity_id: Optional[str] = None
    subject_id: Optional[str] = None  # Specific subject within the course/group
    value: Optional[float] = None
    comments: Optional[str] = ""
    recovery_status: Optional[str] = None  # 'approved', 'rejected', or None

class GradeUpdate(BaseModel):
    value: Optional[float] = None
    comments: Optional[str] = None
    recovery_status: Optional[str] = None  # 'approved', 'rejected', or None

class ClassVideoCreate(BaseModel):
    course_id: str
    subject_id: Optional[str] = None  # Specific subject within the course/group
    title: str
    url: str
    description: Optional[str] = ""
    available_from: Optional[str] = None  # ISO datetime; if set, students only see this after this date

class SubmissionCreate(BaseModel):
    activity_id: str
    content: Optional[str] = ""
    files: Optional[list] = []

class RecoveryEnableRequest(BaseModel):
    student_id: str
    course_id: str
    subject_id: Optional[str] = None

class ModuleCloseDateUpdate(BaseModel):
    module1_close_date: Optional[str] = None
    module2_close_date: Optional[str] = None

# --- Auth Routes ---
@api_router.post("/auth/login")
async def login(req: LoginRequest, request: Request):
    # Get client IP for rate limiting
    client_ip = request.client.host if request.client else "unknown"
    
    # Check rate limit
    identifier = req.cedula if req.role == "estudiante" else req.email
    if not await check_rate_limit(client_ip, identifier):
        log_security_event("RATE_LIMIT_EXCEEDED", {
            "ip": client_ip,
            "role": req.role,
            "identifier": req.email or req.cedula
        })
        raise HTTPException(
            status_code=429, 
            detail="Demasiados intentos de inicio de sesión. Por favor, intente más tarde."
        )
    
    # Record login attempt (with lock)
    async with login_attempts_lock:
        login_attempts[client_ip].append(datetime.now().timestamp())
        if identifier:
            login_attempts_by_identifier[identifier].append(datetime.now().timestamp())
    
    # Validate input and find user
    if req.role == "estudiante":
        if not req.cedula:
            raise HTTPException(status_code=400, detail="Cédula requerida")
        user = await db.users.find_one(
            {"cedula": req.cedula, "role": "estudiante"}, 
            {"_id": 0}
        )
        identifier = req.cedula
    elif req.role in ["profesor", "admin", "editor"]:
        if not req.email:
            raise HTTPException(status_code=400, detail="Correo requerido")
        # Profesor tab continues supporting admin/editor, and direct role login now works too
        role_filter = {"$in": ["profesor", "admin", "editor"]} if req.role == "profesor" else req.role
        user = await db.users.find_one({"email": req.email, "role": role_filter}, {"_id": 0})
        identifier = req.email
    else:
        # Invalid role
        log_security_event("INVALID_ROLE", {"role": req.role, "ip": client_ip})
        raise HTTPException(status_code=400, detail="Rol inválido")
    
    if not user:
        log_security_event("LOGIN_FAILED_USER_NOT_FOUND", {
            "ip": client_ip,
            "role": req.role,
            "identifier": identifier
        })
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    
    if not verify_password(req.password, user["password_hash"]):
        log_security_event("LOGIN_FAILED_WRONG_PASSWORD", {
            "ip": client_ip,
            "role": req.role,
            "user_id": user["id"],
            "identifier": identifier
        })
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    
    if not user.get("active", True):
        log_security_event("LOGIN_FAILED_INACTIVE_ACCOUNT", {
            "ip": client_ip,
            "user_id": user["id"],
            "identifier": identifier
        })
        raise HTTPException(status_code=403, detail="Cuenta desactivada")
    
    # Successful login - clear attempts for this IP (with lock)
    async with login_attempts_lock:
        login_attempts[client_ip] = []
        if identifier:
            login_attempts_by_identifier[identifier] = []
    
    logger.info(f"Successful login: user_id={user['id']}, role={user['role']}, ip={client_ip}")
    await log_audit("login_success", user["id"], user["role"], {"ip": client_ip})
    
    token = create_token(user["id"], user["role"])
    user_data = {k: v for k, v in user.items() if k != "password_hash"}
    return {"token": token, "user": user_data}

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    user_data = {k: v for k, v in user.items() if k != "password_hash"}
    return user_data

@api_router.get("/me/subjects")
async def get_my_subjects(user=Depends(get_current_user)):
    """Return the full Subject objects assigned to the authenticated teacher.
    Only accessible by roles 'profesor' and 'admin'.
    Subject IDs that no longer exist in the DB are silently ignored."""
    if user["role"] not in ["profesor", "admin"]:
        raise HTTPException(status_code=403, detail="Solo profesores pueden acceder a sus materias")
    subject_ids = user.get("subject_ids") or []
    if not subject_ids:
        return []
    subjects = await db.subjects.find({"id": {"$in": subject_ids}}, {"_id": 0}).to_list(500)
    return subjects

# --- Users Routes ---
@api_router.get("/users")
async def get_users(
    role: Optional[str] = None,
    estado: Optional[str] = None,
    search: Optional[str] = None,
    program_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user=Depends(get_current_user)
):
    if user["role"] not in ["admin", "profesor"]:
        raise HTTPException(status_code=403, detail="No autorizado")

    if page < 1:
        page = 1
    if page_size < 1 or page_size > 200:
        page_size = 50

    query = {}
    if role:
        query["role"] = role
    if estado:
        # Treat missing/null estado as 'activo' so newly created users are visible
        if estado == 'activo':
            query["$or"] = [{"estado": "activo"}, {"estado": None}, {"estado": {"$exists": False}}]
        else:
            query["estado"] = estado

    # Program filter
    if program_id:
        program_filter = {"$or": [{"program_id": program_id}, {"program_ids": program_id}]}
        if "$or" in query:
            # Already have $or from estado filter — combine with $and
            estado_or = query.pop("$or")
            query["$and"] = [{"$or": estado_or}, program_filter]
        else:
            query["$or"] = [{"program_id": program_id}, {"program_ids": program_id}]

    # Text search by name or cedula
    if search and search.strip():
        search_regex = {"$regex": search.strip(), "$options": "i"}
        search_cond = {"$or": [{"name": search_regex}, {"cedula": search_regex}]}
        if "$and" in query:
            query["$and"].append(search_cond)
        elif "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, search_cond]
        else:
            query.update(search_cond)

    total = await db.users.count_documents(query)
    skip = (page - 1) * page_size
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("name", 1).skip(skip).limit(page_size).to_list(page_size)

    return {
        "users": users,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size)
    }

@api_router.post("/users")
async def create_user(req: UserCreate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        log_security_event("UNAUTHORIZED_USER_CREATE_ATTEMPT", {
            "attempted_by": user["id"],
            "attempted_role": user["role"]
        })
        raise HTTPException(status_code=403, detail="Solo admin puede crear usuarios")
    
    # Validate unique email for admin/profesor
    if req.role in ["admin", "profesor"] and req.email:
        existing = await db.users.find_one({"email": req.email})
        if existing:
            log_security_event("DUPLICATE_EMAIL_ATTEMPT", {
                "email": req.email,
                "attempted_by": user["id"]
            })
            raise HTTPException(status_code=400, detail="Email ya existe")
    
    # Validate unique cedula for estudiante
    if req.role == "estudiante" and req.cedula:
        existing = await db.users.find_one({"cedula": req.cedula})
        if existing:
            log_security_event("DUPLICATE_CEDULA_ATTEMPT", {
                "cedula": req.cedula,
                "attempted_by": user["id"]
            })
            raise HTTPException(status_code=400, detail="Cédula ya existe")
    
    # Determine program_ids: use provided program_ids, or convert program_id to list, or empty list
    program_ids = []
    if req.program_ids:
        program_ids = req.program_ids
    elif req.program_id:
        program_ids = [req.program_id]
    
    # Rule 1: Students must be enrolled in at least one technical program
    if req.role == "estudiante" and not program_ids:
        raise HTTPException(status_code=400, detail="El estudiante debe estar inscrito en al menos un programa técnico")
    
    # Handle subject_ids for professors
    subject_ids = req.subject_ids if req.subject_ids else []
    
    # Rule 2: Teacher-subject uniqueness — a subject can only be assigned to one professor
    if req.role == "profesor" and subject_ids:
        for subject_id in subject_ids:
            conflict = await db.users.find_one(
                {"role": "profesor", "subject_ids": subject_id},
                {"_id": 0, "id": 1, "name": 1}
            )
            if conflict:
                raise HTTPException(
                    status_code=400,
                    detail=f"La materia ya está asignada al profesor '{conflict['name']}'. Desasígnela primero antes de asignarla a otro profesor."
                )
    
    # Set default estado for students
    estado = req.estado if req.estado else ("activo" if req.role == "estudiante" else None)
    
    # Handle program_modules: if provided, use it; otherwise, if module is provided and we have program_ids, 
    # initialize program_modules with the same module for all programs
    if req.program_modules:
        program_modules = req.program_modules
    elif req.module and program_ids:
        # Initialize all programs with the same module for backward compatibility
        program_modules = {prog_id: req.module for prog_id in program_ids}
    else:
        program_modules = None
    
    # Initialize program_statuses for students: use provided value or default all programs to "activo"
    if req.role == "estudiante":
        if req.program_statuses:
            program_statuses = req.program_statuses
        elif program_ids:
            program_statuses = {prog_id: "activo" for prog_id in program_ids}
        else:
            program_statuses = None
        # Derive global estado from program_statuses if not explicitly given
        if program_statuses and not req.estado:
            estado = derive_estado_from_program_statuses(program_statuses)
    else:
        program_statuses = None
    
    new_user = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "cedula": req.cedula,
        "password_hash": hash_password(req.password),
        "role": req.role,
        "program_id": req.program_id,  # Keep for backward compatibility
        "program_ids": program_ids,
        "subject_ids": subject_ids,
        "phone": req.phone,
        "module": req.module,  # Keep for backward compatibility
        "program_modules": program_modules,
        "program_statuses": program_statuses,
        "estado": estado,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    # Only include email in the document when a value is provided.
    # Omitting the field (rather than storing null) ensures the sparse unique
    # email index skips this document, allowing multiple users without an email.
    if req.email is not None:
        new_user["email"] = req.email
    try:
        await db.users.insert_one(new_user)
    except Exception as exc:
        logger.error(f"Error inserting user into database: {exc}")
        raise HTTPException(status_code=500, detail="No se pudo crear el usuario. Inténtelo de nuevo.")
    
    logger.info(f"User created: id={new_user['id']}, role={req.role}, by={user['id']}")
    await log_audit("student_created" if req.role == "estudiante" else "user_created", user["id"], user["role"], {"new_user_id": new_user["id"], "new_user_role": req.role, "new_user_name": req.name})
    
    del new_user["_id"]
    del new_user["password_hash"]
    return new_user

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        log_security_event("UNAUTHORIZED_USER_UPDATE_ATTEMPT", {
            "attempted_by": user["id"],
            "target_user": user_id
        })
        raise HTTPException(status_code=403, detail="Solo admin puede editar usuarios")
    
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    
    # Hash password if provided and not empty
    if "password" in update_data:
        password = update_data.pop("password")
        if password is not None and password.strip():
            update_data["password_hash"] = hash_password(password)
            logger.info(f"Password updated for user: {user_id} by admin: {user['id']}")
        # If password is empty/whitespace, just ignore it (don't update)
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    # Validate cedula uniqueness if it's being changed
    if "cedula" in update_data and update_data["cedula"]:
        existing = await db.users.find_one({"cedula": update_data["cedula"], "id": {"$ne": user_id}})
        if existing:
            log_security_event("DUPLICATE_CEDULA_UPDATE_ATTEMPT", {
                "cedula": update_data["cedula"],
                "attempted_by": user["id"],
                "target_user": user_id
            })
            raise HTTPException(status_code=400, detail="Esta cédula ya está registrada")
    
    # Validate email uniqueness if it's being changed
    if "email" in update_data and update_data["email"]:
        existing = await db.users.find_one({"email": update_data["email"], "id": {"$ne": user_id}})
        if existing:
            log_security_event("DUPLICATE_EMAIL_UPDATE_ATTEMPT", {
                "email": update_data["email"],
                "attempted_by": user["id"],
                "target_user": user_id
            })
            raise HTTPException(status_code=400, detail="Este correo ya está registrado")
    
    # Rule 2: Teacher-subject uniqueness — a subject can only be assigned to one professor
    if "subject_ids" in update_data and update_data["subject_ids"]:
        target_user = await db.users.find_one({"id": user_id}, {"_id": 0, "role": 1})
        if target_user and target_user.get("role") == "profesor":
            for subject_id in update_data["subject_ids"]:
                conflict = await db.users.find_one(
                    {"role": "profesor", "subject_ids": subject_id, "id": {"$ne": user_id}},
                    {"_id": 0, "id": 1, "name": 1}
                )
                if conflict:
                    raise HTTPException(
                        status_code=400,
                        detail=f"La materia ya está asignada al profesor '{conflict['name']}'. Desasígnela primero antes de asignarla a otro profesor."
                    )

    # Rule 3: When adding new programs to a student, initialize program_statuses and program_modules
    if "program_ids" in update_data and update_data["program_ids"] is not None:
        current_student = await db.users.find_one({"id": user_id, "role": "estudiante"}, {"_id": 0, "program_ids": 1, "program_id": 1, "program_statuses": 1, "program_modules": 1})
        if current_student:
            current_program_ids = current_student.get("program_ids") or (
                [current_student["program_id"]] if current_student.get("program_id") else []
            )
            new_program_ids = update_data["program_ids"]
            added_programs = [p for p in new_program_ids if p not in current_program_ids]
            if added_programs:
                stored_statuses = current_student.get("program_statuses") or {}
                stored_modules = current_student.get("program_modules") or {}
                # Merge: stored as base, frontend-sent values take priority
                sent_statuses = update_data.get("program_statuses") if update_data.get("program_statuses") is not None else {}
                sent_modules = update_data.get("program_modules") if update_data.get("program_modules") is not None else {}
                merged_statuses = {**stored_statuses, **sent_statuses}
                merged_modules = {**stored_modules, **sent_modules}
                for prog_id in added_programs:
                    if prog_id not in merged_statuses:
                        merged_statuses[prog_id] = "activo"
                    if prog_id not in merged_modules:
                        merged_modules[prog_id] = 1
                update_data["program_statuses"] = merged_statuses
                update_data["program_modules"] = merged_modules

    # Always recalculate estado from program_statuses when they are present,
    # ensuring egresado → activo when a new program is added (covers both the
    # added_programs path and edge cases where statuses changed without new programs).
    if update_data.get("program_statuses"):
        update_data["estado"] = derive_estado_from_program_statuses(update_data["program_statuses"])

    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Log important updates for debugging persistence issues
    if "subject_ids" in update_data:
        logger.info(f"User subject assignment updated: user_id={user_id}, subject_ids={update_data['subject_ids']}, by={user['id']}")
    
    logger.info(f"User updated: id={user_id}, by={user['id']}, fields={list(update_data.keys())}")
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    await log_audit("user_updated", user["id"], user["role"], {"target_user_id": user_id})
    return updated

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede eliminar usuarios")
    target = await db.users.find_one({"id": user_id}, {"_id": 0, "name": 1, "role": 1})
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    # Remove the deleted user from all course student_ids arrays so counts stay accurate
    await db.courses.update_many(
        {"student_ids": user_id},
        {"$pull": {"student_ids": user_id}}
    )
    # Delete all data associated with this user to avoid orphaned records
    await db.grades.delete_many({"student_id": user_id})
    await db.submissions.delete_many({"student_id": user_id})
    await db.failed_subjects.delete_many({"student_id": user_id})
    await db.recovery_enabled.delete_many({"student_id": user_id})
    await log_audit("user_deleted", user["id"], user["role"], {"deleted_user_id": user_id, "deleted_user_name": (target or {}).get("name", ""), "deleted_user_role": (target or {}).get("role", "")})
    return {"message": "Usuario eliminado"}

# --- Editor Routes ---
@api_router.post("/editor/create-admin")
async def editor_create_admin(req: AdminCreateByEditor, user=Depends(get_current_user)):
    """Endpoint for editor to create admin users"""
    if user["role"] != "editor":
        raise HTTPException(status_code=403, detail="Solo editor puede crear administradores")
    
    if not req.email or not req.password or not req.name:
        raise HTTPException(status_code=400, detail="Email, password y nombre son requeridos")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": req.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya existe")
    
    new_admin = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "email": req.email,
        "cedula": None,
        "password_hash": hash_password(req.password),
        "role": "admin",
        "program_id": None,
        "program_ids": [],
        "subject_ids": [],
        "phone": None,
        "module": None,
        "grupo": None,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_admin)
    del new_admin["_id"]
    del new_admin["password_hash"]
    return new_admin

@api_router.get("/editor/admins")
async def editor_get_admins(user=Depends(get_current_user)):
    """Endpoint for editor to list all admins"""
    if user["role"] != "editor":
        raise HTTPException(status_code=403, detail="Solo editor puede ver administradores")
    
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return admins

@api_router.put("/editor/admins/{admin_id}")
async def editor_update_admin(admin_id: str, req: AdminUpdateByEditor, user=Depends(get_current_user)):
    """Endpoint for editor to update admin users. At least one field must be provided."""
    if user["role"] != "editor":
        raise HTTPException(status_code=403, detail="Solo editor puede editar administradores")
    
    # Check that the admin exists and is actually an admin
    target_user = await db.users.find_one({"id": admin_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Administrador no encontrado")
    
    if target_user["role"] != "admin":
        raise HTTPException(status_code=400, detail="El usuario no es un administrador")
    
    # Build update data
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    
    # Hash password if provided
    if "password" in update_data:
        password = update_data.pop("password")
        if password and password.strip():
            update_data["password_hash"] = hash_password(password)
            logger.info(f"Password updated for admin: {admin_id} by editor: {user['id']}")
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    # Validate email uniqueness if it's being changed
    if "email" in update_data and update_data["email"]:
        existing = await db.users.find_one({"email": update_data["email"], "id": {"$ne": admin_id}})
        if existing:
            log_security_event("DUPLICATE_EMAIL_UPDATE_ATTEMPT", {
                "email": update_data["email"],
                "attempted_by": user["id"],
                "target_user": admin_id
            })
            raise HTTPException(status_code=400, detail="Este correo ya está registrado")
    
    # Update the admin
    result = await db.users.update_one({"id": admin_id}, {"$set": update_data})
    
    logger.info(f"Admin updated: id={admin_id}, by={user['id']}, fields={list(update_data.keys())}")
    
    # Return updated admin without sensitive data
    updated_admin = await db.users.find_one({"id": admin_id}, {"_id": 0, "password_hash": 0})
    return updated_admin

@api_router.delete("/editor/admins/{admin_id}")
async def editor_delete_admin(admin_id: str, user=Depends(get_current_user)):
    """Endpoint for editor to delete admin users"""
    if user["role"] != "editor":
        raise HTTPException(status_code=403, detail="Solo editor puede eliminar administradores")
    
    # Check that the admin exists and is actually an admin
    target_user = await db.users.find_one({"id": admin_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Administrador no encontrado")
    
    if target_user["role"] != "admin":
        raise HTTPException(status_code=400, detail="El usuario no es un administrador")
    
    # Delete the admin
    result = await db.users.delete_one({"id": admin_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Administrador no encontrado")
    
    logger.info(f"Admin deleted: id={admin_id}, by editor={user['id']}")
    log_security_event("ADMIN_DELETED_BY_EDITOR", {
        "deleted_admin_id": admin_id,
        "deleted_admin_email": target_user.get("email"),
        "editor_id": user["id"]
    })
    
    return {"message": "Administrador eliminado exitosamente"}

# --- Programs Routes ---
@api_router.get("/programs")
async def get_programs():
    programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    return programs

@api_router.post("/programs")
async def create_program(req: ProgramCreate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede crear programas")
    program = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "description": req.description,
        "duration": req.duration,
        "modules": req.modules,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.programs.insert_one(program)
    del program["_id"]
    await log_audit("program_created", user["id"], user["role"], {"program_id": program["id"], "program_name": req.name})
    return program

@api_router.put("/programs/{program_id}")
async def update_program(program_id: str, req: ProgramUpdate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    result = await db.programs.update_one({"id": program_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Programa no encontrado")
    updated = await db.programs.find_one({"id": program_id}, {"_id": 0})
    return updated

@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    await db.programs.delete_one({"id": program_id})
    await log_audit("program_deleted", user["id"], user["role"], {"program_id": program_id})
    return {"message": "Programa eliminado"}

@api_router.get("/student/programs")
async def get_student_programs(user=Depends(get_current_user)):
    """Get programs that a student is enrolled in"""
    if user["role"] != "estudiante":
        raise HTTPException(status_code=403, detail="Solo estudiantes")
    
    program_ids = user.get("program_ids", [])
    if not program_ids and user.get("program_id"):
        # Backward compatibility
        program_ids = [user.get("program_id")]
    
    if not program_ids:
        return []
    
    programs = await db.programs.find({"id": {"$in": program_ids}}, {"_id": 0}).to_list(100)
    return programs

# --- Subjects Routes ---
@api_router.get("/subjects")
async def get_subjects(program_id: Optional[str] = None, teacher_id: Optional[str] = None):
    query = {}
    if program_id:
        query["program_id"] = program_id
    if teacher_id:
        teacher = await db.users.find_one({"id": teacher_id, "role": "profesor"}, {"_id": 0})
        if teacher and teacher.get("subject_ids"):
            query["id"] = {"$in": teacher["subject_ids"]}
        else:
            return []
    subjects = await db.subjects.find(query, {"_id": 0}).to_list(500)
    return subjects

@api_router.get("/subjects/teachers")
async def get_subjects_teachers(user=Depends(get_current_user)):
    """Return a map of subject_id -> comma-separated teacher names for all teachers.
    Accessible by all authenticated users (including students) so the
    StudentCourseSelector can display the teacher's name per subject
    without requiring admin/profesor permissions.
    If multiple teachers share the same subject, all their names are included."""
    teachers = await db.users.find(
        {"role": "profesor"},
        {"_id": 0, "id": 1, "name": 1, "subject_ids": 1}
    ).to_list(500)
    names_by_subject: dict = {}
    for t in teachers:
        for sid in (t.get("subject_ids") or []):
            names_by_subject.setdefault(sid, []).append(t["name"])
    return {sid: ", ".join(names) for sid, names in names_by_subject.items()}

@api_router.post("/subjects")
async def create_subject(req: SubjectCreate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    subject = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "program_id": req.program_id,
        "module_number": req.module_number,
        "description": req.description,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.subjects.insert_one(subject)
    del subject["_id"]
    return subject

@api_router.get("/subjects/{subject_id}")
async def get_subject(subject_id: str):
    subject = await db.subjects.find_one({"id": subject_id}, {"_id": 0})
    if not subject:
        raise HTTPException(status_code=404, detail="Materia no encontrada")
    return subject

@api_router.put("/subjects/{subject_id}")
async def update_subject(subject_id: str, req: SubjectUpdate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    result = await db.subjects.update_one({"id": subject_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Materia no encontrada")
    updated = await db.subjects.find_one({"id": subject_id}, {"_id": 0})
    return updated

@api_router.delete("/subjects/{subject_id}")
async def delete_subject(subject_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    await db.subjects.delete_one({"id": subject_id})
    return {"message": "Materia eliminada"}

# --- Courses Routes ---
@api_router.get("/courses")
async def get_courses(teacher_id: Optional[str] = None, student_id: Optional[str] = None, user=Depends(get_current_user)):
    conditions = []
    
    if teacher_id:
        # Get teacher's assigned subjects
        teacher = await db.users.find_one({"id": teacher_id}, {"_id": 0})
        if teacher and teacher.get("subject_ids"):
            # Match courses that have any subject in common with teacher's subjects
            # Also include courses explicitly assigned to this teacher (backward compatibility)
            conditions.append({
                "$or": [
                    {"teacher_id": teacher_id},
                    {"subject_ids": {"$in": teacher["subject_ids"]}},
                    {"subject_id": {"$in": teacher["subject_ids"]}}  # Backward compatibility
                ]
            })
        else:
            # Fallback to old behavior if teacher has no subjects assigned
            conditions.append({"teacher_id": teacher_id})
    
    if student_id:
        conditions.append({"student_ids": student_id})
    
    # Build final query
    if len(conditions) == 0:
        query = {}
    elif len(conditions) == 1:
        query = conditions[0]
    else:
        query = {"$and": conditions}
    
    courses = await db.courses.find(query, {"_id": 0}).to_list(2000)
    return courses

@api_router.get("/courses/{course_id}")
async def get_course(course_id: str, user=Depends(get_current_user)):
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    return course

@api_router.get("/courses/{course_id}/students")
async def get_course_students(course_id: str, user=Depends(get_current_user)):
    """Return students enrolled in a specific course. Much more efficient than
    fetching all students and filtering client-side."""
    if user["role"] not in ["admin", "profesor"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    student_ids = course.get("student_ids") or []
    if not student_ids:
        return []
    students = await db.users.find(
        {"id": {"$in": student_ids}, "role": "estudiante"},
        {"_id": 0, "password_hash": 0}
    ).to_list(5000)
    return students

@api_router.post("/courses")
async def create_course(req: CourseCreate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    
    # Validate: no duplicate group name for the same program
    existing_group = await db.courses.find_one({"name": req.name, "program_id": req.program_id}, {"_id": 0, "id": 1})
    if existing_group:
        raise HTTPException(status_code=400, detail="Ya existe un grupo con ese nombre para este programa")
    
    # Validate module date order (next module must start after previous module's recovery close)
    module_dates = req.module_dates or {}
    date_order_error = validate_module_dates_order(module_dates)
    if date_order_error:
        raise HTTPException(status_code=400, detail=date_order_error)
    
    # Business rule (2026-02-25): all module_dates entries must define recovery_close.
    recovery_close_error = validate_module_dates_recovery_close(module_dates)
    if recovery_close_error:
        raise HTTPException(status_code=400, detail=recovery_close_error)
    
    # Validate enrollment deadline per module
    student_ids_to_add = req.student_ids or []
    if student_ids_to_add:
        # Determine the current module for this course based on module_dates
        course_current_module = get_open_enrollment_module(module_dates) or get_current_module_from_dates(module_dates) or 1
        if not can_enroll_in_module(module_dates, course_current_module):
            # Enrollment window closed – only allow reingreso for retirado students in the current module
            if req.program_id:
                students_info = await db.users.find(
                    {"id": {"$in": student_ids_to_add}, "role": "estudiante"},
                    {"_id": 0, "id": 1, "program_statuses": 1, "program_modules": 1}
                ).to_list(None)
                student_map = {s["id"]: s for s in students_info}
                for sid in student_ids_to_add:
                    s = student_map.get(sid)
                    if not s:
                        raise HTTPException(status_code=400, detail=f"Estudiante {sid} no encontrado")
                    prog_statuses = s.get("program_statuses") or {}
                    prog_modules = s.get("program_modules") or {}
                    status = prog_statuses.get(req.program_id)
                    student_module = prog_modules.get(req.program_id)
                    if status == "retirado" and student_module is not None and student_module == course_current_module:
                        continue  # Reingreso allowed
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"No se puede matricular estudiantes: el período de matrícula para el "
                            f"Módulo {course_current_module} ha cerrado. Solo se permite reingreso de "
                            "estudiantes retirados en el módulo actual del grupo."
                        )
                    )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"No se puede matricular estudiantes: el período de matrícula para el Módulo {course_current_module} ha cerrado"
                )
        
        # Issue 5 (2026-02-25): validate that each student's program module matches the group's module.
        if req.program_id:
            students_mod_info = await db.users.find(
                {"id": {"$in": student_ids_to_add}, "role": "estudiante"},
                {"_id": 0, "id": 1, "program_modules": 1}
            ).to_list(None)
            student_mod_map = {s["id"]: s for s in students_mod_info}
            for sid in student_ids_to_add:
                s = student_mod_map.get(sid) or {}
                prog_modules = s.get("program_modules") or {}
                student_mod = prog_modules.get(req.program_id)
                if student_mod is not None and student_mod != course_current_module:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"El estudiante está en el Módulo {student_mod} del programa, pero el grupo "
                            f"corresponde al Módulo {course_current_module}. Solo puede inscribirse en "
                            "grupos del módulo en que se encuentra."
                        )
                    )

    # Validate: a student cannot be in 2+ groups of the same program
    if student_ids_to_add and req.program_id:
        conflicting_groups = await db.courses.find(
            {"program_id": req.program_id, "student_ids": {"$in": student_ids_to_add}},
            {"_id": 0, "name": 1, "student_ids": 1}
        ).to_list(None)
        if conflicting_groups:
            conflict_names = [g["name"] for g in conflicting_groups]
            raise HTTPException(
                status_code=400,
                detail=f"Uno o más estudiantes ya están inscritos en otro grupo del mismo programa: {', '.join(conflict_names)}"
            )
    
    # Handle subject_ids: use provided subject_ids, or convert subject_id to list if provided
    # Always ensure subject_ids is a list, never None or empty when subject_id is provided
    subject_ids = []
    if req.subject_ids:
        subject_ids = req.subject_ids
    elif req.subject_id:
        subject_ids = [req.subject_id]
    
    course = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "program_id": req.program_id,
        "subject_id": req.subject_id,  # Keep for backward compatibility
        "subject_ids": subject_ids if subject_ids else [],  # Always set as list, never None
        "teacher_id": req.teacher_id,
        "year": req.year,
        "student_ids": student_ids_to_add,
        "start_date": req.start_date,
        "end_date": req.end_date,
        "grupo": req.grupo,
        "module_dates": module_dates,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Log course creation for debugging persistence issues
    logger.info(f"Creating course: id={course['id']}, name={course['name']}, subject_ids={course['subject_ids']}, student_ids={len(course['student_ids'])} students")
    
    await db.courses.insert_one(course)
    del course["_id"]
    
    # Assign module to enrolled students based on module_dates (current date) or subject module_number
    if course["student_ids"]:
        program_id = course["program_id"]
        # Prefer date-based module determination
        module_number = get_current_module_from_dates(course["module_dates"])
        if module_number is None and course["subject_ids"]:
            # Fall back to minimum module number from subjects
            subject_docs = await db.subjects.find(
                {"id": {"$in": course["subject_ids"]}}, {"_id": 0, "module_number": 1}
            ).to_list(None)
            valid_modules = [s["module_number"] for s in subject_docs if s.get("module_number")]
            if valid_modules:
                module_number = min(valid_modules)
        if module_number is not None:
            await db.users.update_many(
                {"id": {"$in": course["student_ids"]}, "role": "estudiante"},
                {"$set": {
                    "module": module_number,
                    f"program_modules.{program_id}": module_number
                }}
            )
    
    await log_audit("course_created", user["id"], user["role"], {"course_id": course["id"], "course_name": course.get("name", "")})
    return course

@api_router.put("/courses/{course_id}")
async def update_course(course_id: str, req: CourseUpdate, user=Depends(get_current_user)):
    if user["role"] not in ["admin", "profesor"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    module_dates_updated = "module_dates" in update_data
    
    # Validate module date order if module_dates are being updated
    if "module_dates" in update_data and update_data["module_dates"]:
        date_order_error = validate_module_dates_order(update_data["module_dates"])
        if date_order_error:
            raise HTTPException(status_code=400, detail=date_order_error)
        # Business rule (2026-02-25): all module_dates entries must define recovery_close.
        recovery_close_error = validate_module_dates_recovery_close(update_data["module_dates"])
        if recovery_close_error:
            raise HTTPException(status_code=400, detail=recovery_close_error)
    
    # Handle subject_ids backward compatibility: if subject_ids provided, also update subject_id for compatibility
    if "subject_ids" in update_data and update_data["subject_ids"]:
        if "subject_id" not in update_data:
            update_data["subject_id"] = update_data["subject_ids"][0]
    elif "subject_id" in update_data and update_data["subject_id"]:
        # If only subject_id provided, ensure subject_ids includes it
        if "subject_ids" not in update_data:
            update_data["subject_ids"] = [update_data["subject_id"]]
    
    # Validate: a student cannot be in 2+ groups of the same program
    if req.student_ids is not None and user["role"] == "admin":
        # Get the current course to know its program, existing student list, and module_dates
        current_course = await db.courses.find_one({"id": course_id}, {"_id": 0, "program_id": 1, "student_ids": 1, "module_dates": 1})
        if current_course:
            # Check enrollment deadline per module
            current_student_ids = set(current_course.get("student_ids") or [])
            newly_added_ids = list(set(req.student_ids) - current_student_ids)
            # Use updated module_dates if provided (admin may be adjusting dates to allow enrollment)
            course_module_dates = (req.module_dates if req.module_dates is not None else current_course.get("module_dates")) or {}
            course_current_module = get_open_enrollment_module(course_module_dates) or get_current_module_from_dates(course_module_dates) or 1
            if newly_added_ids and not can_enroll_in_module(course_module_dates, course_current_module):
                # Enrollment window closed – only allow reingreso for retirado students in current module
                prog_id = current_course.get("program_id")
                if prog_id:
                    students_info = await db.users.find(
                        {"id": {"$in": newly_added_ids}, "role": "estudiante"},
                        {"_id": 0, "id": 1, "program_statuses": 1, "program_modules": 1}
                    ).to_list(None)
                    student_map = {s["id"]: s for s in students_info}
                    for sid in newly_added_ids:
                        s = student_map.get(sid)
                        if not s:
                            raise HTTPException(status_code=400, detail=f"Estudiante {sid} no encontrado")
                        prog_statuses = s.get("program_statuses") or {}
                        prog_modules = s.get("program_modules") or {}
                        status = prog_statuses.get(prog_id)
                        student_module = prog_modules.get(prog_id)
                        if status == "retirado" and student_module is not None and student_module == course_current_module:
                            continue  # Reingreso allowed
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"No se puede matricular estudiantes: el período de matrícula para el "
                                f"Módulo {course_current_module} ha cerrado. Solo se permite reingreso de "
                                "estudiantes retirados en el módulo actual del grupo."
                            )
                        )
                else:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No se puede matricular estudiantes: el período de matrícula para el Módulo {course_current_module} ha cerrado"
                    )

            program_id = current_course.get("program_id")
            if program_id and req.student_ids:
                conflicting_groups = await db.courses.find(
                    {"program_id": program_id, "id": {"$ne": course_id}, "student_ids": {"$in": req.student_ids}},
                    {"_id": 0, "name": 1, "student_ids": 1}
                ).to_list(None)
                if conflicting_groups:
                    conflict_names = [g["name"] for g in conflicting_groups]
                    raise HTTPException(
                        status_code=400,
                        detail=f"Uno o más estudiantes ya están inscritos en otro grupo del mismo programa: {', '.join(conflict_names)}"
                    )
            
            # Block students who previously failed and were removed from this specific group
            removed_ids = set(current_course.get("removed_student_ids") or [])
            if removed_ids and newly_added_ids:
                blocked = [sid for sid in newly_added_ids if sid in removed_ids]
                if blocked:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"No se puede re-matricular a {len(blocked)} estudiante(s) que fueron retirados "
                            "de este grupo por no pasar la recuperación. Deben inscribirse en un nuevo grupo."
                        )
                    )
            
            # Issue 5 (2026-02-25): validate module match for newly added students.
            if program_id and newly_added_ids:
                students_mod_info = await db.users.find(
                    {"id": {"$in": newly_added_ids}, "role": "estudiante"},
                    {"_id": 0, "id": 1, "program_modules": 1}
                ).to_list(None)
                for s in students_mod_info:
                    prog_modules = s.get("program_modules") or {}
                    student_mod = prog_modules.get(program_id)
                    if student_mod is not None and student_mod != course_current_module:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"El estudiante está en el Módulo {student_mod} del programa, pero el grupo "
                                f"corresponde al Módulo {course_current_module}. Solo puede inscribirse en "
                                "grupos del módulo en que se encuentra."
                            )
                        )
    
    result = await db.courses.update_one({"id": course_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    updated = await db.courses.find_one({"id": course_id}, {"_id": 0})
    
    # Update module for enrolled students when student_ids or module_dates change
    if updated and req.student_ids is not None:
        program_id = updated.get("program_id", "")
        student_ids = updated.get("student_ids") or []
        if student_ids:
            # Prefer date-based module determination
            module_number = get_current_module_from_dates(updated.get("module_dates") or {})
            if module_number is None:
                subject_ids_for_module = updated.get("subject_ids") or []
                if subject_ids_for_module:
                    subject_docs = await db.subjects.find(
                        {"id": {"$in": subject_ids_for_module}}, {"_id": 0, "module_number": 1}
                    ).to_list(None)
                    valid_modules = [s["module_number"] for s in subject_docs if s.get("module_number")]
                    if valid_modules:
                        module_number = min(valid_modules)
            if module_number is not None:
                await db.users.update_many(
                    {"id": {"$in": student_ids}, "role": "estudiante"},
                    {"$set": {
                        "module": module_number,
                        f"program_modules.{program_id}": module_number
                    }}
                )

    # If module dates were adjusted, immediately execute the same automatic
    # closure/recovery-close logic used by the daily scheduler. This prevents
    # students from remaining in groups until 02:00 AM when admins move dates
    # to already-past values (e.g., recovery_close already expired).
    if module_dates_updated:
        await check_and_close_modules()
    
    return updated

@api_router.delete("/courses/{course_id}")
async def delete_course(course_id: str, force: bool = False, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    
    # Find the course to get its student list and program
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    
    program_id = course.get("program_id", "")
    student_ids_in_course = course.get("student_ids", [])
    
    if student_ids_in_course:
        # Check which students are non-egresado
        students = await db.users.find(
            {"id": {"$in": student_ids_in_course}, "role": "estudiante"},
            {"_id": 0, "id": 1, "program_statuses": 1, "estado": 1}
        ).to_list(None)
        blocking_students = []
        for s in students:
            program_statuses = s.get("program_statuses") or {}
            status = program_statuses.get(program_id) if program_id else s.get("estado")
            if not status:
                status = s.get("estado", "activo")
            if status != "egresado":
                blocking_students.append(s["id"])
        if blocking_students and not force:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"No se puede eliminar el grupo: contiene {len(blocking_students)} estudiante(s) "
                    "que aún no han egresado. Use force=true para desmatricularlos y eliminar el grupo."
                )
            )
        # force=True or all students are egresados: unenroll non-egresado students (do not delete them)
        if blocking_students:
            await db.users.update_many(
                {"id": {"$in": blocking_students}},
                {"$unset": {"grupo": ""}}
            )
            logger.info(
                f"Force-deleted course {course_id}: unenrolled {len(blocking_students)} active student(s) "
                f"(students not deleted, only removed from group)"
            )
    
    # Collect file references before deleting (for disk/Cloudinary cleanup)
    activities_for_files = await db.activities.find(
        {"course_id": course_id}, {"_id": 0, "id": 1, "files": 1}
    ).to_list(None)
    activity_ids = [a["id"] for a in activities_for_files if a.get("id")]
    if activity_ids:
        submissions_for_files = await db.submissions.find(
            {"activity_id": {"$in": activity_ids}}, {"_id": 0, "files": 1}
        ).to_list(None)
    else:
        submissions_for_files = []

    # Delete the course (students are never deleted)
    await db.courses.delete_one({"id": course_id})

    # Delete all associated data
    activities_deleted = await db.activities.delete_many({"course_id": course_id})
    grades_deleted = await db.grades.delete_many({"course_id": course_id})
    if activity_ids:
        submissions_deleted = await db.submissions.delete_many({"activity_id": {"$in": activity_ids}})
    else:
        submissions_deleted = type("_DeleteResult", (), {"deleted_count": 0})()
    failed_subjects_deleted = await db.failed_subjects.delete_many({"course_id": course_id})
    recovery_enabled_deleted = await db.recovery_enabled.delete_many({"course_id": course_id})
    videos_deleted = await db.class_videos.delete_many({"course_id": course_id})

    # Clean up uploaded files from disk or Cloudinary
    cloudinary_deleted_count = 0
    for doc in activities_for_files + submissions_for_files:
        for f in (doc.get("files") or []):
            stored_name = f.get("stored_name") if isinstance(f, dict) else None
            if not stored_name:
                continue
            # Determine storage backend: use the 'storage' field set by the upload endpoint,
            # or detect Cloudinary by whether the public_id starts with the known upload folder.
            storage = f.get("storage") if isinstance(f, dict) else None
            is_cloudinary_file = (
                USE_CLOUDINARY and (
                    storage == "cloudinary" or
                    (isinstance(stored_name, str) and stored_name.startswith("educando/"))
                )
            )
            if is_cloudinary_file:
                try:
                    import cloudinary.uploader
                    _rt = f.get("resource_type") if isinstance(f, dict) else None
                    if not _rt:
                        _fext = Path(stored_name).suffix.lower().lstrip(".")
                        if _fext in {"jpg", "jpeg", "png", "gif", "webp", "bmp", "svg"}:
                            _rt = "image"
                        elif _fext in {"mp4", "avi", "mov", "mkv", "webm"}:
                            _rt = "video"
                        else:
                            _rt = "raw"
                    cloudinary.uploader.destroy(stored_name, resource_type=_rt)
                    cloudinary_deleted_count += 1
                except Exception as e:
                    logger.warning(f"Failed to delete Cloudinary file {stored_name}: {e}")
            else:
                file_path = UPLOAD_DIR / stored_name
                if file_path.exists():
                    try:
                        file_path.unlink()
                    except Exception as e:
                        logger.warning(f"Failed to delete file {stored_name}: {e}")

    logger.info(
        f"Course {course_id} deleted with associated data: "
        f"{activities_deleted.deleted_count} activities, {grades_deleted.deleted_count} grades, "
        f"{submissions_deleted.deleted_count} submissions, {failed_subjects_deleted.deleted_count} failed_subjects, "
        f"{recovery_enabled_deleted.deleted_count} recovery_enabled, {videos_deleted.deleted_count} videos, "
        f"{cloudinary_deleted_count} Cloudinary files removed"
    )
    await log_audit("course_deleted", user["id"], user["role"], {"course_id": course_id, "course_name": course.get("name", "")})

    return {
        "message": "Grupo eliminado con todos sus datos asociados",
        "deleted": {
            "activities": activities_deleted.deleted_count,
            "grades": grades_deleted.deleted_count,
            "submissions": submissions_deleted.deleted_count,
            "failed_subjects": failed_subjects_deleted.deleted_count,
            "recovery_enabled": recovery_enabled_deleted.deleted_count,
            "videos": videos_deleted.deleted_count
        }
    }

# --- Activities Routes ---
@api_router.get("/activities")
async def get_activities(course_id: Optional[str] = None, subject_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if course_id:
        query["course_id"] = course_id
    if subject_id:
        query["subject_id"] = subject_id
    activities = await db.activities.find(query, {"_id": 0}).to_list(500)
    
    # For students: show recovery activities only for subjects explicitly approved by admin.
    if user["role"] == "estudiante" and course_id:
        approved_records = await db.failed_subjects.find({
            "student_id": user["id"],
            "course_id": course_id,
            "recovery_approved": True,
            "recovery_processed": {"$ne": True},
            "recovery_rejected": {"$ne": True},
        }, {"_id": 0, "subject_id": 1}).to_list(None)
        approved_subject_ids = {r.get("subject_id") for r in approved_records if r.get("subject_id")}
        has_course_level_approval = any(not r.get("subject_id") for r in approved_records)

        filtered_activities = []
        for a in activities:
            if not a.get("is_recovery"):
                filtered_activities.append(a)
                continue
            act_sid = a.get("subject_id")
            if (act_sid and act_sid in approved_subject_ids) or (not act_sid and has_course_level_approval):
                filtered_activities.append(a)
        activities = filtered_activities
    
    return activities

@api_router.post("/activities")
async def create_activity(req: ActivityCreate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    # Validate: only ONE recovery activity per subject per course
    if req.is_recovery:
        recovery_query = {"course_id": req.course_id, "is_recovery": True}
        if req.subject_id:
            recovery_query["subject_id"] = req.subject_id
        existing_recovery = await db.activities.find_one(recovery_query)
        if existing_recovery:
            raise HTTPException(
                status_code=400,
                detail="Ya existe una actividad de recuperación para esta materia. Solo se permite una por materia."
            )
    # Auto-number: find the max activity_number for this subject within the course (per-subject numbering)
    activity_number_query = {"course_id": req.course_id}
    if req.subject_id:
        activity_number_query["subject_id"] = req.subject_id
    agg_result = await db.activities.aggregate([
        {"$match": activity_number_query},
        {"$group": {"_id": None, "max_num": {"$max": "$activity_number"}}}
    ]).to_list(1)
    max_num = agg_result[0]["max_num"] if agg_result else 0
    activity_number = (max_num or 0) + 1
    activity = {
        "id": str(uuid.uuid4()),
        "course_id": req.course_id,
        "subject_id": req.subject_id,
        "activity_number": activity_number,
        "title": req.title,
        "description": req.description,
        "start_date": req.start_date,
        "due_date": req.due_date,
        "files": req.files,
        "is_recovery": req.is_recovery or False,
        "active": True,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.activities.insert_one(activity)
    del activity["_id"]
    await log_audit("activity_created", user["id"], user["role"], {"activity_id": activity["id"], "course_id": req.course_id, "title": req.title})
    return activity

@api_router.put("/activities/{activity_id}")
async def update_activity(activity_id: str, req: ActivityUpdate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    result = await db.activities.update_one({"id": activity_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    updated = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    return updated

@api_router.delete("/activities/{activity_id}")
async def delete_activity(activity_id: str, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    
    # Cascade delete: remove related grades and submissions
    await db.grades.delete_many({"activity_id": activity_id})
    await db.submissions.delete_many({"activity_id": activity_id})
    await db.activities.delete_one({"id": activity_id})
    await log_audit("activity_deleted", user["id"], user["role"], {"activity_id": activity_id})
    return {"message": "Actividad eliminada con sus notas y entregas"}

# --- Grades Routes ---
@api_router.get("/grades")
async def get_grades(course_id: Optional[str] = None, student_id: Optional[str] = None, subject_id: Optional[str] = None, activity_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if course_id:
        query["course_id"] = course_id
    if student_id:
        query["student_id"] = student_id
    if subject_id:
        query["subject_id"] = subject_id
    if activity_id:
        query["activity_id"] = activity_id
    grades = await db.grades.find(query, {"_id": 0}).to_list(50000)
    return grades

@api_router.post("/grades")
async def create_grade(req: GradeCreate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    
    # Rule E: For recovery activities, only approve/reject is allowed — no numeric grade
    if req.activity_id:
        activity_doc = await db.activities.find_one({"id": req.activity_id}, {"_id": 0, "is_recovery": 1, "course_id": 1, "subject_id": 1})
        if activity_doc and activity_doc.get("is_recovery"):
            if req.value is not None and not req.recovery_status:
                raise HTTPException(
                    status_code=400,
                    detail="Las actividades de recuperación no admiten nota numérica. Use Aprobar o Rechazar."
                )
            if req.recovery_status not in ("approved", "rejected", None):
                raise HTTPException(status_code=400, detail="Estado de recuperación inválido")
            # Recovery must be admin-approved before teacher can grade it
            rec_subject_id = req.subject_id or activity_doc.get("subject_id")
            failed_filter = {
                "student_id": req.student_id,
                "course_id": req.course_id,
                "recovery_approved": True,
                "recovery_processed": {"$ne": True},
                "recovery_rejected": {"$ne": True},
            }
            if rec_subject_id:
                failed_filter["subject_id"] = rec_subject_id
            failed_record = await db.failed_subjects.find_one(failed_filter)
            if not failed_record:
                raise HTTPException(
                    status_code=400,
                    detail="La recuperación debe ser aprobada por el administrador antes de poder calificar"
                )
    
    existing = await db.grades.find_one({
        "student_id": req.student_id,
        "course_id": req.course_id,
        "activity_id": req.activity_id
    })
    
    # If this is a recovery grading, calculate the grade based on approval
    grade_value = req.value
    if req.recovery_status:
        # Build the filter to find the specific failed_subjects record
        rec_subject_id = req.subject_id
        if req.activity_id:
            _act = await db.activities.find_one({"id": req.activity_id}, {"_id": 0, "subject_id": 1})
            rec_subject_id = rec_subject_id or (_act or {}).get("subject_id")

        fs_filter = {
            "student_id": req.student_id,
            "course_id": req.course_id,
            "recovery_approved": True,
            "recovery_processed": {"$ne": True},
            "recovery_rejected": {"$ne": True},
        }
        if rec_subject_id:
            fs_filter["subject_id"] = rec_subject_id

        if req.recovery_status == "approved":
            # Calculate what grade is needed to get exactly 3.0 average
            other_grades = await db.grades.find({
                "student_id": req.student_id,
                "course_id": req.course_id,
                "activity_id": {"$ne": req.activity_id}
            }, {"_id": 0}).to_list(100)
            
            if other_grades:
                total = sum(g["value"] for g in other_grades)
                count = len(other_grades)
                # To get average of 3.0: (total + x) / (count + 1) = 3.0
                # x = 3.0 * (count + 1) - total
                grade_value = 3.0 * (count + 1) - total
                # Ensure it's between 0 and 5
                grade_value = max(0.0, min(5.0, grade_value))
            else:
                grade_value = 3.0
            
            # Mark ONLY this specific failed_subjects record as completed+approved.
            # Promotion/graduation (module advancement) is deferred to check_and_close_modules
            # after recovery_close date, but the student's estado is updated immediately if all
            # their recoveries are now approved (see _check_and_update_recovery_completion).
            await db.failed_subjects.update_one(
                fs_filter,
                {"$set": {
                    "recovery_completed": True,
                    "teacher_graded_status": "approved",
                    "completed_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            # Immediately exit pendiente_recuperacion if all recovery subjects are now approved
            await _check_and_update_recovery_completion(req.student_id, req.course_id)
        else:
            # Recovery rejected by teacher: mark the record as completed+rejected.
            await db.failed_subjects.update_one(
                fs_filter,
                {"$set": {
                    "recovery_completed": True,
                    "teacher_graded_status": "rejected",
                    "completed_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            # When a teacher rejects, check if ALL admin-approved subjects for this
            # student+course have now been teacher-graded. If so and at least one is
            # rejected, the student definitively cannot pass: remove from group and
            # mark as reprobado immediately (don't wait for recovery_close).
            await _check_and_update_recovery_rejection(req.student_id, req.course_id)
            
            # If rejected, don't create/update a numeric grade (keep existing average)
            # Just update the recovery status; if no grade exists, persist a status-only
            # row so teacher UIs (Notas/Entregas) can display "Rechazado" consistently.
            if existing:
                await db.grades.update_one(
                    {"id": existing["id"]},
                    {"$set": {"recovery_status": req.recovery_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                updated = await db.grades.find_one({"id": existing["id"]}, {"_id": 0})
                await log_audit("recovery_graded", user["id"], user["role"], {"student_id": req.student_id, "course_id": req.course_id, "subject_id": req.subject_id, "result": "rejected"})
                return updated
            # Create a status-only grade record (value=None) to expose rejected state in UI
            grade = {
                "id": str(uuid.uuid4()),
                "student_id": req.student_id,
                "course_id": req.course_id,
                "activity_id": req.activity_id,
                "subject_id": req.subject_id,
                "value": None,
                "comments": req.comments,
                "recovery_status": req.recovery_status,
                "graded_by": user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.grades.insert_one(grade)
            await log_audit("recovery_graded", user["id"], user["role"], {"student_id": req.student_id, "course_id": req.course_id, "subject_id": req.subject_id, "result": "rejected"})
            grade.pop("_id", None)
            return grade
    
    if existing:
        update_data = {
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        if grade_value is not None:
            update_data["value"] = grade_value
        if req.comments:
            update_data["comments"] = req.comments
        if req.recovery_status:
            update_data["recovery_status"] = req.recovery_status
            
        await db.grades.update_one(
            {"id": existing["id"]},
            {"$set": update_data}
        )
        updated = await db.grades.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    
    grade = {
        "id": str(uuid.uuid4()),
        "student_id": req.student_id,
        "course_id": req.course_id,
        "activity_id": req.activity_id,
        "subject_id": req.subject_id,
        "value": grade_value if grade_value is not None else 0.0,
        "comments": req.comments,
        "recovery_status": req.recovery_status,
        "graded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.grades.insert_one(grade)
    del grade["_id"]
    if req.recovery_status == "approved":
        await log_audit("recovery_graded", user["id"], user["role"], {"student_id": req.student_id, "course_id": req.course_id, "subject_id": req.subject_id, "result": "approved"})
    else:
        await log_audit("grade_assigned", user["id"], user["role"], {"student_id": req.student_id, "course_id": req.course_id, "subject_id": req.subject_id, "activity_id": req.activity_id})
    return grade

@api_router.put("/grades/{grade_id}")
async def update_grade(grade_id: str, req: GradeUpdate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.grades.update_one({"id": grade_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nota no encontrada")
    updated = await db.grades.find_one({"id": grade_id}, {"_id": 0})
    return updated

# --- Class Videos Routes ---
@api_router.get("/class-videos")
async def get_class_videos(course_id: Optional[str] = None, subject_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if course_id:
        query["course_id"] = course_id
    if subject_id:
        query["subject_id"] = subject_id
    # Students only see videos whose available_from has already passed (or is not set)
    if user["role"] == "estudiante":
        now_iso = datetime.now(timezone.utc).isoformat()
        query["$or"] = [
            {"available_from": {"$exists": False}},
            {"available_from": None},
            {"available_from": ""},
            {"available_from": {"$lte": now_iso}}
        ]
    videos = await db.class_videos.find(query, {"_id": 0}).to_list(500)
    return videos

@api_router.post("/class-videos")
async def create_class_video(req: ClassVideoCreate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    video = {
        "id": str(uuid.uuid4()),
        "course_id": req.course_id,
        "subject_id": req.subject_id,
        "title": req.title,
        "url": req.url,
        "description": req.description,
        "available_from": req.available_from or None,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.class_videos.insert_one(video)
    del video["_id"]
    return video

@api_router.delete("/class-videos/{video_id}")
async def delete_class_video(video_id: str, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    await db.class_videos.delete_one({"id": video_id})
    return {"message": "Video eliminado"}

class ClassVideoUpdate(BaseModel):
    title: Optional[str] = None
    url: Optional[str] = None
    description: Optional[str] = None
    available_from: Optional[str] = None  # ISO datetime; empty string clears the restriction

@api_router.put("/class-videos/{video_id}")
async def update_class_video(video_id: str, req: ClassVideoUpdate, user=Depends(get_current_user)):
    if user["role"] != "profesor":
        raise HTTPException(status_code=403, detail="Solo profesores")
    # Build update dict; allow explicit None/empty for available_from to clear it
    raw = req.model_dump()
    update_data = {k: v for k, v in raw.items() if v is not None}
    # Explicitly handle available_from="" or available_from=None to clear the restriction
    if raw.get("available_from") in ("", None) and "available_from" in raw:
        update_data["available_from"] = None
    result = await db.class_videos.update_one({"id": video_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Video no encontrado")
    updated = await db.class_videos.find_one({"id": video_id}, {"_id": 0})
    return updated

# --- File Upload Route ---
@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["role"] not in ["profesor", "admin", "estudiante"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    original_name = file.filename
    file_content = await file.read()
    file_size = len(file_content)

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="El archivo excede el tamaño máximo permitido (10MB)")

    import re as _re
    _ext = Path(original_name).suffix.lower().lstrip(".")

    ALLOWED_EXTENSIONS = {"pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx", "txt", "jpg", "jpeg", "png", "gif", "webp"}
    if _ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Tipo de archivo no permitido: .{_ext}")

    _safe_basename = _re.sub(r'[^\w\-]', '_', Path(original_name).stem)[:100]

    # Rate limit: max 20 uploads per minute per user
    _user_id = user["id"]
    _now = datetime.now().timestamp()
    upload_attempts[_user_id] = [t for t in upload_attempts[_user_id] if _now - t < UPLOAD_WINDOW]
    if len(upload_attempts[_user_id]) >= MAX_UPLOADS_PER_MINUTE:
        raise HTTPException(status_code=429, detail="Demasiadas subidas. Intente de nuevo en un minuto.")
    upload_attempts[_user_id].append(_now)

    _unique_suffix = str(uuid.uuid4())[:8]

    if USE_S3:
        import boto3

        # Content-type correcto por extensión
        _content_type_map = {
            "pdf":  "application/pdf",
            "doc":  "application/msword",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "xls":  "application/vnd.ms-excel",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "ppt":  "application/vnd.ms-powerpoint",
            "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "txt":  "text/plain",
            "jpg":  "image/jpeg",
            "jpeg": "image/jpeg",
            "png":  "image/png",
            "gif":  "image/gif",
            "webp": "image/webp",
        }
        _content_type = _content_type_map.get(_ext, "application/octet-stream")

        # Subcarpeta por tipo de archivo dentro del bucket
        if _ext == "pdf":
            _folder = "uploads/pdf"
        elif _ext in {"jpg", "jpeg", "png", "gif", "webp"}:
            _folder = "uploads/images"
        else:
            _folder = "uploads/docs"

        _s3_key = f"{_folder}/{_safe_basename}_{_unique_suffix}.{_ext}"

        try:
            s3_client = boto3.client(
                's3',
                aws_access_key_id=AWS_ACCESS_KEY_ID,
                aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                region_name=AWS_S3_REGION
            )
            logger.info(f"Attempting S3 upload: bucket={AWS_S3_BUCKET_NAME}, region={AWS_S3_REGION}, key={_s3_key}")
            s3_client.put_object(
                Bucket=AWS_S3_BUCKET_NAME,
                Key=_s3_key,
                Body=file_content,
                ContentType=_content_type,
                ContentDisposition=f'inline; filename="{original_name}"',
            )
            _file_url = f"https://{AWS_S3_BUCKET_NAME}.s3.{AWS_S3_REGION}.amazonaws.com/{_s3_key}"
            logger.info(f"File uploaded to S3: {_file_url}")
            return {
                "filename": original_name,
                "stored_name": _s3_key,
                "url": _file_url,
                "size": file_size,
                "storage": "s3",
                "resource_type": "raw"
            }
        except Exception as e:
            logger.error(f"S3 upload failed: {e}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Error uploading file to S3: {str(e)}. Please check S3 configuration."
            )

    # Fallback: use Cloudinary if configured
    if USE_CLOUDINARY:
        import cloudinary.uploader
        import io as _io
        if _ext in {"jpg", "jpeg", "png", "gif", "webp", "bmp", "svg"}:
            _resource_type = "image"
        elif _ext in {"mp4", "avi", "mov", "mkv", "webm"}:
            _resource_type = "video"
        else:
            _resource_type = "raw"
        _public_id = f"educando/uploads/{_safe_basename}_{_unique_suffix}.{_ext}"
        result = cloudinary.uploader.upload(
            _io.BytesIO(file_content),
            public_id=_public_id,
            resource_type=_resource_type,
            overwrite=False,
        )
        return {
            "filename": original_name,
            "stored_name": result["public_id"],
            "url": result["secure_url"],
            "size": file_size,
            "storage": "cloudinary",
            "resource_type": _resource_type
        }

    # Fallback: local disk
    _file_id = str(uuid.uuid4())
    _ext_with_dot = Path(original_name).suffix
    safe_name = f"{_file_id}{_ext_with_dot}"
    file_path = UPLOAD_DIR / safe_name
    with open(file_path, "wb") as f:
        f.write(file_content)
    return {
        "filename": original_name,
        "stored_name": safe_name,
        "url": f"/api/files/{safe_name}",
        "size": os.path.getsize(file_path),
        "storage": "local"
    }

@api_router.get("/files/{filename}")
async def get_file(filename: str):
    safe_filename = Path(filename).name
    file_path = UPLOAD_DIR / safe_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    ext = safe_filename.rsplit('.', 1)[-1].lower() if '.' in safe_filename else ''
    mime_map = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
        'png': 'image/png', 'gif': 'image/gif', 'webp': 'image/webp',
        'txt': 'text/plain',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'xls': 'application/vnd.ms-excel',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'ppt': 'application/vnd.ms-powerpoint',
        'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    }
    media_type = mime_map.get(ext, 'application/octet-stream')
    if ext in ('pdf', 'jpg', 'jpeg', 'png', 'gif', 'webp', 'txt'):
        disposition = f'inline; filename="{safe_filename}"'
    else:
        disposition = f'attachment; filename="{safe_filename}"'
    from starlette.responses import FileResponse
    return FileResponse(
        file_path,
        media_type=media_type,
        headers={"Content-Disposition": disposition}
    )

# --- Submissions Routes ---
@api_router.get("/submissions")
async def get_submissions(activity_id: Optional[str] = None, student_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if activity_id:
        query["activity_id"] = activity_id
    if student_id:
        query["student_id"] = student_id
    submissions = await db.submissions.find(query, {"_id": 0}).to_list(50000)
    return submissions

@api_router.post("/submissions")
async def create_submission(req: SubmissionCreate, user=Depends(get_current_user)):
    if user["role"] != "estudiante":
        raise HTTPException(status_code=403, detail="Solo estudiantes")
    
    activity = await db.activities.find_one({"id": req.activity_id}, {"_id": 0})
    if not activity:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    now = datetime.now(timezone.utc)
    
    # Check admin approval for recovery activities (subject-scoped)
    if activity.get("is_recovery"):
        failed_filter = {
            "student_id": user["id"],
            "course_id": activity["course_id"],
            "recovery_approved": True,
            "recovery_processed": {"$ne": True},
            "recovery_rejected": {"$ne": True},
        }
        if activity.get("subject_id"):
            failed_filter["subject_id"] = activity["subject_id"]
        failed_record = await db.failed_subjects.find_one(failed_filter)
        if not failed_record:
            raise HTTPException(
                status_code=400,
                detail="Tu recuperación debe ser aprobada por el administrador antes de poder entregar actividades"
            )
    
    # Check start_date
    if activity.get("start_date"):
        start_date = datetime.fromisoformat(activity["start_date"].replace("Z", "+00:00"))
        if now < start_date:
            raise HTTPException(status_code=400, detail="La actividad aún no está disponible.")
    
    due_date = datetime.fromisoformat(activity["due_date"].replace("Z", "+00:00"))
    if now > due_date:
        raise HTTPException(status_code=400, detail="La fecha límite ha pasado. No se puede entregar.")
    
    # Check module restriction: activity's subject module must match student's current module
    if activity.get("subject_id"):
        subject = await db.subjects.find_one({"id": activity["subject_id"]}, {"_id": 0, "module_number": 1})
        if subject and subject.get("module_number") is not None:
            subject_module = subject["module_number"]
            # Get course's program_id
            course = await db.courses.find_one({"id": activity["course_id"]}, {"_id": 0, "program_id": 1})
            if course and course.get("program_id"):
                program_id = course["program_id"]
                student_module = (user.get("program_modules") or {}).get(program_id)
                if student_module is None:
                    student_module = user.get("module")
                if student_module is not None and int(student_module) != int(subject_module):
                    raise HTTPException(
                        status_code=403,
                        detail=f"Esta actividad pertenece al Módulo {int(subject_module)}, pero estás actualmente en el Módulo {int(student_module)}. Solo puedes entregar actividades de tu módulo actual."
                    )
    
    existing = await db.submissions.find_one({
        "activity_id": req.activity_id,
        "student_id": user["id"]
    })
    if existing:
        # Check if already edited once
        if existing.get("edited", False):
            raise HTTPException(status_code=400, detail="Esta actividad ya ha sido editada. Solo se permite una edición por actividad.")
        
        # Allow editing, mark as edited
        await db.submissions.update_one(
            {"id": existing["id"]},
            {"$set": {"content": req.content, "files": req.files, "submitted_at": datetime.now(timezone.utc).isoformat(), "edited": True}}
        )
        updated = await db.submissions.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    
    submission = {
        "id": str(uuid.uuid4()),
        "activity_id": req.activity_id,
        "student_id": user["id"],
        "content": req.content,
        "files": req.files,
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "edited": False
    }
    await db.submissions.insert_one(submission)
    del submission["_id"]
    return submission

# --- Recovery Management Routes ---
@api_router.post("/recovery/enable")
async def enable_recovery(req: RecoveryEnableRequest, user=Depends(get_current_user)):
    """Admin enables recovery for a specific student in a course"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede habilitar recuperaciones")
    
    # Create or update recovery enablement record
    recovery = {
        "id": str(uuid.uuid4()),
        "student_id": req.student_id,
        "course_id": req.course_id,
        "subject_id": req.subject_id,
        "enabled": True,
        "enabled_by": user["id"],
        "enabled_at": datetime.now(timezone.utc).isoformat()
    }
    
    existing = await db.recovery_enabled.find_one({
        "student_id": req.student_id,
        "course_id": req.course_id,
        "subject_id": req.subject_id
    })
    
    if existing:
        await db.recovery_enabled.update_one(
            {"id": existing["id"]},
            {"$set": {"enabled": True, "enabled_by": user["id"], "enabled_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Recuperación actualizada"}
    
    await db.recovery_enabled.insert_one(recovery)
    return {"message": "Recuperación habilitada para el estudiante"}

@api_router.get("/recovery/enabled")
async def get_recovery_enabled(student_id: Optional[str] = None, course_id: Optional[str] = None, user=Depends(get_current_user)):
    """Get list of students with recovery enabled"""
    # Source of truth: currently approved failed_subjects that are still in process.
    # This prevents old approvals from previous modules from auto-enabling new recoveries.
    fs_query = {
        "recovery_approved": True,
        "recovery_completed": {"$ne": True},
        "recovery_processed": {"$ne": True},
        "recovery_rejected": {"$ne": True},
    }
    if student_id:
        fs_query["student_id"] = student_id
    if course_id:
        fs_query["course_id"] = course_id

    active_records = await db.failed_subjects.find(fs_query, {"_id": 0, "student_id": 1, "course_id": 1, "subject_id": 1}).to_list(1000)
    if active_records:
        seen = set()
        enabled = []
        for r in active_records:
            key = (r.get("student_id"), r.get("course_id"), r.get("subject_id"))
            if key in seen:
                continue
            seen.add(key)
            enabled.append({
                "student_id": r.get("student_id"),
                "course_id": r.get("course_id"),
                "subject_id": r.get("subject_id"),
                "enabled": True,
                "source": "failed_subjects"
            })
        return enabled

    # Backward compatibility fallback
    query = {}
    if student_id:
        query["student_id"] = student_id
    if course_id:
        query["course_id"] = course_id
    return await db.recovery_enabled.find(query, {"_id": 0}).to_list(500)

@api_router.put("/users/{user_id}/promote")
async def promote_student(user_id: str, program_id: Optional[str] = None, user=Depends(get_current_user)):
    """Admin promotes a student to the next module.

    Supports N modules (no hardcoded upper bound).  When program_id is supplied
    only that program's module is advanced; otherwise every program the student
    belongs to that is currently at ``current_module`` is advanced.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede promover estudiantes")
    
    student = await db.users.find_one({"id": user_id, "role": "estudiante"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    program_modules = student.get("program_modules") or {}
    program_statuses = student.get("program_statuses") or {}
    current_module = student.get("module", 1)

    # Determine target programs
    if program_id:
        all_prog_ids = student.get("program_ids") or ([student.get("program_id")] if student.get("program_id") else [])
        if program_id not in all_prog_ids:
            raise HTTPException(status_code=400, detail=f"El estudiante no pertenece al programa {program_id}")
        target_programs = [program_id]
    else:
        target_programs = student.get("program_ids") or ([student.get("program_id")] if student.get("program_id") else [])

    if not target_programs:
        # Legacy fallback: no per-program data; just bump the global module field
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"module": current_module + 1}}
        )
        updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        return updated

    # Load max_modules per program for N-module support
    prog_docs = await db.programs.find(
        {"id": {"$in": target_programs}}, {"_id": 0, "id": 1, "modules": 1}
    ).to_list(None)
    prog_max_map = {p["id"]: max(len(p.get("modules") or []), 2) for p in prog_docs}

    set_fields: dict = {}
    for pid in target_programs:
        prog_current = program_modules.get(pid, current_module)
        prog_max = prog_max_map.get(pid, 2)
        if prog_current >= prog_max:
            raise HTTPException(
                status_code=400,
                detail=f"El estudiante ya está en el módulo final del programa {pid}"
            )
        next_mod = prog_current + 1
        program_modules[pid] = next_mod
        program_statuses[pid] = "activo"
        set_fields[f"program_modules.{pid}"] = next_mod

    new_estado = derive_estado_from_program_statuses(program_statuses)
    set_fields["program_statuses"] = program_statuses
    set_fields["estado"] = new_estado
    # Keep global module field in sync with the first target program
    set_fields["module"] = program_modules.get(target_programs[0], current_module)

    await db.users.update_one({"id": user_id}, {"$set": set_fields})
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated

@api_router.put("/users/{user_id}/graduate")
async def graduate_student(user_id: str, program_id: Optional[str] = None, user=Depends(get_current_user)):
    """Admin marks a student as graduated (egresado).

    Uses dynamic max_modules from the program definition (N-module support).
    Updates both ``program_statuses`` and the derived global ``estado`` so that
    multi-program students are handled correctly.
    When program_id is supplied only that program is graduated; otherwise every
    program the student belongs to is graduated.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede graduar estudiantes")
    
    student = await db.users.find_one({"id": user_id, "role": "estudiante"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    program_modules = student.get("program_modules") or {}
    program_statuses = student.get("program_statuses") or {}
    current_module = student.get("module", 1)

    # Determine target programs
    if program_id:
        all_prog_ids = student.get("program_ids") or ([student.get("program_id")] if student.get("program_id") else [])
        if program_id not in all_prog_ids:
            raise HTTPException(status_code=400, detail=f"El estudiante no pertenece al programa {program_id}")
        target_programs = [program_id]
    else:
        target_programs = student.get("program_ids") or ([student.get("program_id")] if student.get("program_id") else [])

    if not target_programs:
        # Legacy fallback: no per-program data; check global module field
        if current_module < 2:
            raise HTTPException(status_code=400, detail="El estudiante debe estar en el módulo final para graduarse")
        await db.users.update_one({"id": user_id}, {"$set": {"estado": "egresado"}})
        updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
        return updated

    # Load max_modules per program for dynamic last-module check (N-module support)
    prog_docs = await db.programs.find(
        {"id": {"$in": target_programs}}, {"_id": 0, "id": 1, "modules": 1}
    ).to_list(None)
    prog_max_map = {p["id"]: max(len(p.get("modules") or []), 2) for p in prog_docs}

    for pid in target_programs:
        prog_current = program_modules.get(pid, current_module)
        prog_max = prog_max_map.get(pid, 2)
        if prog_current < prog_max:
            raise HTTPException(
                status_code=400,
                detail=f"El estudiante debe estar en el módulo final del programa {pid} para graduarse"
            )
        program_statuses[pid] = "egresado"

    new_estado = derive_estado_from_program_statuses(program_statuses)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"estado": new_estado, "program_statuses": program_statuses}}
    )
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated

@api_router.post("/admin/set-all-students-module-1")
async def set_all_students_module_1(user=Depends(get_current_user)):
    """Admin endpoint to set all existing students to Module 1"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede realizar esta operación")
    
    # Update all students to module 1
    result = await db.users.update_many(
        {"role": "estudiante"},
        {"$set": {"module": 1}}
    )
    
    return {
        "message": f"Se actualizaron {result.modified_count} estudiantes al Módulo 1",
        "modified_count": result.modified_count
    }

async def close_module_internal(module_number: int, program_id: Optional[str] = None):
    """
    Internal function to close a module for a program. 
    Reviews all student grades and:
    - If student passed all subjects: promote to next module or graduate
    - If student failed any subject: mark as 'pendiente_recuperacion'
    
    This function can be called both by the API endpoint and the automatic scheduler.
    """
    # Business rule (2026-02-25): no longer restricting to modules 1-2; supports N modules.
    if not isinstance(module_number, int) or module_number < 1:
        raise ValueError(f"Número de módulo inválido: debe ser un entero >= 1, got {module_number}")
    
    # Get students enrolled in this program who are currently activo for that program.
    # Business rule (2026-02-25): filter by per-program status (not global estado) so that
    # multi-program students in pendiente_recuperacion for another program are not incorrectly
    # included in this module's closure.
    # Dual-filter rationale: program_statuses.{program_id}=="activo" ensures per-program
    # correctness; the $or on program_id/program_ids restricts to students enrolled in the
    # target program (supports both legacy single-program and multi-program data shapes).
    query: dict = {"role": "estudiante"}
    if program_id:
        query[f"program_statuses.{program_id}"] = "activo"
        query["$or"] = [
            {"program_id": program_id},
            {"program_ids": program_id}
        ]
    else:
        query["estado"] = "activo"
    
    students = await db.users.find(query, {"_id": 0}).to_list(5000)
    
    promoted_count = 0
    graduated_count = 0
    recovery_count = 0
    failed_subjects_records = []
    
    # Get all courses/groups
    courses = await db.courses.find({}, {"_id": 0}).to_list(5000)
    
    # Load subjects for per-subject grade calculations
    all_subjects = await db.subjects.find({}, {"_id": 0, "id": 1, "name": 1, "module_number": 1}).to_list(1000)
    subject_name_map = {s["id"]: s["name"] for s in all_subjects}
    subject_module_map = {s["id"]: (s.get("module_number") or 1) for s in all_subjects}
    
    # Pre-load programs for N-module max_modules determination
    all_programs = await db.programs.find({}, {"_id": 0, "id": 1, "modules": 1}).to_list(100)
    program_max_modules_map = {
        p["id"]: max(len(p.get("modules") or []), 2)
        for p in all_programs
    }
    
    # Cargar TODAS las grades de todos los cursos en UNA sola query (optimización crítica)
    all_course_ids = [c["id"] for c in courses]
    all_grades_bulk = await db.grades.find(
        {"course_id": {"$in": all_course_ids}}, {"_id": 0}
    ).to_list(None)
    # Indexar por student_id para acceso O(1) en el loop
    grades_by_student = {}
    for g in all_grades_bulk:
        sid = g.get("student_id")
        if sid:
            grades_by_student.setdefault(sid, []).append(g)
    
    for student in students:
        student_id = student["id"]
        student_program_modules = student.get("program_modules", {})
        
        # Check if student is in the specified module for any of their programs
        programs_in_module = []
        if program_id:
            # Check specific program
            if student_program_modules.get(program_id) == module_number:
                programs_in_module.append(program_id)
        else:
            # Check all programs
            for prog_id, mod_num in student_program_modules.items():
                if mod_num == module_number:
                    programs_in_module.append(prog_id)
        
        if not programs_in_module:
            continue  # Student not in this module
        
        # Get all courses for this student in the specified programs
        student_courses = [c for c in courses if student_id in c.get("student_ids", []) and c["program_id"] in programs_in_module]
        
        # Get all grades for the student
        all_grades = grades_by_student.get(student_id, [])
        
        # Group grades by course; detect failing subjects individually
        failed_courses = []
        for course in student_courses:
            course_grades = [g for g in all_grades if g.get("course_id") == course["id"]]
            
            subject_ids = course.get("subject_ids") or []
            if not subject_ids and course.get("subject_id"):
                subject_ids = [course["subject_id"]]
            
            course_has_failing = False
            if subject_ids:
                # Create one record per failing subject (only subjects belonging to the module being closed)
                for subject_id in subject_ids:
                    # Only process subjects that belong to the module being closed.
                    # Default to module 1 for subjects without module_number, so they're
                    # included in module 1 closures but excluded from module 2+ closures.
                    if subject_module_map.get(subject_id, 1) != module_number:
                        continue
                    subject_grades = [g["value"] for g in course_grades if g.get("subject_id") == subject_id and g.get("value") is not None]
                    subject_avg = sum(subject_grades) / len(subject_grades) if subject_grades else 0.0
                    
                    if subject_avg < 3.0:
                        course_has_failing = True
                        subject_name = subject_name_map.get(subject_id, "Desconocido")
                        failed_subjects_records.append({
                            "id": str(uuid.uuid4()),
                            "student_id": student_id,
                            "student_name": student["name"],
                            "course_id": course["id"],
                            "course_name": course["name"],
                            "subject_id": subject_id,
                            "subject_name": subject_name,
                            "program_id": course["program_id"],
                            "module_number": module_number,
                            "average_grade": round(subject_avg, 2),
                            "recovery_approved": False,
                            "recovery_completed": False,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
            else:
                # Fallback: no subjects defined, use course-level average
                grade_values = [g["value"] for g in course_grades if g.get("value") is not None]
                average = sum(grade_values) / len(grade_values) if grade_values else 0.0
                
                if average < 3.0:
                    course_has_failing = True
                    failed_subjects_records.append({
                        "id": str(uuid.uuid4()),
                        "student_id": student_id,
                        "student_name": student["name"],
                        "course_id": course["id"],
                        "course_name": course["name"],
                        "program_id": course["program_id"],
                        "module_number": module_number,
                        "average_grade": round(average, 2),
                        "recovery_approved": False,
                        "recovery_completed": False,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    })
            
            if course_has_failing:
                failed_courses.append({
                    "course_id": course["id"],
                    "course_name": course["name"],
                    "program_id": course["program_id"],
                })
        
        # Determine student status
        if failed_courses:
            # Student failed some subjects - mark as pending recovery per-program
            recovery_count += 1
            
            # Collect programs where student failed
            failed_program_ids = list({f["program_id"] for f in failed_courses})
            
            # Update program_statuses per-program and derive global estado
            student_program_statuses = student.get("program_statuses") or {}
            for prog_id in failed_program_ids:
                student_program_statuses[prog_id] = "pendiente_recuperacion"
            new_global_estado = derive_estado_from_program_statuses(student_program_statuses)
            await db.users.update_one(
                {"id": student_id},
                {"$set": {
                    "program_statuses": student_program_statuses,
                    "estado": new_global_estado
                }}
            )
        else:
            # Student passed all courses.
            # If any of the student's courses in this module has a recovery_close date
            # configured, defer the promotion/graduation to recovery_close processing.
            # This ensures students advance exactly when the recovery period closes,
            # not at the module close date.
            any_recovery_close = any(
                (c.get("module_dates") or {}).get(str(module_number), {}).get("recovery_close")
                for c in student_courses
            )
            if any_recovery_close:
                # Promotion deferred; the scheduler's recovery_close section will handle it.
                # Business rule (2026-02-25): set program_promotion_pending so the frontend
                # can show "Aprobado (pend. avance)" and the field is cleared at recovery_close.
                pending_set = {f"program_promotion_pending.{p}": True for p in programs_in_module}
                if pending_set:
                    await db.users.update_one({"id": student_id}, {"$set": pending_set})
                promoted_count += 1
            else:
                # No recovery period configured: promote immediately (backward compatibility).
                student_program_statuses = student.get("program_statuses") or {}
                for prog_id in programs_in_module:
                    current_module = student_program_modules.get(prog_id, 1)
                    prog_max_modules = program_max_modules_map.get(prog_id, 2)

                    if current_module < prog_max_modules:
                        # Promote to next module
                        student_program_modules[prog_id] = current_module + 1
                        student_program_statuses[prog_id] = "activo"
                        promoted_count += 1
                    else:
                        # Last module completed - graduate
                        graduated_count += 1
                        student_program_statuses[prog_id] = "egresado"
                
                new_global_estado = derive_estado_from_program_statuses(student_program_statuses)
                # Update program_modules and program_statuses
                await db.users.update_one(
                    {"id": student_id},
                    {"$set": {
                        "program_modules": student_program_modules,
                        "program_statuses": student_program_statuses,
                        "estado": new_global_estado
                    }}
                )
    
    # Bulk insert failed subjects records
    if failed_subjects_records:
        await db.failed_subjects.insert_many(failed_subjects_records)
    
    return {
        "message": "Cierre de módulo completado",
        "module_number": module_number,
        "program_id": program_id,
        "promoted_count": promoted_count,
        "graduated_count": graduated_count,
        "recovery_pending_count": recovery_count,
        "failed_subjects_count": len(failed_subjects_records)
    }

@api_router.post("/admin/close-module")
async def close_module(module_number: int, program_id: Optional[str] = None, user=Depends(get_current_user)):
    """
    Admin manually closes a module for a program. 
    Reviews all student grades and:
    - If student passed all subjects: promote to next module or graduate
    - If student failed any subject: mark as 'pendiente_recuperacion'
    
    NOTE: This is a manual endpoint. Module closure also happens automatically
    when the configured close date is reached.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede cerrar módulos")
    
    try:
        result = await close_module_internal(module_number, program_id)
        await log_audit("module_closed", user["id"], user["role"], {"module_number": module_number, "program_id": program_id or "all", "promoted": result.get("promoted_count", 0), "graduated": result.get("graduated_count", 0), "recovery": result.get("recovery_count", 0)})
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/admin/force-recovery-check")
async def force_recovery_check(user=Depends(get_current_user)):
    """
    Manually trigger the automatic recovery-close check and expulsion logic
    (same logic that the daily scheduler runs at 02:00 AM).

    Use this endpoint when the scheduler may not have run at the expected time
    or when you need to immediately apply expulsions after recovery deadlines have passed.
    Only accessible by admin users.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede forzar la verificación de recuperaciones")
    try:
        await check_and_close_modules()
        await log_audit(
            "force_recovery_check", user["id"], user["role"],
            {"trigger": "manual_admin"}
        )
        return {"message": "Verificación de recuperaciones ejecutada exitosamente"}
    except Exception as e:
        logger.error(f"Error in force_recovery_check endpoint: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error ejecutando verificación: {str(e)}")

@api_router.get("/admin/recovery-panel")
async def get_recovery_panel(user=Depends(get_current_user)):
    """
    Get all students with failed subjects pending recovery approval.
    Returns detailed information for admin to review and approve recoveries.
    Only shows in-process records: not yet decided, or approved-but-not-completed.
    Excluded: recovery_rejected=True or (recovery_approved=True and recovery_completed=True).
    Also detects students in courses where the module close date has passed
    and they have failing averages, even if they haven't been explicitly processed.
    Only shows entries where the group's recovery period (recovery_close) is still open.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede acceder al panel de recuperaciones")
    
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Pre-load all courses and subjects for efficient lookups
    all_courses = await db.courses.find({}, {"_id": 0}).to_list(1000)
    course_map = {c["id"]: c for c in all_courses}
    all_subjects = await db.subjects.find({}, {"_id": 0, "id": 1, "name": 1, "module_number": 1}).to_list(1000)
    subject_map = {s["id"]: s["name"] for s in all_subjects}
    subject_module_map = {s["id"]: (s.get("module_number") or 1) for s in all_subjects}

    def get_subject_names(course_doc):
        """Return list of ALL subject names for a course (fallback)."""
        sids = course_doc.get("subject_ids") or []
        if not sids and course_doc.get("subject_id"):
            sids = [course_doc["subject_id"]]
        names = [subject_map[sid] for sid in sids if sid in subject_map]
        return names if names else [course_doc.get("name", "Sin nombre")]

    def get_failing_subject_names(student_id, course_id, course_doc, grades_index):
        """Return only the subject names where the student's average is < 3.0.
        Falls back to all course subjects if per-subject data is unavailable."""
        sids = course_doc.get("subject_ids") or []
        if not sids and course_doc.get("subject_id"):
            sids = [course_doc["subject_id"]]
        if not sids:
            return [course_doc.get("name", "Sin nombre")]
        failing = []
        for sid in sids:
            values = grades_index.get((student_id, course_id, sid), [])
            # No grades or average < 3.0 → consider failed
            avg = sum(values) / len(values) if values else 0.0
            if avg < 3.0 and sid in subject_map:
                failing.append(subject_map[sid])
        return failing if failing else get_subject_names(course_doc)

    def get_failing_subjects_with_ids(student_id, course_id, course_doc, grades_index, filter_module=None):
        """Return list of (subject_id, subject_name, avg) for each failing subject.
        If filter_module is provided, only subjects belonging to that module are returned."""
        sids = course_doc.get("subject_ids") or []
        if not sids and course_doc.get("subject_id"):
            sids = [course_doc["subject_id"]]
        if not sids:
            return []
        failing = []
        for sid in sids:
            if filter_module is not None and subject_module_map.get(sid, 1) != filter_module:
                continue
            values = grades_index.get((student_id, course_id, sid), [])
            avg = sum(values) / len(values) if values else 0.0
            if avg < 3.0 and sid in subject_map:
                failing.append((sid, subject_map[sid], round(avg, 2)))
        return failing

    def get_recovery_close(course_doc, module_number):
        """Return the recovery_close date string for a given module in a course, or None."""
        module_dates = course_doc.get("module_dates") or {}
        dates = module_dates.get(str(module_number)) or {}
        return dates.get("recovery_close")

    def get_next_module_start(course_doc, module_number):
        """Return the start date of the next module in a course, or None."""
        module_dates = course_doc.get("module_dates") or {}
        next_dates = module_dates.get(str(module_number + 1)) or {}
        return next_dates.get("start")

    # Fetch only records still in-process for admin action.
    failed_records = await db.failed_subjects.find(
        {"recovery_processed": {"$ne": True}},
        {"_id": 0}
    ).to_list(1000)
    
    # Build a lookup of student cedulas for human-readable IDs
    student_ids_in_records = list({r["student_id"] for r in failed_records})
    students_lookup = {}
    if student_ids_in_records:
        student_docs = await db.users.find(
            {"id": {"$in": student_ids_in_records}},
            {"_id": 0, "id": 1, "cedula": 1, "name": 1}
        ).to_list(None)
        students_lookup = {s["id"]: s for s in student_docs}
    
    # Pre-load grades for all relevant student/course pairs to compute per-subject averages
    course_ids_in_records = list({r["course_id"] for r in failed_records})
    grades_for_records = []
    if student_ids_in_records and course_ids_in_records:
        grades_for_records = await db.grades.find(
            {"student_id": {"$in": student_ids_in_records}, "course_id": {"$in": course_ids_in_records}},
            {"_id": 0, "student_id": 1, "course_id": 1, "subject_id": 1, "value": 1, "recovery_status": 1}
        ).to_list(None)
    # Index: (student_id, course_id, subject_id) -> [grade values]
    grades_index = {}
    # Index: (student_id, course_id, subject_id) -> recovery_status (teacher's grading result)
    teacher_graded_index = {}
    for g in grades_for_records:
        key = (g["student_id"], g["course_id"], g.get("subject_id"))
        if g.get("value") is not None:
            grades_index.setdefault(key, []).append(g["value"])
        if g.get("recovery_status"):
            teacher_graded_index[(g["student_id"], g["course_id"], g.get("subject_id"))] = g["recovery_status"]

    # Get all programs for reference
    programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    program_map = {p["id"]: p["name"] for p in programs}
    program_close_map = {}
    for p in programs:
        pid = p.get("id")
        if not pid:
            continue
        max_mod = max(len(p.get("modules") or []), 2)
        for mn in range(1, max_mod + 1):
            close_val = p.get(f"module{mn}_close_date")
            if close_val:
                program_close_map[(pid, mn)] = close_val
    
    # Organize by student; include all non-processed records regardless of recovery_close date
    students_map = {}
    for record in failed_records:
        course_doc = course_map.get(record["course_id"]) or {}
        module_num = record.get("module_number", 1)
        recovery_close = get_recovery_close(course_doc, module_num)
        if not recovery_close:
            recovery_close = program_close_map.get((record.get("program_id", ""), module_num))
        next_module_start = get_next_module_start(course_doc, module_num)

        student_id = record["student_id"]
        if student_id not in students_map:
            student_doc = students_lookup.get(student_id) or {}
            students_map[student_id] = {
                "student_id": student_id,
                "student_name": record["student_name"],
                "student_cedula": student_doc.get("cedula") or "",
                "failed_subjects": []
            }
        
        subject_name = record.get("subject_name") or record.get("course_name", "Sin nombre")
        subject_id_for_record = record.get("subject_id")
        teacher_status = teacher_graded_index.get((student_id, record["course_id"], subject_id_for_record))
        if not teacher_status:
            teacher_status = teacher_graded_index.get((student_id, record["course_id"], None))

        # Compute a human-readable status for the record
        ts = record.get("teacher_graded_status") if record.get("teacher_graded_status") is not None else teacher_status
        if record.get("recovery_completed"):
            rec_status = "teacher_approved" if ts == "approved" else "teacher_rejected"
        else:
            # Not yet graded by teacher (may or may not be admin-approved)
            rec_status = "pending"

        students_map[student_id]["failed_subjects"].append({
            "id": record["id"],
            "course_id": record["course_id"],
            "course_name": record["course_name"],
            "subject_name": subject_name,
            "program_id": record["program_id"],
            "program_name": program_map.get(record["program_id"], "Desconocido"),
            "module_number": record["module_number"],
            "average_grade": next((v for v in (record.get("average_grade"), record.get("avg"), 0.0) if v is not None), 0.0),
            "recovery_approved": record["recovery_approved"],
            "recovery_completed": record["recovery_completed"],
            "recovery_processed": record.get("recovery_processed", False),
            "processed_at": record.get("processed_at"),
            "recovery_close": recovery_close,
            "next_module_start": next_module_start,
            "teacher_graded_status": ts,
            "status": rec_status
        })
    
    # Also detect students in courses with past module close dates who have failing averages
    # but are not yet in the failed_subjects collection.
    # Show entries where the recovery period is still open (auto-detected records can only be approved before close).
    
    # Track which (student_id, course_id, subject_id) combos are already in failed_subjects
    already_tracked = set()
    for record in failed_records:
        already_tracked.add((record["student_id"], record["course_id"], record.get("subject_id")))
    
    # Program close-date fallback map used when a course module has no explicit end date.
    # Key: (program_id, module_number) -> module_close_date
    program_close_map = {}
    for p in programs:
        pid = p.get("id")
        if not pid:
            continue
        max_mod = max(len(p.get("modules") or []), 2)
        for mn in range(1, max_mod + 1):
            close_val = p.get(f"module{mn}_close_date")
            if close_val:
                program_close_map[(pid, mn)] = close_val

    for course in all_courses:
        module_dates = course.get("module_dates") or {}
        # Fallback for legacy/incomplete courses: synthesize minimal module_dates
        # from program close dates so auto-detection still runs.
        if not module_dates:
            pid = course.get("program_id", "")
            fallback_dates = {
                str(mn): {"end": close_val, "recovery_close": close_val}
                for (prog_id, mn), close_val in program_close_map.items()
                if prog_id == pid
            }
            module_dates = fallback_dates
        for module_key, dates in module_dates.items():
            module_number = int(module_key) if str(module_key).isdigit() else None
            if not module_number:
                continue

            close_date = (dates or {}).get("end")
            if not close_date:
                close_date = program_close_map.get((course.get("program_id", ""), module_number))
            if not close_date or close_date > today_str:
                continue  # Module not closed yet

            recovery_close = (dates or {}).get("recovery_close")

            # Get all grades for this course once
            all_grades = await db.grades.find(
                {"course_id": course["id"]}, {"_id": 0}
            ).to_list(5000)

            # Candidate students: enrolled + anyone who already has grades in this course.
            # This covers edge-cases in the last module where enrollments may already
            # have changed but failing records were not persisted yet.
            enrolled_ids = set(course.get("student_ids") or [])
            graded_ids = {g.get("student_id") for g in all_grades if g.get("student_id")}
            candidate_student_ids = list(enrolled_ids | graded_ids)
            if not candidate_student_ids:
                continue

            # Build a per-subject index for this course's grades
            course_grades_index = {}
            for g in all_grades:
                key = (g.get("student_id"), course["id"], g.get("subject_id"))
                if g.get("value") is not None:
                    course_grades_index.setdefault(key, []).append(g["value"])

            for student_id in candidate_student_ids:
                student_grades = [g for g in all_grades if g.get("student_id") == student_id]

                # Prefer per-subject failure detection for the closed module.
                # This avoids masking a failed subject with high grades from other subjects/modules.
                failing_subjects = get_failing_subjects_with_ids(
                    student_id,
                    course["id"],
                    course,
                    course_grades_index,
                    module_number
                )

                # Fallback: if no subject structure exists, use course-level average.
                grade_values = [g["value"] for g in student_grades if g.get("value") is not None]

                average = sum(grade_values) / len(grade_values) if grade_values else 0.0
                if not failing_subjects and average >= 3.0:
                    continue  # Student passed

                # Student has failing grade – look them up and add to panel
                student = await db.users.find_one({"id": student_id, "role": "estudiante"}, {"_id": 0})
                if not student:
                    continue

                if student_id not in students_map:
                    students_map[student_id] = {
                        "student_id": student_id,
                        "student_name": student.get("name", "Desconocido"),
                        "student_cedula": student.get("cedula") or "",
                        "failed_subjects": []
                    }

                failing_subjects = get_failing_subjects_with_ids(student_id, course["id"], course, course_grades_index, module_number)
                if failing_subjects:
                    for subj_id, subj_name, subj_avg in failing_subjects:
                        if (student_id, course["id"], subj_id) in already_tracked:
                            continue  # This subject already has a persisted record
                        temp_record_id = f"auto-{student_id}-{course['id']}-{subj_id}-{module_number}"
                        teacher_status = teacher_graded_index.get((student_id, course["id"], subj_id))
                        if not teacher_status:
                            teacher_status = teacher_graded_index.get((student_id, course["id"], None))
                        students_map[student_id]["failed_subjects"].append({
                            "id": temp_record_id,
                            "course_id": course["id"],
                            "course_name": course.get("name", "Sin nombre"),
                            "subject_id": subj_id,
                            "subject_name": subj_name,
                            "program_id": course.get("program_id", ""),
                            "program_name": program_map.get(course.get("program_id", ""), "Desconocido"),
                            "module_number": module_number,
                            "average_grade": subj_avg,
                            "recovery_approved": False,
                            "recovery_completed": False,
                            "recovery_close": recovery_close,
                            "next_module_start": get_next_module_start(course, module_number),
                            "auto_detected": True,
                            "teacher_graded_status": teacher_status,
                            "status": "pending"
                        })
                        already_tracked.add((student_id, course["id"], subj_id))
                else:
                    # Fallback: no subjects defined, use a single course-level record
                    if (student_id, course["id"], None) in already_tracked:
                        continue  # Already tracked at course level
                    temp_record_id = f"auto-{student_id}-{course['id']}-{module_number}"
                    students_map[student_id]["failed_subjects"].append({
                        "id": temp_record_id,
                        "course_id": course["id"],
                        "course_name": course.get("name", "Sin nombre"),
                        "subject_name": course.get("name", "Sin nombre"),
                        "program_id": course.get("program_id", ""),
                        "program_name": program_map.get(course.get("program_id", ""), "Desconocido"),
                        "module_number": module_number,
                        "average_grade": round(average, 2),
                        "recovery_approved": False,
                        "recovery_completed": False,
                        "recovery_close": recovery_close,
                        "next_module_start": get_next_module_start(course, module_number),
                        "auto_detected": True,
                        "teacher_graded_status": teacher_graded_index.get((student_id, course["id"], None)),
                        "status": "pending"
                    })
                    already_tracked.add((student_id, course["id"], None))
    
    filtered_students = [s for s in students_map.values() if s.get("failed_subjects")]

    return {
        "students": filtered_students,
        "total_students": len(filtered_students),
        "total_failed_subjects": sum(len(s["failed_subjects"]) for s in filtered_students)
    }

@api_router.post("/admin/approve-recovery")
async def approve_recovery_for_subject(failed_subject_id: str, approve: bool, user=Depends(get_current_user)):
    """
    Admin approves or rejects recovery for a specific failed subject.
    If approved (approve=True), student can see and complete recovery activities.
    If rejected (approve=False), student is immediately removed from all program courses
    and marked as 'reprobado'.
    Handles both persisted records and auto-detected entries (id starts with 'auto-').
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede aprobar recuperaciones")

    now = datetime.now(timezone.utc)

    # Handle auto-detected entries (not persisted yet).
    # Formats:
    #   Old: auto-{student_id}-{course_id}-{module_number}           (sc_part = 73 chars)
    #   New: auto-{student_id}-{course_id}-{subject_id}-{module_number} (sc_part = 110 chars)
    if failed_subject_id.startswith("auto-"):
        # UUIDs are always exactly 36 chars. Strip 'auto-' prefix then parse fixed-length fields.
        remainder = failed_subject_id[5:]  # strip 'auto-'
        # last '-' separates the module_number (an integer, no dashes)
        last_dash = remainder.rfind("-")
        if last_dash == -1:
            raise HTTPException(status_code=404, detail="Registro de materia reprobada no encontrado")
        module_str = remainder[last_dash + 1:]
        sc_part = remainder[:last_dash]
        # sc_part is either:
        #   "{student_id}-{course_id}"                   → 36 + 1 + 36 = 73 chars (old)
        #   "{student_id}-{course_id}-{subject_id}"      → 36 + 1 + 36 + 1 + 36 = 110 chars (new)
        subject_id = None
        # UUIDs are 36 chars each. sc_part lengths:
        #   old: 36 (student) + 1 (-) + 36 (course)          = 73
        #   new: 36 (student) + 1 (-) + 36 (course) + 1 (-) + 36 (subject) = 110
        if len(sc_part) == 73:
            student_id = sc_part[:36]
            course_id = sc_part[37:]
        elif len(sc_part) == 110:
            student_id = sc_part[:36]
            course_id = sc_part[37:73]
            subject_id = sc_part[74:]
        else:
            raise HTTPException(status_code=404, detail="Registro de materia reprobada no encontrado")

        # Validate that both IDs exist in the database
        student = await db.users.find_one({"id": student_id, "role": "estudiante"}, {"_id": 0})
        course = await db.courses.find_one({"id": course_id}, {"_id": 0})
        if not student or not course:
            raise HTTPException(status_code=404, detail="Estudiante o grupo no encontrado")

        prog_id = course.get("program_id", "")
        student_module = (student.get("program_modules") or {}).get(prog_id) or student.get("module")
        module_number = int(module_str) if module_str.isdigit() else 1

        all_grades = await db.grades.find({"student_id": student_id, "course_id": course_id}, {"_id": 0}).to_list(100)

        # Compute average: per-subject if subject_id available, otherwise course-level
        if subject_id:
            subject_grade_values = [g["value"] for g in all_grades if g.get("subject_id") == subject_id and g.get("value") is not None]
            average = sum(subject_grade_values) / len(subject_grade_values) if subject_grade_values else 0.0
        else:
            grade_values = [g["value"] for g in all_grades if g.get("value") is not None]
            average = sum(grade_values) / len(grade_values) if grade_values else 0.0

        if not approve:
            # Admin explicitly rejects this auto-detected recovery: create a rejection record,
            # remove student from the affected course and mark as reprobado.
            rejection_record = {
                "id": str(uuid.uuid4()),
                "student_id": student_id,
                "student_name": student.get("name", "Desconocido"),
                "course_id": course_id,
                "course_name": course.get("name", "Sin nombre"),
                "program_id": prog_id,
                "module_number": module_number,
                "average_grade": round(average, 2),
                "recovery_approved": False,
                "recovery_rejected": True,
                "recovery_completed": False,
                "recovery_processed": True,
                "rejected_by": user["id"],
                "rejected_at": now.isoformat(),
                "processed_at": now.isoformat(),
                "created_at": now.isoformat()
            }
            if subject_id:
                subject_doc = await db.subjects.find_one({"id": subject_id}, {"_id": 0, "name": 1})
                rejection_record["subject_id"] = subject_id
                rejection_record["subject_name"] = subject_doc.get("name", "Desconocido") if subject_doc else "Desconocido"
            await db.failed_subjects.insert_one(rejection_record)
            # Remove student only from this course (auto-detected rejection path)
            await _unenroll_student_from_course(student_id, course_id)
            # Mark student as reprobado
            student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
            student_program_statuses = (student_doc or {}).get("program_statuses") or {}
            if prog_id:
                student_program_statuses[prog_id] = "reprobado"
            new_estado = derive_estado_from_program_statuses(student_program_statuses)
            update_fields: dict = {"estado": new_estado}
            if prog_id:
                update_fields["program_statuses"] = student_program_statuses
            await db.users.update_one({"id": student_id, "role": "estudiante"}, {"$set": update_fields})
            logger.info(
                f"Admin {user['id']} rejected recovery for student {student_id} "
                f"in course {course_id} (auto-detected entry)"
            )
            await log_audit(
                "recovery_rejected_by_admin", user["id"], user["role"],
                {"student_id": student_id, "course_id": course_id, "program_id": prog_id,
                 "module_number": module_number}
            )
            return {"message": "Recuperación rechazada. El estudiante ha sido retirado del grupo correspondiente."}

        # Module alignment: recovery can only be approved for the student's current module
        if student_module is not None and student_module != module_number:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"No se puede aprobar: la recuperación es del Módulo {module_number} "
                    f"pero el estudiante está en el Módulo {student_module}."
                )
            )

        new_record = {
            "id": str(uuid.uuid4()),
            "student_id": student_id,
            "student_name": student.get("name", "Desconocido"),
            "course_id": course_id,
            "course_name": course.get("name", "Sin nombre"),
            "program_id": course.get("program_id", ""),
            "module_number": module_number,
            "average_grade": round(average, 2),
            "recovery_approved": True,
            "recovery_rejected": False,
            "recovery_completed": False,
            "approved_by": user["id"],
            "approved_at": now.isoformat(),
            "created_at": now.isoformat()
        }
        if subject_id:
            subject_doc = await db.subjects.find_one({"id": subject_id}, {"_id": 0, "name": 1})
            new_record["subject_id"] = subject_id
            new_record["subject_name"] = subject_doc.get("name", "Desconocido") if subject_doc else "Desconocido"
        await db.failed_subjects.insert_one(new_record)

        # Enable recovery activities for this student/course
        existing = await db.recovery_enabled.find_one({"student_id": student_id, "course_id": course_id, "subject_id": subject_id})
        if not existing:
            await db.recovery_enabled.insert_one({
                "id": str(uuid.uuid4()),
                "student_id": student_id,
                "course_id": course_id,
                "subject_id": subject_id,
                "enabled": True,
                "enabled_by": user["id"],
                "enabled_at": now.isoformat()
            })
        else:
            await db.recovery_enabled.update_one(
                {"id": existing["id"]},
                {"$set": {"enabled": True, "enabled_by": user["id"], "enabled_at": now.isoformat()}}
            )
        # Update program_statuses per-program and derive global estado
        prog_id = course.get("program_id", "")
        student_doc = await db.users.find_one({"id": student_id}, {"_id": 0, "program_statuses": 1})
        student_program_statuses = (student_doc or {}).get("program_statuses") or {}
        if prog_id:
            student_program_statuses[prog_id] = "pendiente_recuperacion"
        new_estado = derive_estado_from_program_statuses(student_program_statuses)
        update_fields = {"estado": new_estado}
        if prog_id:
            update_fields["program_statuses"] = student_program_statuses
        await db.users.update_one({"id": student_id}, {"$set": update_fields})

        return {"message": "Recuperación aprobada exitosamente"}

    # Find the failed subject record
    failed_record = await db.failed_subjects.find_one({"id": failed_subject_id}, {"_id": 0})
    if not failed_record:
        raise HTTPException(status_code=404, detail="Registro de materia reprobada no encontrado")

    if not approve:
        # Admin explicitly rejects this persisted recovery record.
        # Remove student only from this course and mark as reprobado for that program.
        rej_student_id = failed_record["student_id"]
        rej_course_id = failed_record["course_id"]
        rej_prog_id = failed_record.get("program_id", "")
        # Mark this record as rejected/processed
        await db.failed_subjects.update_one(
            {"id": failed_subject_id},
            {"$set": {
                "recovery_approved": False,
                "recovery_rejected": True,
                "recovery_processed": True,
                "rejected_by": user["id"],
                "rejected_at": now.isoformat(),
                "processed_at": now.isoformat()
            }}
        )
        # Also mark all other unprocessed records for this student/course as processed
        await db.failed_subjects.update_many(
            {
                "student_id": rej_student_id,
                "course_id": rej_course_id,
                "id": {"$ne": failed_subject_id},
                "recovery_processed": {"$ne": True}
            },
            {"$set": {"recovery_processed": True, "processed_at": now.isoformat()}}
        )
        # Remove student only from this course (auto-detected rejection path)
        await _unenroll_student_from_course(rej_student_id, rej_course_id)
        # Mark student as reprobado
        rej_student_doc = await db.users.find_one({"id": rej_student_id}, {"_id": 0, "program_statuses": 1})
        rej_program_statuses = (rej_student_doc or {}).get("program_statuses") or {}
        if rej_prog_id:
            rej_program_statuses[rej_prog_id] = "reprobado"
        rej_new_estado = derive_estado_from_program_statuses(rej_program_statuses)
        rej_update: dict = {"estado": rej_new_estado}
        if rej_prog_id:
            rej_update["program_statuses"] = rej_program_statuses
        await db.users.update_one({"id": rej_student_id, "role": "estudiante"}, {"$set": rej_update})
        logger.info(
            f"Admin {user['id']} rejected recovery for student {rej_student_id} "
            f"in course {rej_course_id} (record {failed_subject_id})"
        )
        await log_audit(
            "recovery_rejected_by_admin", user["id"], user["role"],
            {"student_id": rej_student_id, "failed_subject_id": failed_subject_id,
             "course_id": rej_course_id, "program_id": rej_prog_id}
        )
        return {"message": "Recuperación rechazada. El estudiante ha sido retirado del grupo correspondiente."}

    # Module alignment: reject approval if subject module doesn't match student's current module
    record_module = failed_record.get("module_number")
    if record_module is not None:
        stud_doc = await db.users.find_one({"id": failed_record["student_id"]}, {"_id": 0, "program_modules": 1, "module": 1})
        if stud_doc:
            rec_prog_id = failed_record.get("program_id", "")
            stud_mod = (stud_doc.get("program_modules") or {}).get(rec_prog_id) or stud_doc.get("module")
            if stud_mod is not None and stud_mod != record_module:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"No se puede aprobar: la recuperación es del Módulo {record_module} "
                        f"pero el estudiante está en el Módulo {stud_mod}."
                    )
                )
    
    # Update approval status
    await db.failed_subjects.update_one(
        {"id": failed_subject_id},
        {"$set": {
            "recovery_approved": True,
            "recovery_rejected": False,
            "approved_by": user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Enable recovery in the recovery_enabled collection
    recovery = {
        "id": str(uuid.uuid4()),
        "student_id": failed_record["student_id"],
        "course_id": failed_record["course_id"],
        "subject_id": failed_record.get("subject_id"),
        "enabled": True,
        "enabled_by": user["id"],
        "enabled_at": datetime.now(timezone.utc).isoformat()
    }
    
    existing = await db.recovery_enabled.find_one({
        "student_id": failed_record["student_id"],
        "course_id": failed_record["course_id"],
        "subject_id": failed_record.get("subject_id")
    })
    
    if not existing:
        await db.recovery_enabled.insert_one(recovery)
    else:
        await db.recovery_enabled.update_one(
            {"id": existing["id"]},
            {"$set": {"enabled": True, "enabled_by": user["id"], "enabled_at": datetime.now(timezone.utc).isoformat()}}
        )
    # Update program_statuses per-program and derive global estado
    prog_id = failed_record.get("program_id", "")
    student_doc = await db.users.find_one({"id": failed_record["student_id"]}, {"_id": 0, "program_statuses": 1})
    student_program_statuses = (student_doc or {}).get("program_statuses") or {}
    if prog_id:
        student_program_statuses[prog_id] = "pendiente_recuperacion"
    new_estado = derive_estado_from_program_statuses(student_program_statuses)
    update_fields = {"estado": new_estado}
    if prog_id:
        update_fields["program_statuses"] = student_program_statuses
    await db.users.update_one({"id": failed_record["student_id"]}, {"$set": update_fields})
    await log_audit("recovery_approved", user["id"], user["role"], {"student_id": failed_record["student_id"], "failed_subject_id": failed_subject_id, "subject_name": failed_record.get("subject_name", ""), "course_id": failed_record.get("course_id", "")})
    return {"message": "Recuperación aprobada exitosamente"}

@api_router.get("/student/my-recoveries")
async def get_student_recoveries(user=Depends(get_current_user)):
    """
    Get recovery subjects for the current student.
    Returns all failed subjects that are not fully processed by recovery-close,
    including teacher-rejected ones, so students can still see the result
    in the Recuperaciones tab until the module recovery period closes.
    that match the student's current module for each program.
    The 'recovery_approved' field indicates whether admin has approved recovery.
    """
    if user["role"] != "estudiante":
        raise HTTPException(status_code=403, detail="Solo estudiantes pueden acceder a sus recuperaciones")
    
    student_id = user["id"]
    program_modules = user.get("program_modules") or {}
    student_global_module = user.get("module") or 1
    
    # Get all failed subjects not yet fully processed for this student
    # Includes records where teacher has graded (recovery_completed) so student can see the outcome
    failed_subjects = await db.failed_subjects.find({
        "student_id": student_id,
        "recovery_processed": {"$ne": True}
    }, {"_id": 0}).to_list(100)
    
    # Load actual subject module info from DB to validate against stale/incorrect records
    all_subject_docs = await db.subjects.find({}, {"_id": 0, "id": 1, "module_number": 1}).to_list(1000)
    db_subject_module_map = {s["id"]: (s.get("module_number") or 1) for s in all_subject_docs}

    # Filter by student's current module for each program.
    # Use the subject's actual module from the DB when available (defense against
    # stale records where module_number in failed_subjects does not match the subject).
    filtered = []
    for subject in failed_subjects:
        prog_id = subject.get("program_id", "")
        # Prefer actual subject module from DB over the value stored in the record
        sid = subject.get("subject_id")
        if sid and sid in db_subject_module_map:
            subject_module = db_subject_module_map[sid]
        else:
            subject_module = subject.get("module_number")
        if subject_module is None:
            filtered.append(subject)
            continue
        current_module = program_modules.get(prog_id)
        if current_module is None:
            # Fallback: only if not tracked per-program, use the legacy global module
            current_module = user.get("module") or 1
        if subject_module == current_module:
            filtered.append(subject)
    failed_subjects = filtered
    
    # Get program info for each
    programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    program_map = {p["id"]: p["name"] for p in programs}
    
    # Pre-load subject names for records that may be missing the field (backward compat)
    missing_subject_ids = [s["subject_id"] for s in failed_subjects if s.get("subject_id") and not s.get("subject_name")]
    subject_name_lookup = {}
    if missing_subject_ids:
        subj_docs = await db.subjects.find({"id": {"$in": missing_subject_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        subject_name_lookup = {s["id"]: s["name"] for s in subj_docs}
    
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    for subject in failed_subjects:
        subject["program_name"] = program_map.get(subject["program_id"], "Desconocido")
        # Ensure subject_name is populated (fallback for older records)
        if not subject.get("subject_name") and subject.get("subject_id"):
            subject["subject_name"] = subject_name_lookup.get(subject["subject_id"], "")
        # Check if recovery closing date has passed for this module/course
        course = await db.courses.find_one({"id": subject["course_id"]}, {"_id": 0, "module_dates": 1})
        recovery_close = None
        if course and course.get("module_dates"):
            # module_dates keys are stored as strings (JSON object keys are always strings)
            module_key = str(subject.get("module_number", ""))
            module_dates = course["module_dates"].get(module_key)
            if module_dates:
                recovery_close = module_dates.get("recovery_close")
        subject["recovery_close_date"] = recovery_close
        subject["recovery_closed"] = bool(recovery_close and recovery_close < today_str)
        # Ensure recovery_approved field is present (backward-compat with older records)
        subject.setdefault("recovery_approved", False)
        teacher_status = subject.get("teacher_graded_status")
        if subject.get("recovery_completed") and teacher_status == "rejected":
            subject["status_label"] = "reprobado"
        elif subject.get("recovery_completed") and teacher_status == "approved":
            subject["status_label"] = "aprobado"
        elif subject.get("recovery_approved"):
            subject["status_label"] = "habilitada"
        else:
            subject["status_label"] = "pendiente_aprobacion"
    
    return {
        "recoveries": failed_subjects,
        "total": len(failed_subjects)
    }

@api_router.get("/reports/course-results")
async def get_course_results_report(course_id: str, subject_id: Optional[str] = None, format: Optional[str] = None, user=Depends(get_current_user)):
    """
    Returns a report of approved/reproved students per course/group.
    Accessible by admin and professor.
    When subject_id is provided, filters the report to only that subject.
    When format=csv/xlsx, returns a file download; otherwise returns JSON.
    """
    if user["role"] not in ["admin", "profesor"]:
        raise HTTPException(status_code=403, detail="Solo admin o profesor pueden acceder a reportes")
    
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    # Load subjects
    all_subject_ids = course.get("subject_ids") or []
    if not all_subject_ids and course.get("subject_id"):
        all_subject_ids = [course["subject_id"]]
    # When subject_id filter is provided, restrict to that single subject
    subject_ids = [subject_id] if subject_id and subject_id in all_subject_ids else all_subject_ids
    all_subjects = await db.subjects.find({"id": {"$in": subject_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None) if subject_ids else []
    subject_map = {s["id"]: s["name"] for s in all_subjects}
    
    # Load enrolled students
    student_ids = course.get("student_ids") or []
    students = await db.users.find(
        {"id": {"$in": student_ids}, "role": "estudiante"},
        {"_id": 0, "id": 1, "name": 1, "cedula": 1}
    ).to_list(None) if student_ids else []
    
    # Load grades for this course and build an index: (student_id, subject_id) -> [values]
    grades = await db.grades.find({"course_id": course_id}, {"_id": 0, "student_id": 1, "subject_id": 1, "value": 1}).to_list(None)
    grades_index = {}
    for g in grades:
        if g.get("value") is not None:
            key = (g["student_id"], g.get("subject_id"))
            grades_index.setdefault(key, []).append(g["value"])
    
    # Determine module number for the group (from first subject, or from course module_dates)
    # A course group targets a single module; all subjects in the course should share the same
    # module_number in practice. We use the first available value as a best-effort.
    module_number = None
    if all_subjects:
        module_number = all_subjects[0].get("module_number")
    if module_number is None:
        module_dates = course.get("module_dates") or {}
        if module_dates:
            module_number = next(iter(module_dates.keys()), None)

    # Build per-student summary rows (one row per student, columns for each subject average)
    rows = []
    for student in students:
        sid = student["id"]
        subject_avgs = {}
        for subj_id in subject_ids:
            values = grades_index.get((sid, subj_id), [])
            subject_avgs[subj_id] = round(sum(values) / len(values), 2) if values else 0.0
        # General average: use only the filtered subjects
        all_values = [v for subj_id in subject_ids for v in grades_index.get((sid, subj_id), [])]
        if not all_values:
            # Fall back: no subject_ids configured – use all grades for this course
            all_values = [v for (s, _subj), vals in grades_index.items() if s == sid for v in vals]
        general_avg = round(sum(all_values) / len(all_values), 2) if all_values else 0.0
        # Check recovery status
        recovery_record = await db.failed_subjects.find_one(
            {"student_id": sid, "course_id": course_id, "recovery_processed": {"$ne": True}},
            {"_id": 0, "recovery_approved": 1}
        )
        if general_avg >= 3.0:
            status = "Aprobado"
        elif recovery_record:
            status = "En Recuperación"
        else:
            status = "Reprobado"
        row = {
            "student_name": student["name"],
            "student_cedula": student.get("cedula", ""),
            "module": module_number,
            "course_name": course.get("name", ""),
            "general_average": general_avg,
            "status": status,
            "student_id": sid,
            "course_id": course_id,
        }
        for subj_id in subject_ids:
            row[f"subject_{subj_id}"] = subject_avgs[subj_id]
            row[f"subject_name_{subj_id}"] = subject_map.get(subj_id, subj_id)
        rows.append(row)
    
    if format and format.lower() in ("csv", "xlsx"):
        course_name = course.get("name", course_id)
        safe_name = re.sub(r'[^\w\-]', '_', course_name)
        subject_headers = [subject_map.get(sid, sid) for sid in subject_ids]
        base_headers = ["Nombre", "Cédula", "Módulo"]
        end_headers = ["Promedio", "Estado"]
        all_headers = base_headers + subject_headers + end_headers

        if format.lower() == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(all_headers)
            for row in rows:
                csv_row = [row["student_name"], row["student_cedula"], row.get("module", "")]
                for subj_id in subject_ids:
                    csv_row.append(row.get(f"subject_{subj_id}", 0.0))
                csv_row.extend([row["general_average"], row["status"]])
                writer.writerow(csv_row)
            output.seek(0)
            filename = f"resultados_{safe_name}.csv"
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            # XLSX with professional formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados"

            # Title row
            title_label = course_name
            if subject_id and subject_id in subject_map:
                title_label = f"{course_name} – {subject_map[subject_id]}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
            title_cell = ws.cell(row=1, column=1, value=f"Reporte de Resultados – {title_label}")
            title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_cell.font = Font(bold=True, size=13, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Header row (row 2)
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(border_style="thin", color="000000")
            header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, header in enumerate(all_headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[2].height = 18

            # Data rows (starting row 3)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row_idx, row in enumerate(rows, start=3):
                avg = row["general_average"]
                row_fill = green_fill if avg >= 3.0 else red_fill
                data = [row["student_name"], row["student_cedula"], row.get("module", "")]
                for subj_id in subject_ids:
                    data.append(row.get(f"subject_{subj_id}", 0.0))
                data.extend([avg, row["status"]])
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = row_fill
                    cell.border = data_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-adjust column widths
            for col_idx, header in enumerate(all_headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(header)), 10)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
            # Fixed widths for known columns
            ws.column_dimensions["A"].width = 30  # Nombre
            ws.column_dimensions["B"].width = 14  # Cédula

            output_bytes = io.BytesIO()
            wb.save(output_bytes)
            output_bytes.seek(0)
            if subject_id and subject_id in subject_map:
                safe_subject = re.sub(r'[^\w\-]', '_', subject_map[subject_id])
                filename = f"resultados_{safe_name}_{safe_subject}.xlsx"
            else:
                filename = f"resultados_{safe_name}.xlsx"
            return StreamingResponse(
                iter([output_bytes.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    
    return {
        "course_id": course_id,
        "course_name": course.get("name", ""),
        "module": module_number,
        "subjects": [{"id": sid, "name": subject_map.get(sid, sid)} for sid in subject_ids],
        "rows": rows,
        "total": len(rows)
    }


@api_router.get("/reports/recovery-results")
async def get_recovery_results_report(format: Optional[str] = None, user=Depends(get_current_user)):
    """
    Returns a consolidated report of all students in recovery with their status.
    Accessible by admin only.
    When format=xlsx, returns an XLSX file download; otherwise returns JSON.
    """
    if user["role"] not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Solo admin puede acceder a reportes de recuperación")

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Fetch all unprocessed failed subjects
    all_failed = await db.failed_subjects.find(
        {"recovery_processed": {"$ne": True}},
        {"_id": 0}
    ).to_list(None)

    if not all_failed:
        all_failed = []

    # Build lookup maps
    student_ids_set = list({r["student_id"] for r in all_failed})
    course_ids_set = list({r["course_id"] for r in all_failed})
    program_ids_set = list({r.get("program_id") for r in all_failed if r.get("program_id")})
    subject_ids_set = list({r.get("subject_id") for r in all_failed if r.get("subject_id")})

    students_list = await db.users.find({"id": {"$in": student_ids_set}}, {"_id": 0, "id": 1, "name": 1, "cedula": 1}).to_list(None) if student_ids_set else []
    courses_list = await db.courses.find({"id": {"$in": course_ids_set}}, {"_id": 0, "id": 1, "name": 1, "module_dates": 1}).to_list(None) if course_ids_set else []
    programs_list = await db.programs.find({"id": {"$in": program_ids_set}}, {"_id": 0, "id": 1, "name": 1}).to_list(None) if program_ids_set else []
    subjects_list = await db.subjects.find({"id": {"$in": subject_ids_set}}, {"_id": 0, "id": 1, "name": 1}).to_list(None) if subject_ids_set else []

    student_map = {s["id"]: s for s in students_list}
    course_map = {c["id"]: c for c in courses_list}
    program_map = {p["id"]: p["name"] for p in programs_list}
    subject_map = {s["id"]: s["name"] for s in subjects_list}

    rows = []
    for record in all_failed:
        student = student_map.get(record["student_id"], {})
        course = course_map.get(record["course_id"], {})
        # Determine recovery close date
        module_key = str(record.get("module_number", ""))
        module_dates = (course.get("module_dates") or {}).get(module_key) or {}
        recovery_close = module_dates.get("recovery_close")
        is_expired = bool(recovery_close and recovery_close <= today_str)
        # Determine status label
        if is_expired and not record.get("recovery_approved"):
            status_label = "Plazo vencido"
        elif record.get("teacher_graded_status") == "approved":
            status_label = "Calificada por profesor: Aprobado"
        elif record.get("teacher_graded_status") == "rejected":
            status_label = "Calificada por profesor: Reprobado"
        elif record.get("recovery_approved"):
            status_label = "Aprobada por admin"
        else:
            status_label = "Pendiente"
        subject_name = record.get("subject_name") or subject_map.get(record.get("subject_id"), "")
        rows.append({
            "student_name": student.get("name", record.get("student_name", "")),
            "student_cedula": student.get("cedula", ""),
            "subject_name": subject_name,
            "course_name": record.get("course_name", course.get("name", "")),
            "program_name": record.get("program_name", program_map.get(record.get("program_id"), "")),
            "module_number": record.get("module_number", ""),
            "average_grade": record.get("average_grade", 0.0),
            "status": status_label,
        })

    if format and format.lower() == "xlsx":
        all_headers = ["Nombre", "Cédula", "Materia reprobada", "Grupo/Curso", "Programa", "Módulo", "Promedio anterior", "Estado de recuperación"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recuperaciones"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
        title_cell = ws.cell(row=1, column=1, value="Reporte de Recuperaciones")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="000000")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 18

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row_idx, row in enumerate(rows, start=3):
            status = row["status"]
            if "Aprobado" in status:
                row_fill = green_fill
            elif "Reprobado" in status or "vencido" in status:
                row_fill = red_fill
            else:
                row_fill = yellow_fill
            data = [
                row["student_name"], row["student_cedula"], row["subject_name"],
                row["course_name"], row["program_name"], row["module_number"],
                row["average_grade"], row["status"]
            ]
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = data_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 32

        output_bytes = io.BytesIO()
        wb.save(output_bytes)
        output_bytes.seek(0)
        return StreamingResponse(
            iter([output_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=reporte_recuperaciones.xlsx"}
        )

    return {"rows": rows, "total": len(rows)}


async def delete_graduated_students(user=Depends(get_current_user)):
    """
    Admin/Editor deletes all graduated students from the system.
    This removes all their data including grades, submissions, and user records.
    This action is irreversible and should be used to clean up graduated students.
    """
    if user["role"] not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Solo admin/editor pueden eliminar estudiantes egresados")
    
    # Find all graduated students (no limit - get all of them)
    graduated_students = await db.users.find(
        {"role": "estudiante", "estado": "egresado"},
        {"_id": 0, "id": 1}
    ).to_list(None)  # None means no limit
    
    if not graduated_students:
        return {
            "message": "No hay estudiantes egresados para eliminar",
            "deleted_count": 0
        }
    
    student_ids = [s["id"] for s in graduated_students]
    
    # Delete related data
    grades_deleted = await db.grades.delete_many({"student_id": {"$in": student_ids}})
    submissions_deleted = await db.submissions.delete_many({"student_id": {"$in": student_ids}})
    failed_subjects_deleted = await db.failed_subjects.delete_many({"student_id": {"$in": student_ids}})
    recovery_enabled_deleted = await db.recovery_enabled.delete_many({"student_id": {"$in": student_ids}})
    
    # Remove students from course student_ids arrays
    await db.courses.update_many(
        {},
        {"$pull": {"student_ids": {"$in": student_ids}}}
    )
    
    # Delete the students themselves
    students_deleted = await db.users.delete_many({"id": {"$in": student_ids}})
    
    return {
        "message": f"Se eliminaron {students_deleted.deleted_count} estudiantes egresados y sus datos relacionados",
        "deleted_count": students_deleted.deleted_count,
        "grades_deleted": grades_deleted.deleted_count,
        "submissions_deleted": submissions_deleted.deleted_count,
        "failed_subjects_deleted": failed_subjects_deleted.deleted_count,
        "recovery_enabled_deleted": recovery_enabled_deleted.deleted_count
    }

@api_router.get("/admin/graduated-students-count")
async def get_graduated_students_count(user=Depends(get_current_user)):
    """
    Get count of graduated students for display purposes.
    """
    if user["role"] not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Solo admin/editor pueden ver esta información")
    
    count = await db.users.count_documents({"role": "estudiante", "estado": "egresado"})
    
    return {
        "count": count
    }

@api_router.post("/admin/reset-users")
async def reset_users(confirm_token: str = None):
    """
    DANGER: Deletes ALL users and creates fresh default users.
    Creates: 2 admins, 1 editor, 2 professors, 2 students
    
    Requires confirmation token: 'RESET_ALL_USERS_CONFIRM'
    
    This endpoint should be disabled or protected in production.
    Set environment variable ALLOW_USER_RESET=false to disable.
    """
    # Check if running in production environment
    if os.environ.get('RENDER') or os.environ.get('RAILWAY_ENVIRONMENT'):
        raise HTTPException(status_code=403, detail="This endpoint is disabled in production")

    # Check if endpoint is allowed (can be disabled via env var)
    allow_reset = os.environ.get('ALLOW_USER_RESET', 'true').lower() == 'true'
    if not allow_reset:
        raise HTTPException(
            status_code=403, 
            detail="Este endpoint está deshabilitado en producción"
        )
    
    # Require confirmation token
    if confirm_token != "RESET_ALL_USERS_CONFIRM":
        raise HTTPException(
            status_code=400,
            detail="Token de confirmación requerido. Parámetro: confirm_token='RESET_ALL_USERS_CONFIRM'"
        )
    
    # Delete ALL users
    deleted_count = await db.users.delete_many({})
    
    # Create new default users
    default_users = [
        # 2 Admins
        {
            "id": str(uuid.uuid4()),
            "name": "Admin Principal",
            "email": "admin@educando.com",
            "cedula": None,
            "password_hash": hash_password("Admin2026"),
            "role": "admin",
            "program_id": None,
            "phone": "3001234567",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Admin Secundario",
            "email": "admin2@educando.com",
            "cedula": None,
            "password_hash": hash_password("Admin2026"),
            "role": "admin",
            "program_id": None,
            "phone": "3001234568",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        # 1 Editor (logs in through profesor login)
        {
            "id": str(uuid.uuid4()),
            "name": "Editor Principal",
            "email": "editor@educando.com",
            "cedula": None,
            "password_hash": hash_password("Editor2026"),
            "role": "editor",
            "program_id": None,
            "phone": "3002222222",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        # 2 Professors
        {
            "id": str(uuid.uuid4()),
            "name": "María García",
            "email": "profesor@educando.com",
            "cedula": None,
            "password_hash": hash_password("Profe2026"),
            "role": "profesor",
            "program_id": None,
            "phone": "3007654321",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Carlos Rodríguez",
            "email": "profesor2@educando.com",
            "cedula": None,
            "password_hash": hash_password("Profe2026"),
            "role": "profesor",
            "program_id": None,
            "phone": "3009876543",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        # 2 Students
        {
            "id": str(uuid.uuid4()),
            "name": "Juan Martínez",
            "email": None,
            "cedula": "1001",
            "password_hash": hash_password("1001"),
            "role": "estudiante",
            "program_id": None,
            "phone": "3101234567",
            "active": True,
            "estado": "activo",
            "current_module": 1,
            "program_ids": [],
            "curso_ids": [],
            "program_modules": {},
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Ana Hernández",
            "email": None,
            "cedula": "1002",
            "password_hash": hash_password("1002"),
            "role": "estudiante",
            "program_id": None,
            "phone": "3207654321",
            "active": True,
            "estado": "activo",
            "current_module": 1,
            "program_ids": [],
            "curso_ids": [],
            "program_modules": {},
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.users.insert_many(default_users)
    
    return {
        "message": "Usuarios reiniciados exitosamente",
        "deleted_count": deleted_count.deleted_count,
        "created_count": len(default_users),
        "users": [
            {"role": "admin", "login": "admin@educando.com", "password": "Admin2026"},
            {"role": "admin", "login": "admin2@educando.com", "password": "Admin2026"},
            {"role": "editor", "login": "editor@educando.com (usar login de profesor)", "password": "Editor2026"},
            {"role": "profesor", "login": "profesor@educando.com", "password": "Profe2026"},
            {"role": "profesor", "login": "profesor2@educando.com", "password": "Profe2026"},
            {"role": "estudiante", "login": "1001 (cédula)", "password": "1001"},
            {"role": "estudiante", "login": "1002 (cédula)", "password": "1002"}
        ]
    }

# --- Seed Data Route ---
@api_router.post("/seed")
async def seed_data():
    # Block in production
    if os.environ.get('RENDER') or os.environ.get('RAILWAY_ENVIRONMENT'):
        raise HTTPException(status_code=403, detail="Este endpoint está deshabilitado en producción")
    # Check if already seeded
    existing_admin = await db.users.find_one({"email": "admin@educando.com"})
    if existing_admin:
        return {"message": "Base de datos ya tiene datos iniciales"}
    
    # Check for protected production users that must never be re-created
    protected_emails = ["laura.torres@educando.com", "diana.silva@educando.com"]
    protected_cedulas = ["1001234567"]
    for email in protected_emails:
        if await db.users.find_one({"email": email}):
            return {"message": "Datos de producción ya existen. No se recrearán usuarios existentes."}
    for cedula in protected_cedulas:
        if await db.users.find_one({"cedula": cedula}):
            return {"message": "Datos de producción ya existen. No se recrearán usuarios existentes."}
    
    # Create Programs
    programs = [
        {
            "id": "prog-admin",
            "name": "Técnico en Asistencia Administrativa",
            "description": "Formación técnica en gestión administrativa, contabilidad, ofimática y gestión documental.",
            "duration": "12 meses (2 módulos de 6 meses)",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Fundamentos de Administración",
                    "Herramientas Ofimáticas",
                    "Gestión Documental y Archivo",
                    "Atención al Cliente y Comunicación Organizacional",
                    "Legislación Laboral y Ética Profesional"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Contabilidad Básica",
                    "Nómina y Seguridad Social Aplicada",
                    "Control de Inventarios y Logística",
                    "Inglés Técnico / Competencias Ciudadanas",
                    "Proyecto Integrador Virtual"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": "prog-infancia",
            "name": "Técnico Laboral en Atención a la Primera Infancia",
            "description": "Formación para el cuidado y educación de niños en primera infancia.",
            "duration": "12 meses (2 módulos de 6 meses)",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Inglés I",
                    "Proyecto de vida",
                    "Construcción social de la infancia",
                    "Perspectiva del desarrollo infantil",
                    "Salud y nutrición",
                    "Lenguaje y educación infantil",
                    "Juego y otras formas de comunicación",
                    "Educación y pedagogía"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Inglés II",
                    "Construcción del mundo Matemático",
                    "Dificultades en el aprendizaje",
                    "Estrategias del aula",
                    "Trabajo de grado",
                    "Investigación",
                    "Práctica - Informe"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": "prog-sst",
            "name": "Técnico en Seguridad y Salud en el Trabajo",
            "description": "Formación en prevención de riesgos laborales, medicina preventiva y gestión ambiental.",
            "duration": "12 meses (2 módulos)",
            "modules": [
                {"number": 1, "name": "MÓDULO 1", "subjects": [
                    "Fundamentos en Seguridad y Salud en el Trabajo",
                    "Administración en salud",
                    "Condiciones de seguridad",
                    "Matemáticas",
                    "Psicología del Trabajo"
                ]},
                {"number": 2, "name": "MÓDULO 2", "subjects": [
                    "Comunicación oral y escrita",
                    "Sistema de gestión de seguridad y salud del trabajo",
                    "Anatomía y fisiología",
                    "Medicina preventiva del trabajo",
                    "Ética profesional",
                    "Gestión ambiental",
                    "Proyecto de grado"
                ]}
            ],
            "module1_close_date": None,
            "module2_close_date": None,
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.programs.insert_many(programs)
    
    # Create Subjects
    subjects = []
    for prog in programs:
        for module in prog["modules"]:
            for subj_name in module["subjects"]:
                subjects.append({
                    "id": str(uuid.uuid4()),
                    "name": subj_name,
                    "program_id": prog["id"],
                    "module_number": module["number"],
                    "description": "",
                    "active": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
    await db.subjects.insert_many(subjects)
    
    # Create Users
    admin_id = "user-admin"
    editor_id = "user-editor-2"
    teacher1_id = "user-teacher1"
    teacher2_id = "user-teacher2"
    student1_id = "user-student1"
    student2_id = "user-student2"
    student3_id = "user-student3"
    
    users = [
        {
            "id": editor_id,
            "name": "Editor General",
            "email": "editorgeneral@educando.com",
            "cedula": None,
            "password_hash": hash_password("EditorSeguro2025"),
            "role": "editor",
            "program_id": None,
            "phone": "3002222222",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": admin_id,
            "name": "Administrador General",
            "email": "admin@educando.com",
            "cedula": None,
            "password_hash": hash_password("Admin2026*Seed"),
            "role": "admin",
            "program_id": None,
            "phone": "3001234567",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": teacher1_id,
            "name": "María García López",
            "email": "profesor@educando.com",
            "cedula": None,
            "password_hash": hash_password("Profe2026*Seed1"),
            "role": "profesor",
            "program_id": None,
            "phone": "3007654321",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": teacher2_id,
            "name": "Carlos Rodríguez Pérez",
            "email": "profesor2@educando.com",
            "cedula": None,
            "password_hash": hash_password("Profe2026*Seed2"),
            "role": "profesor",
            "program_id": None,
            "phone": "3009876543",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": student1_id,
            "name": "Juan Martínez Ruiz",
            "email": None,
            "cedula": "1234567890",
            "password_hash": hash_password("Estud2026*Seed1"),
            "role": "estudiante",
            "program_id": "prog-admin",
            "phone": "3101234567",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": student2_id,
            "name": "Ana Sofía Hernández",
            "email": None,
            "cedula": "0987654321",
            "password_hash": hash_password("Estud2026*Seed2"),
            "role": "estudiante",
            "program_id": "prog-infancia",
            "phone": "3207654321",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": student3_id,
            "name": "Pedro López Castro",
            "email": None,
            "cedula": "1122334455",
            "password_hash": hash_password("Estud2026*Seed3"),
            "role": "estudiante",
            "program_id": "prog-sst",
            "phone": "3159876543",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    await db.users.insert_many(users)
    
    return {
        "message": "Datos iniciales creados exitosamente",
        "users_created": 7,
        "programs_created": 3
    }

# --- Stats Route ---
@api_router.delete("/admin/purge-group-data")
async def purge_all_group_data(user=Depends(get_current_user)):
    """Delete ALL courses/groups and every dependent record (activities, grades,
    submissions, videos, recovery_enabled, failed_subjects). Admin only."""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")

    courses_deleted = await db.courses.delete_many({})
    activities_deleted = await db.activities.delete_many({})
    grades_deleted = await db.grades.delete_many({})
    submissions_deleted = await db.submissions.delete_many({})
    videos_deleted = await db.class_videos.delete_many({})
    recovery_deleted = await db.recovery_enabled.delete_many({})
    failed_deleted = await db.failed_subjects.delete_many({})

    await log_audit(
        "purge_group_data",
        user["id"],
        user["role"],
        {
            "courses_deleted": courses_deleted.deleted_count,
            "activities_deleted": activities_deleted.deleted_count,
            "grades_deleted": grades_deleted.deleted_count,
            "submissions_deleted": submissions_deleted.deleted_count,
            "videos_deleted": videos_deleted.deleted_count,
            "recovery_deleted": recovery_deleted.deleted_count,
            "failed_subjects_deleted": failed_deleted.deleted_count,
        }
    )

    logger.info(
        f"Admin {user['id']} purged all group data: "
        f"{courses_deleted.deleted_count} courses, "
        f"{activities_deleted.deleted_count} activities, "
        f"{grades_deleted.deleted_count} grades, "
        f"{submissions_deleted.deleted_count} submissions, "
        f"{videos_deleted.deleted_count} videos"
    )

    return {
        "message": "Todos los grupos y sus datos dependientes fueron eliminados",
        "courses_deleted": courses_deleted.deleted_count,
        "activities_deleted": activities_deleted.deleted_count,
        "grades_deleted": grades_deleted.deleted_count,
        "submissions_deleted": submissions_deleted.deleted_count,
        "videos_deleted": videos_deleted.deleted_count,
        "recovery_enabled_deleted": recovery_deleted.deleted_count,
        "failed_subjects_deleted": failed_deleted.deleted_count,
    }

@api_router.get("/stats")
async def get_stats(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    
    students = await db.users.count_documents({"role": "estudiante", "active": True})
    teachers = await db.users.count_documents({"role": "profesor", "active": True})
    programs = await db.programs.count_documents({"active": True})
    courses = await db.courses.count_documents({"active": True})
    # Only count activities that belong to an existing course
    existing_course_ids = [c["id"] for c in await db.courses.find({}, {"_id": 0, "id": 1}).to_list(5000)]
    if existing_course_ids:
        activities = await db.activities.count_documents({"course_id": {"$in": existing_course_ids}})
    else:
        activities = 0
    
    return {
        "students": students,
        "teachers": teachers,
        "programs": programs,
        "courses": courses,
        "activities": activities
    }

@api_router.get("/health")
async def health_check():
    """Health check endpoint for monitoring and Railway deployment"""
    try:
        # Test MongoDB connection
        await db.command('ping')
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "database": "disconnected",
                "error": str(e),
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
        )

@api_router.get("/")
async def root():
    return {"message": "Corporación Social Educando API"}

@api_router.get("/admin/audit-logs")
async def get_audit_logs(
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user=Depends(get_current_user)
):
    """
    Returns paginated audit logs. Admin only.
    Supports filters: action, user_id, from_date (ISO), to_date (ISO).
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede acceder a los registros de auditoría")
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 100:
        page_size = 20

    query: dict = {}
    if action:
        query["action"] = action
    if user_id:
        query["user_id"] = user_id
    if from_date or to_date:
        ts_filter: dict = {}
        if from_date:
            # Normalize: treat date-only strings as start of day UTC
            ts_filter["$gte"] = from_date if "T" in from_date else from_date + "T00:00:00+00:00"
        if to_date:
            # Normalize: treat date-only strings as end of day UTC
            ts_filter["$lte"] = to_date if "T" in to_date else to_date + "T23:59:59+00:00"
        query["timestamp"] = ts_filter

    total = await db.audit_logs.count_documents(query)
    skip = (page - 1) * page_size
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(page_size).to_list(page_size)

    # Enrich each log with the actor's name where possible
    actor_ids = list({log["user_id"] for log in logs if log.get("user_id") and log["user_id"] != "system"})
    actor_map: dict = {}
    if actor_ids:
        actors = await db.users.find({"id": {"$in": actor_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        actor_map = {a["id"]: a["name"] for a in actors}

    for log in logs:
        uid = log.get("user_id", "")
        if uid == "system":
            log["user_name"] = "Sistema"
        else:
            log["user_name"] = actor_map.get(uid, uid or "-")

    return {
        "logs": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

app.include_router(api_router)

@app.get("/")
async def app_root():
    return {"message": "Corporación Social Educando API"}

@app.on_event("shutdown")
async def shutdown_db_client():
    # Shutdown scheduler gracefully
    if scheduler.running:
        scheduler.shutdown()
        logger.info("Scheduler shut down")
    client.close()
