import asyncio
import logging
import io
import csv
import re
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import JSONResponse, StreamingResponse

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database import db
from utils.security import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/student/dashboard/{course_id}")
async def get_student_dashboard(
    course_id: str,
    subject_id: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Consolidated endpoint that returns all data needed for the student course dashboard
    in a single request, reducing 4-5 API calls to 1."""
    if user["role"] != "estudiante":
        raise HTTPException(status_code=403, detail="Solo estudiantes")

    student_id = user["id"]

    # Build all queries
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course or student_id not in (course.get("student_ids") or []):
        raise HTTPException(status_code=404, detail="Curso no encontrado o no inscrito")

    # Activities query
    activities_query = {"course_id": course_id}
    if subject_id:
        activities_query["subject_id"] = subject_id

    # Videos query (with available_from filter for students)
    videos_query = {"course_id": course_id}
    if subject_id:
        videos_query["subject_id"] = subject_id
    now_iso = datetime.now(timezone.utc).isoformat()
    videos_query["$or"] = [
        {"available_from": {"$exists": False}},
        {"available_from": None},
        {"available_from": ""},
        {"available_from": {"$lte": now_iso}}
    ]

    # Grades query
    grades_query = {"student_id": student_id, "course_id": course_id}
    if subject_id:
        grades_query["subject_id"] = subject_id

    # Execute all queries in parallel
    activities, videos, grades = await asyncio.gather(
        db.activities.find(activities_query, {"_id": 0}).to_list(500),
        db.class_videos.find(videos_query, {"_id": 0}).to_list(500),
        db.grades.find(grades_query, {"_id": 0}).to_list(10000)
    )

    # Filter recovery activities for students
    if activities:
        approved_records = await db.failed_subjects.find({
            "student_id": student_id,
            "course_id": course_id,
            "recovery_approved": True,
            "recovery_processed": {"$ne": True},
            "recovery_expired": {"$ne": True}
        }, {"_id": 0, "subject_id": 1}).to_list(100)
        approved_subject_ids = {r.get("subject_id") for r in approved_records}
        activities = [
            a for a in activities
            if not a.get("is_recovery") or a.get("subject_id") in approved_subject_ids
        ]

    # Optionally include subject info
    subject = None
    if subject_id:
        subject = await db.subjects.find_one({"id": subject_id}, {"_id": 0})

    return {
        "course": course,
        "activities": activities,
        "videos": videos,
        "grades": grades,
        "subject": subject
    }


@router.get("/teacher/dashboard/{course_id}")
async def get_teacher_dashboard(
    course_id: str,
    subject_id: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Consolidated endpoint that returns all data needed for the teacher course dashboard
    in a single request, reducing 3-4 API calls to 1."""
    if user["role"] not in ["profesor", "admin"]:
        raise HTTPException(status_code=403, detail="Solo profesores o admin")

    # Build all queries
    course_fut = db.courses.find_one({"id": course_id}, {"_id": 0})

    activities_query = {"course_id": course_id}
    if subject_id:
        activities_query["subject_id"] = subject_id

    videos_query = {"course_id": course_id}
    if subject_id:
        videos_query["subject_id"] = subject_id

    # Execute all in parallel
    course, activities, videos = await asyncio.gather(
        course_fut,
        db.activities.find(activities_query, {"_id": 0}).to_list(500),
        db.class_videos.find(videos_query, {"_id": 0}).sort("created_at", -1).to_list(500),
    )

    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")

    # Optionally include subject info (only fetches 1 doc, not all subjects)
    subject = None
    if subject_id:
        subject = await db.subjects.find_one({"id": subject_id}, {"_id": 0})

    return {
        "course": course,
        "activities": activities,
        "videos": videos,
        "subject": subject
    }


@router.get("/teacher/grades-data/{course_id}")
async def get_teacher_grades_data(
    course_id: str,
    subject_id: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Consolidated endpoint that returns all data needed for the teacher grades page
    in a single request, reducing 5 API calls to 1."""
    if user["role"] not in ["profesor", "admin"]:
        raise HTTPException(status_code=403, detail="Solo profesores o admin")

    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")

    # Build queries
    activities_query = {"course_id": course_id}
    if subject_id:
        activities_query["subject_id"] = subject_id

    grades_query = {"course_id": course_id}
    if subject_id:
        grades_query["subject_id"] = subject_id

    # All student IDs (enrolled + removed)
    student_ids = list(set(
        (course.get("student_ids") or []) +
        (course.get("removed_student_ids") or [])
    ))

    # Recovery enabled query
    fs_query = {
        "recovery_approved": True,
        "recovery_completed": {"$ne": True},
        "recovery_processed": {"$ne": True},
        "recovery_rejected": {"$ne": True},
        "course_id": course_id,
    }

    # Build futures before gather
    activities_fut = db.activities.find(activities_query, {"_id": 0}).to_list(500)
    grades_fut = db.grades.find(grades_query, {"_id": 0}).to_list(10000)

    async def _empty_list():
        return []

    students_fut = (
        db.users.find(
            {"id": {"$in": student_ids}, "role": "estudiante"},
            {"_id": 0, "password_hash": 0}
        ).to_list(5000)
        if student_ids else _empty_list()
    )
    recovery_fut = db.failed_subjects.find(
        fs_query, {"_id": 0, "student_id": 1, "course_id": 1, "subject_id": 1}
    ).to_list(1000)

    results = await asyncio.gather(activities_fut, grades_fut, students_fut, recovery_fut)
    activities = results[0]
    grades = results[1]
    students_docs = results[2] if student_ids else []
    recovery_records = results[3]

    # Mark removed students
    removed_set = set(course.get("removed_student_ids") or []) - set(course.get("student_ids") or [])
    for s in students_docs:
        if s["id"] in removed_set:
            s["_removed_from_group"] = True

    # Deduplicate recovery enabled
    seen = set()
    recovery_enabled = []
    for r in recovery_records:
        key = (r.get("student_id"), r.get("course_id"), r.get("subject_id"))
        if key not in seen:
            seen.add(key)
            recovery_enabled.append({
                "student_id": r.get("student_id"),
                "course_id": r.get("course_id"),
                "subject_id": r.get("subject_id"),
                "enabled": True,
                "source": "failed_subjects"
            })

    return {
        "course": course,
        "activities": activities,
        "grades": grades,
        "students": students_docs,
        "recovery_enabled": recovery_enabled
    }


@router.get("/reports/course-results")
async def get_course_results_report(course_id: str, subject_id: Optional[str] = None, format: Optional[str] = None, user=Depends(get_current_user)):
    """
    Returns a report of approved/reproved students per course/group.
    Accessible by admin and professor.
    When subject_id is provided, filters the report to only that subject.
    When format=csv/xlsx, returns a file download; otherwise returns JSON.
    """
    if user["role"] not in ["admin", "profesor"]:
        raise HTTPException(status_code=403, detail="Solo admin o profesor pueden acceder a reportes")
    
    course = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    # Load subjects
    all_subject_ids = course.get("subject_ids") or []
    if not all_subject_ids and course.get("subject_id"):
        all_subject_ids = [course["subject_id"]]
    # When subject_id filter is provided, restrict to that single subject
    subject_ids = [subject_id] if subject_id and subject_id in all_subject_ids else all_subject_ids
    all_subjects = await db.subjects.find({"id": {"$in": subject_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500) if subject_ids else []
    subject_map = {s["id"]: s["name"] for s in all_subjects}
    
    # Load enrolled students
    student_ids = course.get("student_ids") or []
    students = await db.users.find(
        {"id": {"$in": student_ids}, "role": "estudiante"},
        {"_id": 0, "id": 1, "name": 1, "cedula": 1}
    ).to_list(5000) if student_ids else []
    
    # Load grades for this course and build an index: (student_id, subject_id) -> [values]
    grades = await db.grades.find({"course_id": course_id}, {"_id": 0, "student_id": 1, "subject_id": 1, "value": 1}).to_list(200000)
    grades_index = {}
    for g in grades:
        if g.get("value") is not None:
            key = (g["student_id"], g.get("subject_id"))
            grades_index.setdefault(key, []).append(g["value"])
    
    # Determine module number for the group (from first subject, or from course module_dates)
    # A course group targets a single module; all subjects in the course should share the same
    # module_number in practice. We use the first available value as a best-effort.
    module_number = None
    if all_subjects:
        module_number = all_subjects[0].get("module_number")
    if module_number is None:
        module_dates = course.get("module_dates") or {}
        if module_dates:
            module_number = next(iter(module_dates.keys()), None)

    # Pre-load all active recovery records for this course in one query (avoids N+1)
    _recovery_docs = await db.failed_subjects.find(
        {"course_id": course_id, "recovery_processed": {"$ne": True}},
        {"_id": 0, "student_id": 1}
    ).to_list(10000)
    # Use a set: we only need to know whether a student has any active recovery record
    recovery_student_ids = {r["student_id"] for r in _recovery_docs}

    # Build per-student summary rows (one row per student, columns for each subject average)
    rows = []
    for student in students:
        sid = student["id"]
        subject_avgs = {}
        for subj_id in subject_ids:
            values = grades_index.get((sid, subj_id), [])
            subject_avgs[subj_id] = round(sum(values) / len(values), 2) if values else 0.0
        # General average: use only the filtered subjects
        all_values = [v for subj_id in subject_ids for v in grades_index.get((sid, subj_id), [])]
        if not all_values:
            # Fall back: no subject_ids configured – use all grades for this course
            all_values = [v for (s, _subj), vals in grades_index.items() if s == sid for v in vals]
        general_avg = round(sum(all_values) / len(all_values), 2) if all_values else 0.0
        # Check recovery status using pre-loaded set (no per-student DB query)
        recovery_record = sid in recovery_student_ids
        if general_avg >= 3.0:
            status = "Aprobado"
        elif recovery_record:
            status = "En Recuperación"
        else:
            status = "Reprobado"
        row = {
            "student_name": student["name"],
            "student_cedula": student.get("cedula", ""),
            "module": module_number,
            "course_name": course.get("name", ""),
            "general_average": general_avg,
            "status": status,
            "student_id": sid,
            "course_id": course_id,
        }
        for subj_id in subject_ids:
            row[f"subject_{subj_id}"] = subject_avgs[subj_id]
            row[f"subject_name_{subj_id}"] = subject_map.get(subj_id, subj_id)
        rows.append(row)
    
    if format and format.lower() in ("csv", "xlsx"):
        course_name = course.get("name", course_id)
        safe_name = re.sub(r'[^\w\-]', '_', course_name)
        subject_headers = [subject_map.get(sid, sid) for sid in subject_ids]
        base_headers = ["Nombre", "Cédula", "Módulo"]
        end_headers = ["Promedio", "Estado"]
        all_headers = base_headers + subject_headers + end_headers

        if format.lower() == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(all_headers)
            for row in rows:
                csv_row = [row["student_name"], row["student_cedula"], row.get("module", "")]
                for subj_id in subject_ids:
                    csv_row.append(row.get(f"subject_{subj_id}", 0.0))
                csv_row.extend([row["general_average"], row["status"]])
                writer.writerow(csv_row)
            output.seek(0)
            filename = f"resultados_{safe_name}.csv"
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            # XLSX with professional formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados"

            # Title row
            title_label = course_name
            if subject_id and subject_id in subject_map:
                title_label = f"{course_name} – {subject_map[subject_id]}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
            title_cell = ws.cell(row=1, column=1, value=f"Reporte de Resultados – {title_label}")
            title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_cell.font = Font(bold=True, size=13, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Header row (row 2)
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(border_style="thin", color="000000")
            header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, header in enumerate(all_headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[2].height = 18

            # Data rows (starting row 3)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row_idx, row in enumerate(rows, start=3):
                avg = row["general_average"]
                row_fill = green_fill if avg >= 3.0 else red_fill
                data = [row["student_name"], row["student_cedula"], row.get("module", "")]
                for subj_id in subject_ids:
                    data.append(row.get(f"subject_{subj_id}", 0.0))
                data.extend([avg, row["status"]])
                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = row_fill
                    cell.border = data_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-adjust column widths
            for col_idx, header in enumerate(all_headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(header)), 10)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
            # Fixed widths for known columns
            ws.column_dimensions["A"].width = 30  # Nombre
            ws.column_dimensions["B"].width = 14  # Cédula

            output_bytes = io.BytesIO()
            wb.save(output_bytes)
            output_bytes.seek(0)
            if subject_id and subject_id in subject_map:
                safe_subject = re.sub(r'[^\w\-]', '_', subject_map[subject_id])
                filename = f"resultados_{safe_name}_{safe_subject}.xlsx"
            else:
                filename = f"resultados_{safe_name}.xlsx"
            return StreamingResponse(
                iter([output_bytes.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    
    return {
        "course_id": course_id,
        "course_name": course.get("name", ""),
        "module": module_number,
        "subjects": [{"id": sid, "name": subject_map.get(sid, sid)} for sid in subject_ids],
        "rows": rows,
        "total": len(rows)
    }


@router.get("/reports/recovery-results")
async def get_recovery_results_report(format: Optional[str] = None, user=Depends(get_current_user)):
    """
    Returns a consolidated report of all students in recovery with their status.
    Accessible by admin only.
    When format=xlsx, returns an XLSX file download; otherwise returns JSON.
    """
    if user["role"] not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Solo admin puede acceder a reportes de recuperación")

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Fetch all unprocessed failed subjects
    all_failed = await db.failed_subjects.find(
        {"recovery_processed": {"$ne": True}},
        {"_id": 0}
    ).to_list(50000)

    if not all_failed:
        all_failed = []

    # Build lookup maps
    student_ids_set = list({r["student_id"] for r in all_failed})
    course_ids_set = list({r["course_id"] for r in all_failed})
    program_ids_set = list({r.get("program_id") for r in all_failed if r.get("program_id")})
    subject_ids_set = list({r.get("subject_id") for r in all_failed if r.get("subject_id")})

    students_list = await db.users.find({"id": {"$in": student_ids_set}}, {"_id": 0, "id": 1, "name": 1, "cedula": 1}).to_list(5000) if student_ids_set else []
    courses_list = await db.courses.find({"id": {"$in": course_ids_set}}, {"_id": 0, "id": 1, "name": 1, "module_dates": 1}).to_list(1000) if course_ids_set else []
    programs_list = await db.programs.find({"id": {"$in": program_ids_set}}, {"_id": 0, "id": 1, "name": 1}).to_list(100) if program_ids_set else []
    subjects_list = await db.subjects.find({"id": {"$in": subject_ids_set}}, {"_id": 0, "id": 1, "name": 1}).to_list(500) if subject_ids_set else []

    student_map = {s["id"]: s for s in students_list}
    course_map = {c["id"]: c for c in courses_list}
    program_map = {p["id"]: p["name"] for p in programs_list}
    subject_map = {s["id"]: s["name"] for s in subjects_list}

    rows = []
    for record in all_failed:
        student = student_map.get(record["student_id"], {})
        course = course_map.get(record["course_id"], {})
        # Determine recovery close date
        module_key = str(record.get("module_number", ""))
        module_dates = (course.get("module_dates") or {}).get(module_key) or {}
        recovery_close = module_dates.get("recovery_close")
        is_expired = bool(recovery_close and recovery_close <= today_str)
        # Determine status label
        if is_expired and not record.get("recovery_approved"):
            status_label = "Plazo vencido"
        elif record.get("teacher_graded_status") == "approved":
            status_label = "Calificada por profesor: Aprobado"
        elif record.get("teacher_graded_status") == "rejected":
            status_label = "Calificada por profesor: Reprobado"
        elif record.get("recovery_approved"):
            status_label = "Aprobada por admin"
        else:
            status_label = "Pendiente"
        subject_name = record.get("subject_name") or subject_map.get(record.get("subject_id"), "")
        rows.append({
            "student_name": student.get("name", record.get("student_name", "")),
            "student_cedula": student.get("cedula", ""),
            "subject_name": subject_name,
            "course_name": record.get("course_name", course.get("name", "")),
            "program_name": record.get("program_name", program_map.get(record.get("program_id"), "")),
            "module_number": record.get("module_number", ""),
            "average_grade": record.get("average_grade", 0.0),
            "status": status_label,
        })

    if format and format.lower() == "xlsx":
        all_headers = ["Nombre", "Cédula", "Materia reprobada", "Grupo/Curso", "Programa", "Módulo", "Promedio anterior", "Estado de recuperación"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recuperaciones"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
        title_cell = ws.cell(row=1, column=1, value="Reporte de Recuperaciones")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="000000")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 18

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row_idx, row in enumerate(rows, start=3):
            status = row["status"]
            if "Aprobado" in status:
                row_fill = green_fill
            elif "Reprobado" in status or "vencido" in status:
                row_fill = red_fill
            else:
                row_fill = yellow_fill
            data = [
                row["student_name"], row["student_cedula"], row["subject_name"],
                row["course_name"], row["program_name"], row["module_number"],
                row["average_grade"], row["status"]
            ]
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = data_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 32

        output_bytes = io.BytesIO()
        wb.save(output_bytes)
        output_bytes.seek(0)
        return StreamingResponse(
            iter([output_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=reporte_recuperaciones.xlsx"}
        )

    return {"rows": rows, "total": len(rows)}

@router.get("/stats")
async def get_stats(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")

    # Conteo de estudiantes por estado en una sola aggregation query
    student_status_agg = await db.users.aggregate([
        {"$match": {"role": "estudiante"}},
        {"$group": {
            "_id": {"$ifNull": ["$estado", "activo"]},
            "count": {"$sum": 1}
        }}
    ]).to_list(20)

    students_by_status = {doc["_id"]: doc["count"] for doc in student_status_agg}
    total_students = sum(students_by_status.values())

    teachers, programs_count, courses_count, existing_courses = await asyncio.gather(
        db.users.count_documents({"role": "profesor"}),
        db.programs.count_documents({}),
        db.courses.count_documents({}),
        db.courses.find({}, {"_id": 0, "id": 1}).to_list(5000),
    )
    existing_course_ids = [c["id"] for c in existing_courses]
    activities = await db.activities.count_documents({"course_id": {"$in": existing_course_ids}}) if existing_course_ids else 0

    return {
        "students": total_students,
        "teachers": teachers,
        "programs": programs_count,
        "courses": courses_count,
        "activities": activities,
        "students_activo": students_by_status.get("activo", 0),
        "students_pendiente_recuperacion": students_by_status.get("pendiente_recuperacion", 0),
        "students_egresado": students_by_status.get("egresado", 0),
        "students_retirado": students_by_status.get("retirado", 0),
        "students_reprobado": students_by_status.get("reprobado", 0),
    }

@router.get("/health")
async def health_check():
    """Health check endpoint for monitoring and Railway deployment"""
    try:
        # Test MongoDB connection
        await db.command('ping')
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "database": "disconnected",
                "error": str(e),
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
        )

@router.get("/")
async def root():
    return {"message": "Corporación Social Educando API"}

@router.get("/admin/audit-logs")
async def get_audit_logs(
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user=Depends(get_current_user)
):
    """
    Returns paginated audit logs. Admin only.
    Supports filters: action, user_id, from_date (ISO), to_date (ISO).
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede acceder a los registros de auditoría")
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 100:
        page_size = 20

    query: dict = {}
    if action:
        query["action"] = action
    if user_id:
        query["user_id"] = user_id
    if from_date or to_date:
        ts_filter: dict = {}
        if from_date:
            # Normalize: treat date-only strings as start of day UTC
            ts_filter["$gte"] = from_date if "T" in from_date else from_date + "T00:00:00+00:00"
        if to_date:
            # Normalize: treat date-only strings as end of day UTC
            ts_filter["$lte"] = to_date if "T" in to_date else to_date + "T23:59:59+00:00"
        query["timestamp"] = ts_filter

    total = await db.audit_logs.count_documents(query)
    skip = (page - 1) * page_size
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(page_size).to_list(page_size)

    # Enrich each log with the actor's name where possible
    actor_ids = list({log["user_id"] for log in logs if log.get("user_id") and log["user_id"] != "system"})
    actor_map: dict = {}
    if actor_ids:
        actors = await db.users.find({"id": {"$in": actor_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
        actor_map = {a["id"]: a["name"] for a in actors}

    for log in logs:
        uid = log.get("user_id", "")
        if uid == "system":
            log["user_name"] = "Sistema"
        else:
            log["user_name"] = actor_map.get(uid, uid or "-")

    return {
        "logs": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }